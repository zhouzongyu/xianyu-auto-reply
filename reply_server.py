from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from typing import List, Tuple, Optional, Dict, Any
from pathlib import Path
from urllib.parse import unquote
import hashlib
import secrets
import time
import json
import os
import re
import uvicorn
import pandas as pd
import io
import asyncio
from collections import defaultdict

import cookie_manager
from db_manager import db_manager
from file_log_collector import setup_file_logging, get_file_log_collector
from ai_reply_engine import ai_reply_engine
from utils.qr_login import qr_login_manager
from utils.xianyu_utils import trans_cookies
from utils.image_utils import image_manager

from loguru import logger

# åˆ®åˆ®ä¹è¿œç¨‹æ§åˆ¶è·¯ç”±
try:
    from api_captcha_remote import router as captcha_router
    CAPTCHA_ROUTER_AVAILABLE = True
except ImportError:
    logger.warning("âš ï¸ api_captcha_remote æœªæ‰¾åˆ°ï¼Œåˆ®åˆ®ä¹è¿œç¨‹æ§åˆ¶åŠŸèƒ½ä¸å¯ç”¨")
    CAPTCHA_ROUTER_AVAILABLE = False

# å…³é”®å­—æ–‡ä»¶è·¯å¾„
KEYWORDS_FILE = Path(__file__).parent / "å›å¤å…³é”®å­—.txt"

# ç®€å•çš„ç”¨æˆ·è®¤è¯é…ç½®
ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin123"  # ç³»ç»Ÿåˆå§‹åŒ–æ—¶çš„é»˜è®¤å¯†ç 
SESSION_TOKENS = {}  # å­˜å‚¨ä¼šè¯token: {token: {'user_id': int, 'username': str, 'timestamp': float}}
TOKEN_EXPIRE_TIME = 24 * 60 * 60  # tokenè¿‡æœŸæ—¶é—´ï¼š24å°æ—¶

# HTTP Bearerè®¤è¯
security = HTTPBearer(auto_error=False)

# æ‰«ç ç™»å½•æ£€æŸ¥é” - é˜²æ­¢å¹¶å‘å¤„ç†åŒä¸€ä¸ªsession
qr_check_locks = defaultdict(lambda: asyncio.Lock())
qr_check_processed = {}  # è®°å½•å·²å¤„ç†çš„session: {session_id: {'processed': bool, 'timestamp': float}}

# è´¦å·å¯†ç ç™»å½•ä¼šè¯ç®¡ç†
password_login_sessions = {}  # {session_id: {'account_id': str, 'account': str, 'password': str, 'show_browser': bool, 'status': str, 'verification_url': str, 'qr_code_url': str, 'slider_instance': object, 'task': asyncio.Task, 'timestamp': float}}
password_login_locks = defaultdict(lambda: asyncio.Lock())

# ä¸å†éœ€è¦å•ç‹¬çš„å¯†ç åˆå§‹åŒ–ï¼Œç”±æ•°æ®åº“åˆå§‹åŒ–æ—¶å¤„ç†


def cleanup_qr_check_records():
    """æ¸…ç†è¿‡æœŸçš„æ‰«ç æ£€æŸ¥è®°å½•"""
    current_time = time.time()
    expired_sessions = []

    for session_id, record in qr_check_processed.items():
        # æ¸…ç†è¶…è¿‡1å°æ—¶çš„è®°å½•
        if current_time - record['timestamp'] > 3600:
            expired_sessions.append(session_id)

    for session_id in expired_sessions:
        if session_id in qr_check_processed:
            del qr_check_processed[session_id]
        if session_id in qr_check_locks:
            del qr_check_locks[session_id]


def load_keywords() -> List[Tuple[str, str]]:
    """è¯»å–å…³é”®å­—â†’å›å¤æ˜ å°„è¡¨

    æ–‡ä»¶æ ¼å¼æ”¯æŒï¼š
        å…³é”®å­—<ç©ºæ ¼/åˆ¶è¡¨ç¬¦/å†’å·>å›å¤å†…å®¹
    å¿½ç•¥ç©ºè¡Œå’Œä»¥ # å¼€å¤´çš„æ³¨é‡Šè¡Œ
    """
    mapping: List[Tuple[str, str]] = []
    if not KEYWORDS_FILE.exists():
        return mapping

    with KEYWORDS_FILE.open('r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            # å°è¯•ç”¨\tã€ç©ºæ ¼ã€å†’å·åˆ†éš”
            if '\t' in line:
                key, reply = line.split('\t', 1)
            elif ' ' in line:
                key, reply = line.split(' ', 1)
            elif ':' in line:
                key, reply = line.split(':', 1)
            else:
                # æ— æ³•è§£æçš„è¡Œï¼Œè·³è¿‡
                continue
            mapping.append((key.strip(), reply.strip()))
    return mapping


KEYWORDS_MAPPING = load_keywords()


# è®¤è¯ç›¸å…³æ¨¡å‹
class LoginRequest(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    email: Optional[str] = None
    verification_code: Optional[str] = None


class LoginResponse(BaseModel):
    success: bool
    token: Optional[str] = None
    message: str
    user_id: Optional[int] = None
    username: Optional[str] = None
    is_admin: Optional[bool] = None


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class RegisterRequest(BaseModel):
    username: str
    email: str
    password: str
    verification_code: str


class RegisterResponse(BaseModel):
    success: bool
    message: str


class SendCodeRequest(BaseModel):
    email: str
    session_id: Optional[str] = None
    type: Optional[str] = 'register'  # 'register' æˆ– 'login'


class SendCodeResponse(BaseModel):
    success: bool
    message: str


class CaptchaRequest(BaseModel):
    session_id: str


class CaptchaResponse(BaseModel):
    success: bool
    captcha_image: str
    session_id: str
    message: str


class VerifyCaptchaRequest(BaseModel):
    session_id: str
    captcha_code: str


class VerifyCaptchaResponse(BaseModel):
    success: bool
    message: str


def generate_token() -> str:
    """ç”Ÿæˆéšæœºtoken"""
    return secrets.token_urlsafe(32)


def verify_token(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Optional[Dict[str, Any]]:
    """éªŒè¯tokenå¹¶è¿”å›ç”¨æˆ·ä¿¡æ¯"""
    if not credentials:
        return None

    token = credentials.credentials
    if token not in SESSION_TOKENS:
        return None

    token_data = SESSION_TOKENS[token]

    # æ£€æŸ¥tokenæ˜¯å¦è¿‡æœŸ
    if time.time() - token_data['timestamp'] > TOKEN_EXPIRE_TIME:
        del SESSION_TOKENS[token]
        return None

    return token_data


def verify_admin_token(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Dict[str, Any]:
    """éªŒè¯ç®¡ç†å‘˜token"""
    user_info = verify_token(credentials)
    if not user_info:
        raise HTTPException(status_code=401, detail="æœªæˆæƒè®¿é—®")

    # æ£€æŸ¥æ˜¯å¦æ˜¯ç®¡ç†å‘˜
    if user_info['username'] != ADMIN_USERNAME:
        raise HTTPException(status_code=403, detail="éœ€è¦ç®¡ç†å‘˜æƒé™")

    return user_info


def require_auth(user_info: Optional[Dict[str, Any]] = Depends(verify_token)):
    """éœ€è¦è®¤è¯çš„ä¾èµ–ï¼Œè¿”å›ç”¨æˆ·ä¿¡æ¯"""
    if not user_info:
        raise HTTPException(status_code=401, detail="æœªæˆæƒè®¿é—®")
    return user_info


def get_current_user(user_info: Dict[str, Any] = Depends(require_auth)) -> Dict[str, Any]:
    """è·å–å½“å‰ç™»å½•ç”¨æˆ·ä¿¡æ¯"""
    return user_info


def get_current_user_optional(user_info: Optional[Dict[str, Any]] = Depends(verify_token)) -> Optional[Dict[str, Any]]:
    """è·å–å½“å‰ç”¨æˆ·ä¿¡æ¯ï¼ˆå¯é€‰ï¼Œä¸å¼ºåˆ¶è¦æ±‚ç™»å½•ï¼‰"""
    return user_info


def get_user_log_prefix(user_info: Dict[str, Any] = None) -> str:
    """è·å–ç”¨æˆ·æ—¥å¿—å‰ç¼€"""
    if user_info:
        return f"ã€{user_info['username']}#{user_info['user_id']}ã€‘"
    return "ã€ç³»ç»Ÿã€‘"


def require_admin(current_user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    """è¦æ±‚ç®¡ç†å‘˜æƒé™"""
    if current_user['username'] != 'admin':
        raise HTTPException(status_code=403, detail="éœ€è¦ç®¡ç†å‘˜æƒé™")
    return current_user


def log_with_user(level: str, message: str, user_info: Dict[str, Any] = None):
    """å¸¦ç”¨æˆ·ä¿¡æ¯çš„æ—¥å¿—è®°å½•"""
    prefix = get_user_log_prefix(user_info)
    full_message = f"{prefix} {message}"

    if level.lower() == 'info':
        logger.info(full_message)
    elif level.lower() == 'error':
        logger.error(full_message)
    elif level.lower() == 'warning':
        logger.warning(full_message)
    elif level.lower() == 'debug':
        logger.debug(full_message)
    else:
        logger.info(full_message)


def match_reply(cookie_id: str, message: str) -> Optional[str]:
    """æ ¹æ® cookie_id åŠæ¶ˆæ¯å†…å®¹åŒ¹é…å›å¤
    åªæœ‰å¯ç”¨çš„è´¦å·æ‰ä¼šåŒ¹é…å…³é”®å­—å›å¤
    """
    mgr = cookie_manager.manager
    if mgr is None:
        return None

    # æ£€æŸ¥è´¦å·æ˜¯å¦å¯ç”¨
    if not mgr.get_cookie_status(cookie_id):
        return None  # ç¦ç”¨çš„è´¦å·ä¸å‚ä¸è‡ªåŠ¨å›å¤

    # ä¼˜å…ˆè´¦å·çº§å…³é”®å­—
    if mgr.get_keywords(cookie_id):
        for k, r in mgr.get_keywords(cookie_id):
            if k in message:
                return r

    # å…¨å±€å…³é”®å­—
    for k, r in KEYWORDS_MAPPING:
        if k in message:
            return r
    return None


class RequestModel(BaseModel):
    cookie_id: str
    msg_time: str
    user_url: str
    send_user_id: str
    send_user_name: str
    item_id: str
    send_message: str
    chat_id: str


class ResponseData(BaseModel):
    send_msg: str


class ResponseModel(BaseModel):
    code: int
    data: ResponseData


app = FastAPI(
    title="Xianyu Auto Reply API",
    version="1.0.0",
    description="é—²é±¼è‡ªåŠ¨å›å¤ç³»ç»ŸAPI",
    docs_url="/docs",
    redoc_url="/redoc"
)

# æ³¨å†Œåˆ®åˆ®ä¹è¿œç¨‹æ§åˆ¶è·¯ç”±
if CAPTCHA_ROUTER_AVAILABLE:
    app.include_router(captcha_router)
    logger.info("âœ… å·²æ³¨å†Œåˆ®åˆ®ä¹è¿œç¨‹æ§åˆ¶è·¯ç”±: /api/captcha")
else:
    logger.warning("âš ï¸ åˆ®åˆ®ä¹è¿œç¨‹æ§åˆ¶è·¯ç”±æœªæ³¨å†Œ")

# åˆå§‹åŒ–æ–‡ä»¶æ—¥å¿—æ”¶é›†å™¨
setup_file_logging()

# æ·»åŠ ä¸€æ¡æµ‹è¯•æ—¥å¿—
from loguru import logger
logger.info("WebæœåŠ¡å™¨å¯åŠ¨ï¼Œæ–‡ä»¶æ—¥å¿—æ”¶é›†å™¨å·²åˆå§‹åŒ–")

# æ·»åŠ è¯·æ±‚æ—¥å¿—ä¸­é—´ä»¶
@app.middleware("http")
async def log_requests(request, call_next):
    start_time = time.time()

    # è·å–ç”¨æˆ·ä¿¡æ¯
    user_info = "æœªç™»å½•"
    try:
        # ä»è¯·æ±‚å¤´ä¸­è·å–Authorization
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]
            if token in SESSION_TOKENS:
                token_data = SESSION_TOKENS[token]
                # æ£€æŸ¥tokenæ˜¯å¦è¿‡æœŸ
                if time.time() - token_data['timestamp'] <= TOKEN_EXPIRE_TIME:
                    user_info = f"ã€{token_data['username']}#{token_data['user_id']}ã€‘"
    except Exception:
        pass

    logger.info(f"ğŸŒ {user_info} APIè¯·æ±‚: {request.method} {request.url.path}")

    response = await call_next(request)

    process_time = time.time() - start_time
    logger.info(f"âœ… {user_info} APIå“åº”: {request.method} {request.url.path} - {response.status_code} ({process_time:.3f}s)")

    return response

# æä¾›å‰ç«¯é™æ€æ–‡ä»¶
import os
static_dir = os.path.join(os.path.dirname(__file__), 'static')
if not os.path.exists(static_dir):
    os.makedirs(static_dir, exist_ok=True)

app.mount('/static', StaticFiles(directory=static_dir), name='static')

# ç¡®ä¿å›¾ç‰‡ä¸Šä¼ ç›®å½•å­˜åœ¨
uploads_dir = os.path.join(static_dir, 'uploads', 'images')
if not os.path.exists(uploads_dir):
    os.makedirs(uploads_dir, exist_ok=True)
    logger.info(f"åˆ›å»ºå›¾ç‰‡ä¸Šä¼ ç›®å½•: {uploads_dir}")

# å¥åº·æ£€æŸ¥ç«¯ç‚¹
@app.get('/health')
async def health_check():
    """å¥åº·æ£€æŸ¥ç«¯ç‚¹ï¼Œç”¨äºDockerå¥åº·æ£€æŸ¥å’Œè´Ÿè½½å‡è¡¡å™¨"""
    try:
        # æ£€æŸ¥Cookieç®¡ç†å™¨çŠ¶æ€
        manager_status = "ok" if cookie_manager.manager is not None else "error"

        # æ£€æŸ¥æ•°æ®åº“è¿æ¥
        from db_manager import db_manager
        try:
            db_manager.get_all_cookies()
            db_status = "ok"
        except Exception:
            db_status = "error"

        # è·å–ç³»ç»ŸçŠ¶æ€
        import psutil
        cpu_percent = psutil.cpu_percent(interval=1)
        memory_info = psutil.virtual_memory()

        status = {
            "status": "healthy" if manager_status == "ok" and db_status == "ok" else "unhealthy",
            "timestamp": time.time(),
            "services": {
                "cookie_manager": manager_status,
                "database": db_status
            },
            "system": {
                "cpu_percent": cpu_percent,
                "memory_percent": memory_info.percent,
                "memory_available": memory_info.available
            }
        }

        if status["status"] == "unhealthy":
            raise HTTPException(status_code=503, detail=status)

        return status

    except Exception as e:
        return {
            "status": "unhealthy",
            "timestamp": time.time(),
            "error": str(e)
        }


# é‡å®šå‘æ ¹è·¯å¾„åˆ°ç™»å½•é¡µé¢
@app.get('/', response_class=HTMLResponse)
async def root():
    login_path = os.path.join(static_dir, 'login.html')
    if os.path.exists(login_path):
        with open(login_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Login page not found</h3>')


# ç™»å½•é¡µé¢è·¯ç”±
@app.get('/login.html', response_class=HTMLResponse)
async def login_page():
    login_path = os.path.join(static_dir, 'login.html')
    if os.path.exists(login_path):
        with open(login_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Login page not found</h3>')


# æ³¨å†Œé¡µé¢è·¯ç”±
@app.get('/register.html', response_class=HTMLResponse)
async def register_page():
    # æ£€æŸ¥æ³¨å†Œæ˜¯å¦å¼€å¯
    from db_manager import db_manager
    registration_enabled = db_manager.get_system_setting('registration_enabled')
    if registration_enabled != 'true':
        return HTMLResponse('''
        <!DOCTYPE html>
        <html>
        <head>
            <title>æ³¨å†Œå·²å…³é—­</title>
            <meta charset="utf-8">
            <style>
                body { font-family: Arial, sans-serif; text-align: center; padding: 50px; }
                .message { color: #666; font-size: 18px; }
                .back-link { margin-top: 20px; }
                .back-link a { color: #007bff; text-decoration: none; }
            </style>
        </head>
        <body>
            <h2>ğŸš« æ³¨å†ŒåŠŸèƒ½å·²å…³é—­</h2>
            <p class="message">ç³»ç»Ÿç®¡ç†å‘˜å·²å…³é—­ç”¨æˆ·æ³¨å†ŒåŠŸèƒ½</p>
            <div class="back-link">
                <a href="/">â† è¿”å›é¦–é¡µ</a>
            </div>
        </body>
        </html>
        ''', status_code=403)

    register_path = os.path.join(static_dir, 'register.html')
    if os.path.exists(register_path):
        with open(register_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Register page not found</h3>')


# ç®¡ç†é¡µé¢ï¼ˆä¸éœ€è¦æœåŠ¡å™¨ç«¯è®¤è¯ï¼Œç”±å‰ç«¯JavaScriptå¤„ç†ï¼‰
@app.get('/admin', response_class=HTMLResponse)
async def admin_page():
    index_path = os.path.join(static_dir, 'index.html')
    if not os.path.exists(index_path):
        return HTMLResponse('<h3>No front-end found</h3>')
    
    # è·å–é™æ€æ–‡ä»¶çš„ä¿®æ”¹æ—¶é—´ä½œä¸ºç‰ˆæœ¬å·ï¼Œè§£å†³æµè§ˆå™¨ç¼“å­˜é—®é¢˜
    def get_file_version(file_path, default='1.0.0'):
        """è·å–æ–‡ä»¶çš„ç‰ˆæœ¬å·ï¼ˆåŸºäºä¿®æ”¹æ—¶é—´ï¼‰"""
        if os.path.exists(file_path):
            try:
                mtime = os.path.getmtime(file_path)
                return str(int(mtime))
            except Exception as e:
                logger.warning(f"è·å–æ–‡ä»¶ {file_path} ä¿®æ”¹æ—¶é—´å¤±è´¥: {e}")
        return default
    
    app_js_path = os.path.join(static_dir, 'js', 'app.js')
    app_css_path = os.path.join(static_dir, 'css', 'app.css')
    
    js_version = get_file_version(app_js_path, '2.2.0')
    css_version = get_file_version(app_css_path, '1.0.0')
    
    try:
        with open(index_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
            
            # æ›¿æ¢ app.js çš„ç‰ˆæœ¬å·å‚æ•°
            js_pattern = r'/static/js/app\.js\?v=[^"\'\s>]+'
            js_new_url = f'/static/js/app.js?v={js_version}'
            if re.search(js_pattern, html_content):
                html_content = re.sub(js_pattern, js_new_url, html_content)
                logger.debug(f"å·²æ›¿æ¢ app.js ç‰ˆæœ¬å·: {js_version}")
            
            # ä¸º app.css æ·»åŠ æˆ–æ›´æ–°ç‰ˆæœ¬å·å‚æ•°
            css_pattern = r'/static/css/app\.css(\?v=[^"\'\s>]+)?'
            css_new_url = f'/static/css/app.css?v={css_version}'
            html_content = re.sub(css_pattern, css_new_url, html_content)
            
            return HTMLResponse(html_content)
    except Exception as e:
        logger.error(f"è¯»å–æˆ–å¤„ç† index.html å¤±è´¥: {e}")
        return HTMLResponse('<h3>Error loading page</h3>')
















# ç™»å½•æ¥å£
@app.post('/login')
async def login(request: LoginRequest):
    from db_manager import db_manager

    # åˆ¤æ–­ç™»å½•æ–¹å¼
    if request.username and request.password:
        # ç”¨æˆ·å/å¯†ç ç™»å½•
        logger.info(f"ã€{request.username}ã€‘å°è¯•ç”¨æˆ·åç™»å½•")

        # ç»Ÿä¸€ä½¿ç”¨ç”¨æˆ·è¡¨éªŒè¯ï¼ˆåŒ…æ‹¬adminç”¨æˆ·ï¼‰
        if db_manager.verify_user_password(request.username, request.password):
            user = db_manager.get_user_by_username(request.username)
            if user:
                # ç”Ÿæˆtoken
                token = generate_token()
                SESSION_TOKENS[token] = {
                    'user_id': user['id'],
                    'username': user['username'],
                    'timestamp': time.time()
                }

                # åŒºåˆ†ç®¡ç†å‘˜å’Œæ™®é€šç”¨æˆ·çš„æ—¥å¿—
                if user['username'] == ADMIN_USERNAME:
                    logger.info(f"ã€{user['username']}#{user['id']}ã€‘ç™»å½•æˆåŠŸï¼ˆç®¡ç†å‘˜ï¼‰")
                else:
                    logger.info(f"ã€{user['username']}#{user['id']}ã€‘ç™»å½•æˆåŠŸ")

                return LoginResponse(
                    success=True,
                    token=token,
                    message="ç™»å½•æˆåŠŸ",
                    user_id=user['id'],
                    username=user['username'],
                    is_admin=(user['username'] == ADMIN_USERNAME)
                )

        logger.warning(f"ã€{request.username}ã€‘ç™»å½•å¤±è´¥ï¼šç”¨æˆ·åæˆ–å¯†ç é”™è¯¯")
        return LoginResponse(
            success=False,
            message="ç”¨æˆ·åæˆ–å¯†ç é”™è¯¯"
        )

    elif request.email and request.password:
        # é‚®ç®±/å¯†ç ç™»å½•
        logger.info(f"ã€{request.email}ã€‘å°è¯•é‚®ç®±å¯†ç ç™»å½•")

        user = db_manager.get_user_by_email(request.email)
        if user and db_manager.verify_user_password(user['username'], request.password):
            # ç”Ÿæˆtoken
            token = generate_token()
            SESSION_TOKENS[token] = {
                'user_id': user['id'],
                'username': user['username'],
                'timestamp': time.time()
            }

            logger.info(f"ã€{user['username']}#{user['id']}ã€‘é‚®ç®±ç™»å½•æˆåŠŸ")

            return LoginResponse(
                success=True,
                token=token,
                message="ç™»å½•æˆåŠŸ",
                user_id=user['id'],
                username=user['username'],
                is_admin=(user['username'] == ADMIN_USERNAME)
            )

        logger.warning(f"ã€{request.email}ã€‘é‚®ç®±ç™»å½•å¤±è´¥ï¼šé‚®ç®±æˆ–å¯†ç é”™è¯¯")
        return LoginResponse(
            success=False,
            message="é‚®ç®±æˆ–å¯†ç é”™è¯¯"
        )

    elif request.email and request.verification_code:
        # é‚®ç®±/éªŒè¯ç ç™»å½•
        logger.info(f"ã€{request.email}ã€‘å°è¯•é‚®ç®±éªŒè¯ç ç™»å½•")

        # éªŒè¯é‚®ç®±éªŒè¯ç 
        if not db_manager.verify_email_code(request.email, request.verification_code, 'login'):
            logger.warning(f"ã€{request.email}ã€‘éªŒè¯ç ç™»å½•å¤±è´¥ï¼šéªŒè¯ç é”™è¯¯æˆ–å·²è¿‡æœŸ")
            return LoginResponse(
                success=False,
                message="éªŒè¯ç é”™è¯¯æˆ–å·²è¿‡æœŸ"
            )

        # è·å–ç”¨æˆ·ä¿¡æ¯
        user = db_manager.get_user_by_email(request.email)
        if not user:
            logger.warning(f"ã€{request.email}ã€‘éªŒè¯ç ç™»å½•å¤±è´¥ï¼šç”¨æˆ·ä¸å­˜åœ¨")
            return LoginResponse(
                success=False,
                message="ç”¨æˆ·ä¸å­˜åœ¨"
            )

        # ç”Ÿæˆtoken
        token = generate_token()
        SESSION_TOKENS[token] = {
            'user_id': user['id'],
            'username': user['username'],
            'timestamp': time.time()
        }

        logger.info(f"ã€{user['username']}#{user['id']}ã€‘éªŒè¯ç ç™»å½•æˆåŠŸ")

        return LoginResponse(
            success=True,
            token=token,
            message="ç™»å½•æˆåŠŸ",
            user_id=user['id'],
            username=user['username'],
            is_admin=(user['username'] == ADMIN_USERNAME)
        )

    else:
        return LoginResponse(
            success=False,
            message="è¯·æä¾›æœ‰æ•ˆçš„ç™»å½•ä¿¡æ¯"
        )


# éªŒè¯tokenæ¥å£
@app.get('/verify')
async def verify(user_info: Optional[Dict[str, Any]] = Depends(verify_token)):
    if user_info:
        return {
            "authenticated": True,
            "user_id": user_info['user_id'],
            "username": user_info['username'],
            "is_admin": user_info['username'] == ADMIN_USERNAME
        }
    return {"authenticated": False}


# ç™»å‡ºæ¥å£
@app.post('/logout')
async def logout(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    if credentials and credentials.credentials in SESSION_TOKENS:
        del SESSION_TOKENS[credentials.credentials]
    return {"message": "å·²ç™»å‡º"}


# ä¿®æ”¹ç®¡ç†å‘˜å¯†ç æ¥å£
@app.post('/change-admin-password')
async def change_admin_password(request: ChangePasswordRequest, admin_user: Dict[str, Any] = Depends(verify_admin_token)):
    from db_manager import db_manager

    try:
        # éªŒè¯å½“å‰å¯†ç ï¼ˆä½¿ç”¨ç”¨æˆ·è¡¨éªŒè¯ï¼‰
        if not db_manager.verify_user_password('admin', request.current_password):
            return {"success": False, "message": "å½“å‰å¯†ç é”™è¯¯"}

        # æ›´æ–°å¯†ç ï¼ˆä½¿ç”¨ç”¨æˆ·è¡¨æ›´æ–°ï¼‰
        success = db_manager.update_user_password('admin', request.new_password)

        if success:
            logger.info(f"ã€admin#{admin_user['user_id']}ã€‘ç®¡ç†å‘˜å¯†ç ä¿®æ”¹æˆåŠŸ")
            return {"success": True, "message": "å¯†ç ä¿®æ”¹æˆåŠŸ"}
        else:
            return {"success": False, "message": "å¯†ç ä¿®æ”¹å¤±è´¥"}

    except Exception as e:
        logger.error(f"ä¿®æ”¹ç®¡ç†å‘˜å¯†ç å¼‚å¸¸: {e}")
        return {"success": False, "message": "ç³»ç»Ÿé”™è¯¯"}


# ç”Ÿæˆå›¾å½¢éªŒè¯ç æ¥å£
@app.post('/generate-captcha')
async def generate_captcha(request: CaptchaRequest):
    from db_manager import db_manager

    try:
        # ç”Ÿæˆå›¾å½¢éªŒè¯ç 
        captcha_text, captcha_image = db_manager.generate_captcha()

        if not captcha_image:
            return CaptchaResponse(
                success=False,
                captcha_image="",
                session_id=request.session_id,
                message="å›¾å½¢éªŒè¯ç ç”Ÿæˆå¤±è´¥"
            )

        # ä¿å­˜éªŒè¯ç åˆ°æ•°æ®åº“
        if db_manager.save_captcha(request.session_id, captcha_text):
            return CaptchaResponse(
                success=True,
                captcha_image=captcha_image,
                session_id=request.session_id,
                message="å›¾å½¢éªŒè¯ç ç”ŸæˆæˆåŠŸ"
            )
        else:
            return CaptchaResponse(
                success=False,
                captcha_image="",
                session_id=request.session_id,
                message="å›¾å½¢éªŒè¯ç ä¿å­˜å¤±è´¥"
            )

    except Exception as e:
        logger.error(f"ç”Ÿæˆå›¾å½¢éªŒè¯ç å¤±è´¥: {e}")
        return CaptchaResponse(
            success=False,
            captcha_image="",
            session_id=request.session_id,
            message="å›¾å½¢éªŒè¯ç ç”Ÿæˆå¤±è´¥"
        )


# éªŒè¯å›¾å½¢éªŒè¯ç æ¥å£
@app.post('/verify-captcha')
async def verify_captcha(request: VerifyCaptchaRequest):
    from db_manager import db_manager

    try:
        if db_manager.verify_captcha(request.session_id, request.captcha_code):
            return VerifyCaptchaResponse(
                success=True,
                message="å›¾å½¢éªŒè¯ç éªŒè¯æˆåŠŸ"
            )
        else:
            return VerifyCaptchaResponse(
                success=False,
                message="å›¾å½¢éªŒè¯ç é”™è¯¯æˆ–å·²è¿‡æœŸ"
            )

    except Exception as e:
        logger.error(f"éªŒè¯å›¾å½¢éªŒè¯ç å¤±è´¥: {e}")
        return VerifyCaptchaResponse(
            success=False,
            message="å›¾å½¢éªŒè¯ç éªŒè¯å¤±è´¥"
        )


# å‘é€éªŒè¯ç æ¥å£ï¼ˆéœ€è¦å…ˆéªŒè¯å›¾å½¢éªŒè¯ç ï¼‰
@app.post('/send-verification-code')
async def send_verification_code(request: SendCodeRequest):
    from db_manager import db_manager

    try:
        # æ£€æŸ¥æ˜¯å¦å·²éªŒè¯å›¾å½¢éªŒè¯ç 
        # é€šè¿‡æ£€æŸ¥æ•°æ®åº“ä¸­æ˜¯å¦å­˜åœ¨å·²éªŒè¯çš„å›¾å½¢éªŒè¯ç è®°å½•
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            current_time = time.time()

            # æŸ¥æ‰¾æœ€è¿‘5åˆ†é’Ÿå†…è¯¥session_idçš„éªŒè¯è®°å½•
            # ç”±äºéªŒè¯æˆåŠŸåéªŒè¯ç ä¼šè¢«åˆ é™¤ï¼Œæˆ‘ä»¬éœ€è¦å¦ä¸€ç§æ–¹å¼æ¥è·Ÿè¸ªéªŒè¯çŠ¶æ€
            # è¿™é‡Œæˆ‘ä»¬æ£€æŸ¥è¯¥session_idæ˜¯å¦åœ¨æœ€è¿‘éªŒè¯è¿‡ï¼ˆé€šè¿‡æ£€æŸ¥æ˜¯å¦æœ‰å·²åˆ é™¤çš„è®°å½•ï¼‰

            # ä¸ºäº†ç®€åŒ–ï¼Œæˆ‘ä»¬è¦æ±‚å‰ç«¯åœ¨éªŒè¯å›¾å½¢éªŒè¯ç æˆåŠŸåç«‹å³å‘é€é‚®ä»¶éªŒè¯ç 
            # æˆ–è€…æˆ‘ä»¬å¯ä»¥åœ¨éªŒè¯æˆåŠŸåè®¾ç½®ä¸€ä¸ªä¸´æ—¶æ ‡è®°
            pass

        # æ ¹æ®éªŒè¯ç ç±»å‹è¿›è¡Œä¸åŒçš„æ£€æŸ¥
        if request.type == 'register':
            # æ³¨å†ŒéªŒè¯ç ï¼šæ£€æŸ¥é‚®ç®±æ˜¯å¦å·²æ³¨å†Œ
            existing_user = db_manager.get_user_by_email(request.email)
            if existing_user:
                return SendCodeResponse(
                    success=False,
                    message="è¯¥é‚®ç®±å·²è¢«æ³¨å†Œ"
                )
        elif request.type == 'login':
            # ç™»å½•éªŒè¯ç ï¼šæ£€æŸ¥é‚®ç®±æ˜¯å¦å­˜åœ¨
            existing_user = db_manager.get_user_by_email(request.email)
            if not existing_user:
                return SendCodeResponse(
                    success=False,
                    message="è¯¥é‚®ç®±æœªæ³¨å†Œ"
                )

        # ç”ŸæˆéªŒè¯ç 
        code = db_manager.generate_verification_code()

        # ä¿å­˜éªŒè¯ç åˆ°æ•°æ®åº“
        if not db_manager.save_verification_code(request.email, code, request.type):
            return SendCodeResponse(
                success=False,
                message="éªŒè¯ç ä¿å­˜å¤±è´¥ï¼Œè¯·ç¨åé‡è¯•"
            )

        # å‘é€éªŒè¯ç é‚®ä»¶
        if await db_manager.send_verification_email(request.email, code):
            return SendCodeResponse(
                success=True,
                message="éªŒè¯ç å·²å‘é€åˆ°æ‚¨çš„é‚®ç®±ï¼Œè¯·æŸ¥æ”¶"
            )
        else:
            return SendCodeResponse(
                success=False,
                message="éªŒè¯ç å‘é€å¤±è´¥ï¼Œè¯·æ£€æŸ¥é‚®ç®±åœ°å€æˆ–ç¨åé‡è¯•"
            )

    except Exception as e:
        logger.error(f"å‘é€éªŒè¯ç å¤±è´¥: {e}")
        return SendCodeResponse(
            success=False,
            message="å‘é€éªŒè¯ç å¤±è´¥ï¼Œè¯·ç¨åé‡è¯•"
        )


# ç”¨æˆ·æ³¨å†Œæ¥å£
@app.post('/register')
async def register(request: RegisterRequest):
    from db_manager import db_manager

    # æ£€æŸ¥æ³¨å†Œæ˜¯å¦å¼€å¯
    registration_enabled = db_manager.get_system_setting('registration_enabled')
    if registration_enabled != 'true':
        logger.warning(f"ã€{request.username}ã€‘æ³¨å†Œå¤±è´¥: æ³¨å†ŒåŠŸèƒ½å·²å…³é—­")
        return RegisterResponse(
            success=False,
            message="æ³¨å†ŒåŠŸèƒ½å·²å…³é—­ï¼Œè¯·è”ç³»ç®¡ç†å‘˜"
        )

    try:
        logger.info(f"ã€{request.username}ã€‘å°è¯•æ³¨å†Œï¼Œé‚®ç®±: {request.email}")

        # éªŒè¯é‚®ç®±éªŒè¯ç 
        if not db_manager.verify_email_code(request.email, request.verification_code):
            logger.warning(f"ã€{request.username}ã€‘æ³¨å†Œå¤±è´¥: éªŒè¯ç é”™è¯¯æˆ–å·²è¿‡æœŸ")
            return RegisterResponse(
                success=False,
                message="éªŒè¯ç é”™è¯¯æˆ–å·²è¿‡æœŸ"
            )

        # æ£€æŸ¥ç”¨æˆ·åæ˜¯å¦å·²å­˜åœ¨
        existing_user = db_manager.get_user_by_username(request.username)
        if existing_user:
            logger.warning(f"ã€{request.username}ã€‘æ³¨å†Œå¤±è´¥: ç”¨æˆ·åå·²å­˜åœ¨")
            return RegisterResponse(
                success=False,
                message="ç”¨æˆ·åå·²å­˜åœ¨"
            )

        # æ£€æŸ¥é‚®ç®±æ˜¯å¦å·²æ³¨å†Œ
        existing_email = db_manager.get_user_by_email(request.email)
        if existing_email:
            logger.warning(f"ã€{request.username}ã€‘æ³¨å†Œå¤±è´¥: é‚®ç®±å·²è¢«æ³¨å†Œ")
            return RegisterResponse(
                success=False,
                message="è¯¥é‚®ç®±å·²è¢«æ³¨å†Œ"
            )

        # åˆ›å»ºç”¨æˆ·
        if db_manager.create_user(request.username, request.email, request.password):
            logger.info(f"ã€{request.username}ã€‘æ³¨å†ŒæˆåŠŸ")
            return RegisterResponse(
                success=True,
                message="æ³¨å†ŒæˆåŠŸï¼Œè¯·ç™»å½•"
            )
        else:
            logger.error(f"ã€{request.username}ã€‘æ³¨å†Œå¤±è´¥: æ•°æ®åº“æ“ä½œå¤±è´¥")
            return RegisterResponse(
                success=False,
                message="æ³¨å†Œå¤±è´¥ï¼Œè¯·ç¨åé‡è¯•"
            )

    except Exception as e:
        logger.error(f"ã€{request.username}ã€‘æ³¨å†Œå¼‚å¸¸: {e}")
        return RegisterResponse(
            success=False,
            message="æ³¨å†Œå¤±è´¥ï¼Œè¯·ç¨åé‡è¯•"
        )


# ------------------------- å‘é€æ¶ˆæ¯æ¥å£ -------------------------

# å›ºå®šçš„APIç§˜é’¥ï¼ˆç”Ÿäº§ç¯å¢ƒä¸­åº”è¯¥ä»é…ç½®æ–‡ä»¶æˆ–ç¯å¢ƒå˜é‡è¯»å–ï¼‰
# æ³¨æ„ï¼šç°åœ¨ä»ç³»ç»Ÿè®¾ç½®ä¸­è¯»å–QQå›å¤æ¶ˆæ¯ç§˜é’¥
API_SECRET_KEY = "xianyu_api_secret_2024"  # ä¿ç•™ä½œä¸ºåå¤‡

class SendMessageRequest(BaseModel):
    api_key: str
    cookie_id: str
    chat_id: str
    to_user_id: str
    message: str


class SendMessageResponse(BaseModel):
    success: bool
    message: str


def verify_api_key(api_key: str) -> bool:
    """éªŒè¯APIç§˜é’¥"""
    try:
        # ä»ç³»ç»Ÿè®¾ç½®ä¸­è·å–QQå›å¤æ¶ˆæ¯ç§˜é’¥
        from db_manager import db_manager
        qq_secret_key = db_manager.get_system_setting('qq_reply_secret_key')

        # å¦‚æœç³»ç»Ÿè®¾ç½®ä¸­æ²¡æœ‰é…ç½®ï¼Œä½¿ç”¨é»˜è®¤å€¼
        if not qq_secret_key:
            qq_secret_key = API_SECRET_KEY

        return api_key == qq_secret_key
    except Exception as e:
        logger.error(f"éªŒè¯APIç§˜é’¥æ—¶å‘ç”Ÿå¼‚å¸¸: {e}")
        # å¼‚å¸¸æƒ…å†µä¸‹ä½¿ç”¨é»˜è®¤ç§˜é’¥éªŒè¯
        return api_key == API_SECRET_KEY


@app.post('/send-message', response_model=SendMessageResponse)
async def send_message_api(request: SendMessageRequest):
    """å‘é€æ¶ˆæ¯APIæ¥å£ï¼ˆä½¿ç”¨ç§˜é’¥éªŒè¯ï¼‰"""
    try:
        # æ¸…ç†æ‰€æœ‰å‚æ•°ä¸­çš„æ¢è¡Œç¬¦
        def clean_param(param_str):
            """æ¸…ç†å‚æ•°ä¸­çš„æ¢è¡Œç¬¦"""
            if isinstance(param_str, str):
                return param_str.replace('\\n', '').replace('\n', '')
            return param_str

        # æ¸…ç†æ‰€æœ‰å‚æ•°
        cleaned_api_key = clean_param(request.api_key)
        cleaned_cookie_id = clean_param(request.cookie_id)
        cleaned_chat_id = clean_param(request.chat_id)
        cleaned_to_user_id = clean_param(request.to_user_id)
        cleaned_message = clean_param(request.message)

        # éªŒè¯APIç§˜é’¥ä¸èƒ½ä¸ºç©º
        if not cleaned_api_key:
            logger.warning("APIç§˜é’¥ä¸ºç©º")
            return SendMessageResponse(
                success=False,
                message="APIç§˜é’¥ä¸èƒ½ä¸ºç©º"
            )

        # ç‰¹æ®Šæµ‹è¯•ç§˜é’¥å¤„ç†
        if cleaned_api_key == "zhinina_test_key":
            logger.info("ä½¿ç”¨æµ‹è¯•ç§˜é’¥ï¼Œç›´æ¥è¿”å›æˆåŠŸ")
            return SendMessageResponse(
                success=True,
                message="æ¥å£éªŒè¯æˆåŠŸ"
            )

        # éªŒè¯APIç§˜é’¥
        if not verify_api_key(cleaned_api_key):
            logger.warning(f"APIç§˜é’¥éªŒè¯å¤±è´¥: {cleaned_api_key}")
            return SendMessageResponse(
                success=False,
                message="APIç§˜é’¥éªŒè¯å¤±è´¥"
            )

        # éªŒè¯å¿…éœ€å‚æ•°ä¸èƒ½ä¸ºç©º
        required_params = {
            'cookie_id': cleaned_cookie_id,
            'chat_id': cleaned_chat_id,
            'to_user_id': cleaned_to_user_id,
            'message': cleaned_message
        }

        for param_name, param_value in required_params.items():
            if not param_value:
                logger.warning(f"å¿…éœ€å‚æ•° {param_name} ä¸ºç©º")
                return SendMessageResponse(
                    success=False,
                    message=f"å‚æ•° {param_name} ä¸èƒ½ä¸ºç©º"
                )

        # ç›´æ¥è·å–XianyuLiveå®ä¾‹ï¼Œè·³è¿‡cookie_manageræ£€æŸ¥
        from XianyuAutoAsync import XianyuLive
        live_instance = XianyuLive.get_instance(cleaned_cookie_id)

        if not live_instance:
            logger.warning(f"è´¦å·å®ä¾‹ä¸å­˜åœ¨æˆ–æœªè¿æ¥: {cleaned_cookie_id}")
            return SendMessageResponse(
                success=False,
                message="è´¦å·å®ä¾‹ä¸å­˜åœ¨æˆ–æœªè¿æ¥ï¼Œè¯·æ£€æŸ¥è´¦å·çŠ¶æ€"
            )

        # æ£€æŸ¥WebSocketè¿æ¥çŠ¶æ€
        if not live_instance.ws or live_instance.ws.closed:
            logger.warning(f"è´¦å·WebSocketè¿æ¥å·²æ–­å¼€: {cleaned_cookie_id}")
            return SendMessageResponse(
                success=False,
                message="è´¦å·WebSocketè¿æ¥å·²æ–­å¼€ï¼Œè¯·ç­‰å¾…é‡è¿"
            )

        # å‘é€æ¶ˆæ¯ï¼ˆä½¿ç”¨æ¸…ç†åçš„æ‰€æœ‰å‚æ•°ï¼‰
        await live_instance.send_msg(
            live_instance.ws,
            cleaned_chat_id,
            cleaned_to_user_id,
            cleaned_message
        )

        logger.info(f"APIæˆåŠŸå‘é€æ¶ˆæ¯: {cleaned_cookie_id} -> {cleaned_to_user_id}, å†…å®¹: {cleaned_message[:50]}{'...' if len(cleaned_message) > 50 else ''}")

        return SendMessageResponse(
            success=True,
            message="æ¶ˆæ¯å‘é€æˆåŠŸ"
        )

    except Exception as e:
        # ä½¿ç”¨æ¸…ç†åçš„å‚æ•°è®°å½•æ—¥å¿—
        cookie_id_for_log = clean_param(request.cookie_id) if 'clean_param' in locals() else request.cookie_id
        to_user_id_for_log = clean_param(request.to_user_id) if 'clean_param' in locals() else request.to_user_id
        logger.error(f"APIå‘é€æ¶ˆæ¯å¼‚å¸¸: {cookie_id_for_log} -> {to_user_id_for_log}, é”™è¯¯: {str(e)}")
        return SendMessageResponse(
            success=False,
            message=f"å‘é€æ¶ˆæ¯å¤±è´¥: {str(e)}"
        )


@app.post("/xianyu/reply", response_model=ResponseModel)
async def xianyu_reply(req: RequestModel):
    msg_template = match_reply(req.cookie_id, req.send_message)
    is_default_reply = False

    if not msg_template:
        # ä»æ•°æ®åº“è·å–é»˜è®¤å›å¤
        from db_manager import db_manager
        default_reply_settings = db_manager.get_default_reply(req.cookie_id)

        if default_reply_settings and default_reply_settings.get('enabled', False):
            # æ£€æŸ¥æ˜¯å¦å¼€å¯äº†"åªå›å¤ä¸€æ¬¡"åŠŸèƒ½
            if default_reply_settings.get('reply_once', False):
                # æ£€æŸ¥æ˜¯å¦å·²ç»å›å¤è¿‡è¿™ä¸ªchat_id
                if db_manager.has_default_reply_record(req.cookie_id, req.chat_id):
                    raise HTTPException(status_code=404, detail="è¯¥å¯¹è¯å·²ä½¿ç”¨é»˜è®¤å›å¤ï¼Œä¸å†é‡å¤å›å¤")

            msg_template = default_reply_settings.get('reply_content', '')
            is_default_reply = True

        # å¦‚æœæ•°æ®åº“ä¸­æ²¡æœ‰è®¾ç½®æˆ–ä¸ºç©ºï¼Œè¿”å›é”™è¯¯
        if not msg_template:
            raise HTTPException(status_code=404, detail="æœªæ‰¾åˆ°åŒ¹é…çš„å›å¤è§„åˆ™ä¸”æœªè®¾ç½®é»˜è®¤å›å¤")

    # æŒ‰å ä½ç¬¦æ ¼å¼åŒ–
    try:
        send_msg = msg_template.format(
            send_user_id=req.send_user_id,
            send_user_name=req.send_user_name,
            send_message=req.send_message,
        )
    except Exception:
        # å¦‚æœæ ¼å¼åŒ–å¤±è´¥ï¼Œè¿”å›åŸå§‹å†…å®¹
        send_msg = msg_template

    # å¦‚æœæ˜¯é»˜è®¤å›å¤ä¸”å¼€å¯äº†"åªå›å¤ä¸€æ¬¡"ï¼Œè®°å½•å›å¤è®°å½•
    if is_default_reply:
        from db_manager import db_manager
        default_reply_settings = db_manager.get_default_reply(req.cookie_id)
        if default_reply_settings and default_reply_settings.get('reply_once', False):
            db_manager.add_default_reply_record(req.cookie_id, req.chat_id)

    return {"code": 200, "data": {"send_msg": send_msg}}

# ------------------------- è´¦å· / å…³é”®å­—ç®¡ç†æ¥å£ -------------------------


class CookieIn(BaseModel):
    id: str
    value: str


class CookieStatusIn(BaseModel):
    enabled: bool


class DefaultReplyIn(BaseModel):
    enabled: bool
    reply_content: Optional[str] = None
    reply_once: bool = False


class NotificationChannelIn(BaseModel):
    name: str
    type: str = "qq"
    config: str


class NotificationChannelUpdate(BaseModel):
    name: str
    config: str
    enabled: bool = True


class MessageNotificationIn(BaseModel):
    channel_id: int
    enabled: bool = True


class SystemSettingIn(BaseModel):
    value: str
    description: Optional[str] = None


class SystemSettingCreateIn(BaseModel):
    key: str
    value: str
    description: Optional[str] = None





@app.get("/cookies")
def list_cookies(current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        return []

    # è·å–å½“å‰ç”¨æˆ·çš„cookies
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)
    return list(user_cookies.keys())


@app.get("/cookies/details")
def get_cookies_details(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–æ‰€æœ‰Cookieçš„è¯¦ç»†ä¿¡æ¯ï¼ˆåŒ…æ‹¬å€¼å’ŒçŠ¶æ€ï¼‰"""
    if cookie_manager.manager is None:
        return []

    # è·å–å½“å‰ç”¨æˆ·çš„cookies
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    result = []
    for cookie_id, cookie_value in user_cookies.items():
        cookie_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
        auto_confirm = db_manager.get_auto_confirm(cookie_id)
        # è·å–å¤‡æ³¨ä¿¡æ¯
        cookie_details = db_manager.get_cookie_details(cookie_id)
        remark = cookie_details.get('remark', '') if cookie_details else ''

        result.append({
            'id': cookie_id,
            'value': cookie_value,
            'enabled': cookie_enabled,
            'auto_confirm': auto_confirm,
            'remark': remark,
            'pause_duration': cookie_details.get('pause_duration', 10) if cookie_details else 10
        })
    return result


@app.post("/cookies")
def add_cookie(item: CookieIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")
    try:
        # æ·»åŠ cookieæ—¶ç»‘å®šåˆ°å½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager

        log_with_user('info', f"å°è¯•æ·»åŠ Cookie: {item.id}, å½“å‰ç”¨æˆ·ID: {user_id}, ç”¨æˆ·å: {current_user.get('username', 'unknown')}", current_user)

        # æ£€æŸ¥cookieæ˜¯å¦å·²å­˜åœ¨ä¸”å±äºå…¶ä»–ç”¨æˆ·
        existing_cookies = db_manager.get_all_cookies()
        if item.id in existing_cookies:
            # æ£€æŸ¥æ˜¯å¦å±äºå½“å‰ç”¨æˆ·
            user_cookies = db_manager.get_all_cookies(user_id)
            if item.id not in user_cookies:
                log_with_user('warning', f"Cookie IDå†²çª: {item.id} å·²è¢«å…¶ä»–ç”¨æˆ·ä½¿ç”¨", current_user)
                raise HTTPException(status_code=400, detail="è¯¥Cookie IDå·²è¢«å…¶ä»–ç”¨æˆ·ä½¿ç”¨")

        # ä¿å­˜åˆ°æ•°æ®åº“æ—¶æŒ‡å®šç”¨æˆ·ID
        db_manager.save_cookie(item.id, item.value, user_id)

        # æ·»åŠ åˆ°CookieManagerï¼ŒåŒæ—¶æŒ‡å®šç”¨æˆ·ID
        cookie_manager.manager.add_cookie(item.id, item.value, user_id=user_id)
        log_with_user('info', f"Cookieæ·»åŠ æˆåŠŸ: {item.id}", current_user)
        return {"msg": "success"}
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"æ·»åŠ Cookieå¤±è´¥: {item.id} - {str(e)}", current_user)
        raise HTTPException(status_code=400, detail=str(e))


@app.put('/cookies/{cid}')
def update_cookie(cid: str, item: CookieIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager æœªå°±ç»ª')
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # è·å–æ—§çš„ cookie å€¼ï¼Œç”¨äºåˆ¤æ–­æ˜¯å¦éœ€è¦é‡å¯ä»»åŠ¡
        old_cookie_details = db_manager.get_cookie_details(cid)
        old_cookie_value = old_cookie_details.get('value') if old_cookie_details else None

        # ä½¿ç”¨ update_cookie_account_info æ›´æ–°ï¼ˆåªæ›´æ–°cookieå€¼ï¼Œä¸è¦†ç›–å…¶ä»–å­—æ®µï¼‰
        success = db_manager.update_cookie_account_info(cid, cookie_value=item.value)
        
        if not success:
            raise HTTPException(status_code=400, detail="æ›´æ–°Cookieå¤±è´¥")
        
        # åªæœ‰å½“ cookie å€¼çœŸçš„å‘ç”Ÿå˜åŒ–æ—¶æ‰é‡å¯ä»»åŠ¡
        if item.value != old_cookie_value:
            logger.info(f"Cookieå€¼å·²å˜åŒ–ï¼Œé‡å¯ä»»åŠ¡: {cid}")
            cookie_manager.manager.update_cookie(cid, item.value, save_to_db=False)
        else:
            logger.info(f"Cookieå€¼æœªå˜åŒ–ï¼Œæ— éœ€é‡å¯ä»»åŠ¡: {cid}")
        
        return {'msg': 'updated'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class CookieAccountInfo(BaseModel):
    """è´¦å·ä¿¡æ¯æ›´æ–°æ¨¡å‹"""
    value: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    show_browser: Optional[bool] = None


@app.post("/cookie/{cid}/account-info")
def update_cookie_account_info(cid: str, info: CookieAccountInfo, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ›´æ–°è´¦å·ä¿¡æ¯ï¼ˆCookieã€ç”¨æˆ·åã€å¯†ç ã€æ˜¾ç¤ºæµè§ˆå™¨è®¾ç½®ï¼‰"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager æœªå°±ç»ª')
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # è·å–æ—§çš„ cookie å€¼ï¼Œç”¨äºåˆ¤æ–­æ˜¯å¦éœ€è¦é‡å¯ä»»åŠ¡
        old_cookie_details = db_manager.get_cookie_details(cid)
        old_cookie_value = old_cookie_details.get('value') if old_cookie_details else None
        
        # æ›´æ–°æ•°æ®åº“
        success = db_manager.update_cookie_account_info(
            cid, 
            cookie_value=info.value,
            username=info.username,
            password=info.password,
            show_browser=info.show_browser
        )
        
        if not success:
            raise HTTPException(status_code=400, detail="æ›´æ–°è´¦å·ä¿¡æ¯å¤±è´¥")
        
        # åªæœ‰å½“ cookie å€¼çœŸçš„å‘ç”Ÿå˜åŒ–æ—¶æ‰é‡å¯ä»»åŠ¡
        if info.value is not None and info.value != old_cookie_value:
            logger.info(f"Cookieå€¼å·²å˜åŒ–ï¼Œé‡å¯ä»»åŠ¡: {cid}")
            cookie_manager.manager.update_cookie(cid, info.value, save_to_db=False)
        else:
            logger.info(f"Cookieå€¼æœªå˜åŒ–ï¼Œæ— éœ€é‡å¯ä»»åŠ¡: {cid}")
        
        return {'msg': 'updated', 'success': True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"æ›´æ–°è´¦å·ä¿¡æ¯å¤±è´¥: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/cookie/{cid}/details")
def get_cookie_account_details(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–è´¦å·è¯¦ç»†ä¿¡æ¯ï¼ˆåŒ…æ‹¬ç”¨æˆ·åã€å¯†ç ã€æ˜¾ç¤ºæµè§ˆå™¨è®¾ç½®ï¼‰"""
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # è·å–è¯¦ç»†ä¿¡æ¯
        details = db_manager.get_cookie_details(cid)
        
        if not details:
            raise HTTPException(status_code=404, detail="è´¦å·ä¸å­˜åœ¨")
        
        return details
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"è·å–è´¦å·è¯¦æƒ…å¤±è´¥: {e}")
        raise HTTPException(status_code=400, detail=str(e))


# ========================= è´¦å·å¯†ç ç™»å½•ç›¸å…³æ¥å£ =========================

async def _execute_password_login(session_id: str, account_id: str, account: str, password: str, show_browser: bool, user_id: int, current_user: Dict[str, Any]):
    """åå°æ‰§è¡Œè´¦å·å¯†ç ç™»å½•ä»»åŠ¡"""
    try:
        log_with_user('info', f"å¼€å§‹æ‰§è¡Œè´¦å·å¯†ç ç™»å½•ä»»åŠ¡: {session_id}, è´¦å·: {account_id}", current_user)
        
        # å¯¼å…¥ XianyuSliderStealth
        from utils.xianyu_slider_stealth import XianyuSliderStealth
        import base64
        import io
        
        # åˆ›å»º XianyuSliderStealth å®ä¾‹
        slider_instance = XianyuSliderStealth(
            user_id=account_id,
            enable_learning=True,
            headless=not show_browser
        )
        
        # æ›´æ–°ä¼šè¯ä¿¡æ¯
        password_login_sessions[session_id]['slider_instance'] = slider_instance
        
        # å®šä¹‰é€šçŸ¥å›è°ƒå‡½æ•°ï¼Œç”¨äºæ£€æµ‹åˆ°äººè„¸è®¤è¯æ—¶è¿”å›éªŒè¯é“¾æ¥æˆ–æˆªå›¾ï¼ˆåŒæ­¥å‡½æ•°ï¼‰
        def notification_callback(message: str, screenshot_path: str = None, verification_url: str = None, screenshot_path_new: str = None):
            """äººè„¸è®¤è¯é€šçŸ¥å›è°ƒï¼ˆåŒæ­¥ï¼‰
            
            Args:
                message: é€šçŸ¥æ¶ˆæ¯
                screenshot_path: æ—§ç‰ˆæˆªå›¾è·¯å¾„ï¼ˆå…¼å®¹å‚æ•°ï¼‰
                verification_url: éªŒè¯é“¾æ¥
                screenshot_path_new: æ–°ç‰ˆæˆªå›¾è·¯å¾„ï¼ˆæ–°å‚æ•°ï¼Œä¼˜å…ˆä½¿ç”¨ï¼‰
            """
            try:
                # ä¼˜å…ˆä½¿ç”¨æ–°çš„æˆªå›¾è·¯å¾„å‚æ•°
                actual_screenshot_path = screenshot_path_new if screenshot_path_new else screenshot_path
                
                # ä¼˜å…ˆä½¿ç”¨æˆªå›¾è·¯å¾„ï¼Œå¦‚æœæ²¡æœ‰æˆªå›¾åˆ™ä½¿ç”¨éªŒè¯é“¾æ¥
                if actual_screenshot_path and os.path.exists(actual_screenshot_path):
                    # æ›´æ–°ä¼šè¯çŠ¶æ€ï¼Œä¿å­˜æˆªå›¾è·¯å¾„
                    password_login_sessions[session_id]['status'] = 'verification_required'
                    password_login_sessions[session_id]['screenshot_path'] = actual_screenshot_path
                    password_login_sessions[session_id]['verification_url'] = None
                    password_login_sessions[session_id]['qr_code_url'] = None
                    log_with_user('info', f"äººè„¸è®¤è¯æˆªå›¾å·²ä¿å­˜: {session_id}, è·¯å¾„: {actual_screenshot_path}", current_user)
                    
                    # å‘é€é€šçŸ¥åˆ°ç”¨æˆ·é…ç½®çš„æ¸ é“
                    def send_face_verification_notification():
                        """åœ¨åå°çº¿ç¨‹ä¸­å‘é€äººè„¸éªŒè¯é€šçŸ¥"""
                        try:
                            from XianyuAutoAsync import XianyuLive
                            log_with_user('info', f"å¼€å§‹å°è¯•å‘é€äººè„¸éªŒè¯é€šçŸ¥: {account_id}", current_user)
                            
                            # å°è¯•è·å–XianyuLiveå®ä¾‹ï¼ˆå¦‚æœè´¦å·å·²ç»å­˜åœ¨ï¼‰
                            live_instance = XianyuLive.get_instance(account_id)
                            
                            if live_instance:
                                log_with_user('info', f"æ‰¾åˆ°è´¦å·å®ä¾‹ï¼Œå‡†å¤‡å‘é€é€šçŸ¥: {account_id}", current_user)
                                # åˆ›å»ºæ–°çš„äº‹ä»¶å¾ªç¯æ¥è¿è¡Œå¼‚æ­¥é€šçŸ¥
                                new_loop = asyncio.new_event_loop()
                                asyncio.set_event_loop(new_loop)
                                try:
                                    new_loop.run_until_complete(
                                        live_instance.send_token_refresh_notification(
                                            error_message=message,
                                            notification_type="face_verification",
                                            verification_url=None,
                                            attachment_path=actual_screenshot_path
                                        )
                                    )
                                    log_with_user('info', f"âœ… å·²å‘é€äººè„¸éªŒè¯é€šçŸ¥: {account_id}", current_user)
                                except Exception as notify_err:
                                    log_with_user('error', f"å‘é€äººè„¸éªŒè¯é€šçŸ¥å¤±è´¥: {str(notify_err)}", current_user)
                                    import traceback
                                    log_with_user('error', f"é€šçŸ¥é”™è¯¯è¯¦æƒ…: {traceback.format_exc()}", current_user)
                                finally:
                                    new_loop.close()
                            else:
                                # å¦‚æœè´¦å·å®ä¾‹ä¸å­˜åœ¨ï¼Œè®°å½•è­¦å‘Šå¹¶å°è¯•ä»æ•°æ®åº“è·å–é€šçŸ¥é…ç½®
                                log_with_user('warning', f"è´¦å·å®ä¾‹ä¸å­˜åœ¨: {account_id}ï¼Œå°è¯•ä»æ•°æ®åº“è·å–é€šçŸ¥é…ç½®", current_user)
                                try:
                                    # å°è¯•ä»æ•°æ®åº“è·å–é€šçŸ¥é…ç½®
                                    notifications = db_manager.get_account_notifications(account_id)
                                    if notifications:
                                        log_with_user('info', f"æ‰¾åˆ° {len(notifications)} ä¸ªé€šçŸ¥é…ç½®ï¼Œä½†éœ€è¦è´¦å·å®ä¾‹æ‰èƒ½å‘é€", current_user)
                                        log_with_user('warning', f"è´¦å·å®ä¾‹ä¸å­˜åœ¨ï¼Œæ— æ³•å‘é€é€šçŸ¥: {account_id}ã€‚è¯·ç¡®ä¿è´¦å·å·²ç™»å½•å¹¶è¿è¡Œä¸­ã€‚", current_user)
                                    else:
                                        log_with_user('warning', f"è´¦å· {account_id} æœªé…ç½®é€šçŸ¥æ¸ é“", current_user)
                                except Exception as db_err:
                                    log_with_user('error', f"è·å–é€šçŸ¥é…ç½®å¤±è´¥: {str(db_err)}", current_user)
                        except Exception as notify_err:
                            log_with_user('error', f"å‘é€äººè„¸éªŒè¯é€šçŸ¥æ—¶å‡ºé”™: {str(notify_err)}", current_user)
                            import traceback
                            log_with_user('error', f"é€šçŸ¥é”™è¯¯è¯¦æƒ…: {traceback.format_exc()}", current_user)
                    
                    # åœ¨åå°çº¿ç¨‹ä¸­å‘é€é€šçŸ¥ï¼Œé¿å…é˜»å¡ç™»å½•æµç¨‹
                    import threading
                    notification_thread = threading.Thread(target=send_face_verification_notification)
                    notification_thread.daemon = True
                    notification_thread.start()
                    log_with_user('info', f"å·²å¯åŠ¨äººè„¸éªŒè¯é€šçŸ¥å‘é€çº¿ç¨‹: {account_id}", current_user)
                elif verification_url:
                    # å¦‚æœæ²¡æœ‰æˆªå›¾ï¼Œä½¿ç”¨éªŒè¯é“¾æ¥ï¼ˆå…¼å®¹æ—§ç‰ˆæœ¬ï¼‰
                    password_login_sessions[session_id]['status'] = 'verification_required'
                    password_login_sessions[session_id]['verification_url'] = verification_url
                    password_login_sessions[session_id]['screenshot_path'] = None
                    password_login_sessions[session_id]['qr_code_url'] = None
                    log_with_user('info', f"äººè„¸è®¤è¯éªŒè¯é“¾æ¥å·²ä¿å­˜: {session_id}, URL: {verification_url}", current_user)
                    
                    # å‘é€é€šçŸ¥åˆ°ç”¨æˆ·é…ç½®çš„æ¸ é“
                    def send_face_verification_notification():
                        """åœ¨åå°çº¿ç¨‹ä¸­å‘é€äººè„¸éªŒè¯é€šçŸ¥"""
                        try:
                            from XianyuAutoAsync import XianyuLive
                            log_with_user('info', f"å¼€å§‹å°è¯•å‘é€äººè„¸éªŒè¯é€šçŸ¥: {account_id}", current_user)
                            
                            # å°è¯•è·å–XianyuLiveå®ä¾‹ï¼ˆå¦‚æœè´¦å·å·²ç»å­˜åœ¨ï¼‰
                            live_instance = XianyuLive.get_instance(account_id)
                            
                            if live_instance:
                                log_with_user('info', f"æ‰¾åˆ°è´¦å·å®ä¾‹ï¼Œå‡†å¤‡å‘é€é€šçŸ¥: {account_id}", current_user)
                                # åˆ›å»ºæ–°çš„äº‹ä»¶å¾ªç¯æ¥è¿è¡Œå¼‚æ­¥é€šçŸ¥
                                new_loop = asyncio.new_event_loop()
                                asyncio.set_event_loop(new_loop)
                                try:
                                    new_loop.run_until_complete(
                                        live_instance.send_token_refresh_notification(
                                            error_message=message,
                                            notification_type="face_verification",
                                            verification_url=verification_url
                                        )
                                    )
                                    log_with_user('info', f"âœ… å·²å‘é€äººè„¸éªŒè¯é€šçŸ¥: {account_id}", current_user)
                                except Exception as notify_err:
                                    log_with_user('error', f"å‘é€äººè„¸éªŒè¯é€šçŸ¥å¤±è´¥: {str(notify_err)}", current_user)
                                    import traceback
                                    log_with_user('error', f"é€šçŸ¥é”™è¯¯è¯¦æƒ…: {traceback.format_exc()}", current_user)
                                finally:
                                    new_loop.close()
                            else:
                                # å¦‚æœè´¦å·å®ä¾‹ä¸å­˜åœ¨ï¼Œè®°å½•è­¦å‘Šå¹¶å°è¯•ä»æ•°æ®åº“è·å–é€šçŸ¥é…ç½®
                                log_with_user('warning', f"è´¦å·å®ä¾‹ä¸å­˜åœ¨: {account_id}ï¼Œå°è¯•ä»æ•°æ®åº“è·å–é€šçŸ¥é…ç½®", current_user)
                                try:
                                    # å°è¯•ä»æ•°æ®åº“è·å–é€šçŸ¥é…ç½®
                                    notifications = db_manager.get_account_notifications(account_id)
                                    if notifications:
                                        log_with_user('info', f"æ‰¾åˆ° {len(notifications)} ä¸ªé€šçŸ¥é…ç½®ï¼Œä½†éœ€è¦è´¦å·å®ä¾‹æ‰èƒ½å‘é€", current_user)
                                        log_with_user('warning', f"è´¦å·å®ä¾‹ä¸å­˜åœ¨ï¼Œæ— æ³•å‘é€é€šçŸ¥: {account_id}ã€‚è¯·ç¡®ä¿è´¦å·å·²ç™»å½•å¹¶è¿è¡Œä¸­ã€‚", current_user)
                                    else:
                                        log_with_user('warning', f"è´¦å· {account_id} æœªé…ç½®é€šçŸ¥æ¸ é“", current_user)
                                except Exception as db_err:
                                    log_with_user('error', f"è·å–é€šçŸ¥é…ç½®å¤±è´¥: {str(db_err)}", current_user)
                        except Exception as notify_err:
                            log_with_user('error', f"å‘é€äººè„¸éªŒè¯é€šçŸ¥æ—¶å‡ºé”™: {str(notify_err)}", current_user)
                            import traceback
                            log_with_user('error', f"é€šçŸ¥é”™è¯¯è¯¦æƒ…: {traceback.format_exc()}", current_user)
                    
                    # åœ¨åå°çº¿ç¨‹ä¸­å‘é€é€šçŸ¥ï¼Œé¿å…é˜»å¡ç™»å½•æµç¨‹
                    import threading
                    notification_thread = threading.Thread(target=send_face_verification_notification)
                    notification_thread.daemon = True
                    notification_thread.start()
                    log_with_user('info', f"å·²å¯åŠ¨äººè„¸éªŒè¯é€šçŸ¥å‘é€çº¿ç¨‹: {account_id}", current_user)
            except Exception as e:
                log_with_user('error', f"å¤„ç†äººè„¸è®¤è¯é€šçŸ¥å¤±è´¥: {str(e)}", current_user)
        
        # è°ƒç”¨ç™»å½•æ–¹æ³•ï¼ˆåŒæ­¥æ–¹æ³•ï¼Œéœ€è¦åœ¨åå°çº¿ç¨‹ä¸­æ‰§è¡Œï¼‰
        import threading
        
        def run_login():
            try:
                cookies_dict = slider_instance.login_with_password_playwright(
                    account=account,
                    password=password,
                    show_browser=show_browser,
                    notification_callback=notification_callback
                )
                
                if cookies_dict is None:
                    password_login_sessions[session_id]['status'] = 'failed'
                    password_login_sessions[session_id]['error'] = 'ç™»å½•å¤±è´¥ï¼Œè¯·æ£€æŸ¥è´¦å·å¯†ç æ˜¯å¦æ­£ç¡®'
                    log_with_user('error', f"è´¦å·å¯†ç ç™»å½•å¤±è´¥: {account_id}", current_user)
                    return
                
                # å°†cookieå­—å…¸è½¬æ¢ä¸ºå­—ç¬¦ä¸²æ ¼å¼
                cookies_str = '; '.join([f"{k}={v}" for k, v in cookies_dict.items()])
                
                log_with_user('info', f"è´¦å·å¯†ç ç™»å½•æˆåŠŸï¼Œè·å–åˆ° {len(cookies_dict)} ä¸ªCookieå­—æ®µ: {account_id}", current_user)
                
                # æ£€æŸ¥æ˜¯å¦å·²å­˜åœ¨ç›¸åŒè´¦å·IDçš„Cookie
                existing_cookies = db_manager.get_all_cookies(user_id)
                is_new_account = account_id not in existing_cookies
                
                # ä¿å­˜è´¦å·å¯†ç å’ŒCookieåˆ°æ•°æ®åº“
                # ä½¿ç”¨ update_cookie_account_info æ¥ä¿å­˜ï¼Œå®ƒä¼šè‡ªåŠ¨å¤„ç†æ–°è´¦å·å’Œç°æœ‰è´¦å·çš„æƒ…å†µ
                update_success = db_manager.update_cookie_account_info(
                    account_id,
                    cookie_value=cookies_str,
                    username=account,
                    password=password,
                    show_browser=show_browser,
                    user_id=user_id  # æ–°è´¦å·æ—¶éœ€è¦æä¾›user_id
                )
                
                if update_success:
                    if is_new_account:
                        log_with_user('info', f"æ–°è´¦å·Cookieå’Œè´¦å·å¯†ç å·²ä¿å­˜: {account_id}", current_user)
                    else:
                        log_with_user('info', f"ç°æœ‰è´¦å·Cookieå’Œè´¦å·å¯†ç å·²æ›´æ–°: {account_id}", current_user)
                else:
                    log_with_user('error', f"ä¿å­˜è´¦å·ä¿¡æ¯å¤±è´¥: {account_id}", current_user)
                
                # æ·»åŠ åˆ°æˆ–æ›´æ–°cookie_managerï¼ˆæ³¨æ„ï¼šä¸è¦åœ¨è¿™é‡Œè°ƒç”¨add_cookieæˆ–update_cookieï¼Œå› ä¸ºå®ƒä»¬ä¼šè¦†ç›–è´¦å·å¯†ç ï¼‰
                # è´¦å·å¯†ç å·²ç»åœ¨ä¸Šé¢é€šè¿‡update_cookie_account_infoä¿å­˜äº†
                # è¿™é‡Œåªéœ€è¦æ›´æ–°å†…å­˜ä¸­çš„cookieå€¼ï¼Œä¸ä¿å­˜åˆ°æ•°æ®åº“ï¼ˆé¿å…è¦†ç›–è´¦å·å¯†ç ï¼‰
                if cookie_manager.manager:
                    # æ›´æ–°å†…å­˜ä¸­çš„cookieå€¼
                    cookie_manager.manager.cookies[account_id] = cookies_str
                    log_with_user('info', f"å·²æ›´æ–°cookie_managerä¸­çš„Cookieï¼ˆå†…å­˜ï¼‰: {account_id}", current_user)
                    
                    # å¦‚æœæ˜¯æ–°è´¦å·ï¼Œéœ€è¦å¯åŠ¨ä»»åŠ¡
                    if is_new_account:
                        # ä½¿ç”¨å¼‚æ­¥æ–¹å¼å¯åŠ¨ä»»åŠ¡ï¼Œä½†ä¸ä¿å­˜åˆ°æ•°æ®åº“ï¼ˆé¿å…è¦†ç›–è´¦å·å¯†ç ï¼‰
                        try:
                            import asyncio
                            loop = cookie_manager.manager.loop
                            if loop:
                                # ç¡®ä¿å…³é”®è¯åˆ—è¡¨å­˜åœ¨
                                if account_id not in cookie_manager.manager.keywords:
                                    cookie_manager.manager.keywords[account_id] = []
                                
                                # åœ¨åå°å¯åŠ¨ä»»åŠ¡ï¼ˆä½¿ç”¨çº¿ç¨‹å®‰å…¨çš„æ–¹å¼ï¼Œå› ä¸ºrun_loginæ˜¯åœ¨åå°çº¿ç¨‹ä¸­è¿è¡Œçš„ï¼‰
                                try:
                                    # å°è¯•ä½¿ç”¨run_coroutine_threadsafeï¼Œè¿™æ˜¯çº¿ç¨‹å®‰å…¨çš„æ–¹å¼
                                    fut = asyncio.run_coroutine_threadsafe(
                                        cookie_manager.manager._run_xianyu(account_id, cookies_str, user_id),
                                        loop
                                    )
                                    # ä¸ç­‰å¾…ç»“æœï¼Œè®©å®ƒåœ¨åå°è¿è¡Œ
                                    log_with_user('info', f"å·²å¯åŠ¨æ–°è´¦å·ä»»åŠ¡: {account_id}", current_user)
                                except RuntimeError as e:
                                    # å¦‚æœäº‹ä»¶å¾ªç¯æœªè¿è¡Œï¼Œè®°å½•è­¦å‘Šä½†ä¸å½±å“ç™»å½•æˆåŠŸ
                                    log_with_user('warning', f"äº‹ä»¶å¾ªç¯æœªè¿è¡Œï¼Œæ— æ³•å¯åŠ¨æ–°è´¦å·ä»»åŠ¡: {account_id}, é”™è¯¯: {str(e)}", current_user)
                                    log_with_user('info', f"è´¦å·å·²ä¿å­˜ï¼Œå°†åœ¨ç³»ç»Ÿé‡å¯åè‡ªåŠ¨å¯åŠ¨ä»»åŠ¡: {account_id}", current_user)
                        except Exception as task_err:
                            log_with_user('warning', f"å¯åŠ¨æ–°è´¦å·ä»»åŠ¡å¤±è´¥: {account_id}, é”™è¯¯: {str(task_err)}", current_user)
                            import traceback
                            logger.error(traceback.format_exc())
                
                # ç™»å½•æˆåŠŸåï¼Œè°ƒç”¨_refresh_cookies_via_browseråˆ·æ–°Cookie
                try:
                    log_with_user('info', f"å¼€å§‹è°ƒç”¨_refresh_cookies_via_browseråˆ·æ–°Cookie: {account_id}", current_user)
                    from XianyuAutoAsync import XianyuLive
                    
                    # åˆ›å»ºä¸´æ—¶çš„XianyuLiveå®ä¾‹æ¥åˆ·æ–°Cookie
                    temp_xianyu = XianyuLive(
                        cookies_str=cookies_str,
                        cookie_id=account_id,
                        user_id=user_id
                    )
                    
                    # é‡ç½®æ‰«ç ç™»å½•Cookieåˆ·æ–°æ ‡å¿—ï¼Œç¡®ä¿è´¦å·å¯†ç ç™»å½•åèƒ½ç«‹å³åˆ·æ–°
                    try:
                        temp_xianyu.reset_qr_cookie_refresh_flag()
                        log_with_user('info', f"å·²é‡ç½®æ‰«ç ç™»å½•Cookieåˆ·æ–°æ ‡å¿—: {account_id}", current_user)
                    except Exception as reset_err:
                        log_with_user('debug', f"é‡ç½®æ‰«ç ç™»å½•Cookieåˆ·æ–°æ ‡å¿—å¤±è´¥ï¼ˆä¸å½±å“åˆ·æ–°ï¼‰: {str(reset_err)}", current_user)
                    
                    # åœ¨åå°å¼‚æ­¥æ‰§è¡Œåˆ·æ–°ï¼ˆä¸é˜»å¡ä¸»æµç¨‹ï¼‰
                    async def refresh_cookies_task():
                        try:
                            refresh_success = await temp_xianyu._refresh_cookies_via_browser(triggered_by_refresh_token=False)
                            if refresh_success:
                                log_with_user('info', f"Cookieåˆ·æ–°æˆåŠŸ: {account_id}", current_user)
                                # åˆ·æ–°æˆåŠŸåï¼Œä»æ•°æ®åº“è·å–æ›´æ–°åçš„Cookie
                                updated_cookie_info = db_manager.get_cookie_details(account_id)
                                if updated_cookie_info:
                                    refreshed_cookies = updated_cookie_info.get('value', '')
                                    if refreshed_cookies:
                                        # æ›´æ–°cookie_managerä¸­çš„Cookie
                                        if cookie_manager.manager:
                                            cookie_manager.manager.update_cookie(account_id, refreshed_cookies, save_to_db=False)
                                        log_with_user('info', f"å·²æ›´æ–°åˆ·æ–°åçš„Cookieåˆ°cookie_manager: {account_id}", current_user)
                            else:
                                log_with_user('warning', f"Cookieåˆ·æ–°å¤±è´¥æˆ–è·³è¿‡: {account_id}", current_user)
                        except Exception as refresh_e:
                            log_with_user('error', f"åˆ·æ–°Cookieæ—¶å‡ºé”™: {account_id}, é”™è¯¯: {str(refresh_e)}", current_user)
                            import traceback
                            logger.error(traceback.format_exc())
                    
                    # åœ¨åå°çº¿ç¨‹ä¸­è¿è¡Œå¼‚æ­¥ä»»åŠ¡
                    # ç”±äºrun_loginæ˜¯åœ¨çº¿ç¨‹ä¸­è¿è¡Œçš„ï¼Œéœ€è¦åˆ›å»ºæ–°çš„äº‹ä»¶å¾ªç¯
                    def run_async_refresh():
                        try:
                            import asyncio
                            # åˆ›å»ºæ–°çš„äº‹ä»¶å¾ªç¯
                            new_loop = asyncio.new_event_loop()
                            asyncio.set_event_loop(new_loop)
                            try:
                                new_loop.run_until_complete(refresh_cookies_task())
                            finally:
                                new_loop.close()
                        except Exception as e:
                            log_with_user('error', f"è¿è¡Œå¼‚æ­¥åˆ·æ–°ä»»åŠ¡å¤±è´¥: {account_id}, é”™è¯¯: {str(e)}", current_user)
                    
                    # åœ¨åå°çº¿ç¨‹ä¸­æ‰§è¡Œåˆ·æ–°ä»»åŠ¡
                    refresh_thread = threading.Thread(target=run_async_refresh, daemon=True)
                    refresh_thread.start()
                    
                except Exception as refresh_err:
                    log_with_user('warning', f"è°ƒç”¨_refresh_cookies_via_browserå¤±è´¥: {account_id}, é”™è¯¯: {str(refresh_err)}", current_user)
                    # åˆ·æ–°å¤±è´¥ä¸å½±å“ç™»å½•æˆåŠŸ
                
                # æ›´æ–°ä¼šè¯çŠ¶æ€
                password_login_sessions[session_id]['status'] = 'success'
                password_login_sessions[session_id]['account_id'] = account_id
                password_login_sessions[session_id]['is_new_account'] = is_new_account
                password_login_sessions[session_id]['cookie_count'] = len(cookies_dict)
                
            except Exception as e:
                error_msg = str(e)
                password_login_sessions[session_id]['status'] = 'failed'
                password_login_sessions[session_id]['error'] = error_msg
                log_with_user('error', f"è´¦å·å¯†ç ç™»å½•å¤±è´¥: {account_id}, é”™è¯¯: {error_msg}", current_user)
                logger.info(f"ä¼šè¯ {session_id} çŠ¶æ€å·²æ›´æ–°ä¸º failedï¼Œé”™è¯¯æ¶ˆæ¯: {error_msg}")  # æ·»åŠ æ—¥å¿—ç¡®è®¤çŠ¶æ€æ›´æ–°
                import traceback
                logger.error(traceback.format_exc())
            finally:
                # æ¸…ç†å®ä¾‹ï¼ˆé‡Šæ”¾å¹¶å‘æ§½ä½ï¼‰
                try:
                    from utils.xianyu_slider_stealth import concurrency_manager
                    concurrency_manager.unregister_instance(account_id)
                    log_with_user('debug', f"å·²é‡Šæ”¾å¹¶å‘æ§½ä½: {account_id}", current_user)
                except Exception as cleanup_e:
                    log_with_user('warning', f"æ¸…ç†å®ä¾‹æ—¶å‡ºé”™: {str(cleanup_e)}", current_user)
        
        # åœ¨åå°çº¿ç¨‹ä¸­æ‰§è¡Œç™»å½•
        login_thread = threading.Thread(target=run_login, daemon=True)
        login_thread.start()
        
    except Exception as e:
        password_login_sessions[session_id]['status'] = 'failed'
        password_login_sessions[session_id]['error'] = str(e)
        log_with_user('error', f"æ‰§è¡Œè´¦å·å¯†ç ç™»å½•ä»»åŠ¡å¼‚å¸¸: {str(e)}", current_user)
        import traceback
        logger.error(traceback.format_exc())


@app.post("/password-login")
async def password_login(
    request: Dict[str, Any],
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """è´¦å·å¯†ç ç™»å½•æ¥å£ï¼ˆå¼‚æ­¥ï¼Œæ”¯æŒäººè„¸è®¤è¯ï¼‰"""
    try:
        account_id = request.get('account_id')
        account = request.get('account')
        password = request.get('password')
        show_browser = request.get('show_browser', False)
        
        if not account_id or not account or not password:
            return {'success': False, 'message': 'è´¦å·IDã€ç™»å½•è´¦å·å’Œå¯†ç ä¸èƒ½ä¸ºç©º'}
        
        log_with_user('info', f"å¼€å§‹è´¦å·å¯†ç ç™»å½•: {account_id}, è´¦å·: {account}", current_user)
        
        # ç”Ÿæˆä¼šè¯ID
        import secrets
        session_id = secrets.token_urlsafe(16)
        
        user_id = current_user['user_id']
        
        # åˆ›å»ºç™»å½•ä¼šè¯
        password_login_sessions[session_id] = {
            'account_id': account_id,
            'account': account,
            'password': password,
            'show_browser': show_browser,
            'status': 'processing',
            'verification_url': None,
            'screenshot_path': None,
            'qr_code_url': None,
            'slider_instance': None,
            'task': None,
            'timestamp': time.time(),
            'user_id': user_id
        }
        
        # å¯åŠ¨åå°ç™»å½•ä»»åŠ¡
        task = asyncio.create_task(_execute_password_login(
            session_id, account_id, account, password, show_browser, user_id, current_user
        ))
        password_login_sessions[session_id]['task'] = task
        
        return {
            'success': True,
            'session_id': session_id,
            'status': 'processing',
            'message': 'ç™»å½•ä»»åŠ¡å·²å¯åŠ¨ï¼Œè¯·ç­‰å¾…...'
        }
        
    except Exception as e:
        log_with_user('error', f"è´¦å·å¯†ç ç™»å½•å¼‚å¸¸: {str(e)}", current_user)
        import traceback
        logger.error(traceback.format_exc())
        return {'success': False, 'message': f'ç™»å½•å¤±è´¥: {str(e)}'}


@app.get("/password-login/check/{session_id}")
async def check_password_login_status(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """æ£€æŸ¥è´¦å·å¯†ç ç™»å½•çŠ¶æ€"""
    try:
        # æ¸…ç†è¿‡æœŸä¼šè¯ï¼ˆè¶…è¿‡1å°æ—¶ï¼‰
        current_time = time.time()
        expired_sessions = [
            sid for sid, session in password_login_sessions.items()
            if current_time - session['timestamp'] > 3600
        ]
        for sid in expired_sessions:
            if sid in password_login_sessions:
                del password_login_sessions[sid]
        
        if session_id not in password_login_sessions:
            return {'status': 'not_found', 'message': 'ä¼šè¯ä¸å­˜åœ¨æˆ–å·²è¿‡æœŸ'}
        
        session = password_login_sessions[session_id]
        
        # æ£€æŸ¥ç”¨æˆ·æƒé™
        if session['user_id'] != current_user['user_id']:
            return {'status': 'forbidden', 'message': 'æ— æƒé™è®¿é—®è¯¥ä¼šè¯'}
        
        status = session['status']
        
        if status == 'verification_required':
            # éœ€è¦äººè„¸è®¤è¯
            screenshot_path = session.get('screenshot_path')
            verification_url = session.get('verification_url')
            return {
                'status': 'verification_required',
                'verification_url': verification_url,
                'screenshot_path': screenshot_path,
                'qr_code_url': session.get('qr_code_url'),  # ä¿ç•™å…¼å®¹æ€§
                'message': 'éœ€è¦äººè„¸éªŒè¯ï¼Œè¯·æŸ¥çœ‹éªŒè¯æˆªå›¾' if screenshot_path else 'éœ€è¦äººè„¸éªŒè¯ï¼Œè¯·ç‚¹å‡»éªŒè¯é“¾æ¥'
            }
        elif status == 'success':
            # ç™»å½•æˆåŠŸ
            # åˆ é™¤æˆªå›¾ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
            screenshot_path = session.get('screenshot_path')
            if screenshot_path:
                try:
                    from utils.image_utils import image_manager
                    if image_manager.delete_image(screenshot_path):
                        log_with_user('info', f"éªŒè¯æˆåŠŸåå·²åˆ é™¤æˆªå›¾: {screenshot_path}", current_user)
                    else:
                        log_with_user('warning', f"åˆ é™¤æˆªå›¾å¤±è´¥: {screenshot_path}", current_user)
                except Exception as e:
                    log_with_user('error', f"åˆ é™¤æˆªå›¾æ—¶å‡ºé”™: {str(e)}", current_user)
            
            result = {
                'status': 'success',
                'message': f'è´¦å· {session["account_id"]} ç™»å½•æˆåŠŸ',
                'account_id': session['account_id'],
                'is_new_account': session.get('is_new_account', False),
                'cookie_count': session.get('cookie_count', 0)
            }
            # æ¸…ç†ä¼šè¯
            del password_login_sessions[session_id]
            return result
        elif status == 'failed':
            # ç™»å½•å¤±è´¥
            # åˆ é™¤æˆªå›¾ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
            screenshot_path = session.get('screenshot_path')
            if screenshot_path:
                try:
                    from utils.image_utils import image_manager
                    if image_manager.delete_image(screenshot_path):
                        log_with_user('info', f"éªŒè¯å¤±è´¥åå·²åˆ é™¤æˆªå›¾: {screenshot_path}", current_user)
                    else:
                        log_with_user('warning', f"åˆ é™¤æˆªå›¾å¤±è´¥: {screenshot_path}", current_user)
                except Exception as e:
                    log_with_user('error', f"åˆ é™¤æˆªå›¾æ—¶å‡ºé”™: {str(e)}", current_user)
            
            error_msg = session.get('error', 'ç™»å½•å¤±è´¥')
            log_with_user('info', f"è¿”å›ç™»å½•å¤±è´¥çŠ¶æ€: {session_id}, é”™è¯¯æ¶ˆæ¯: {error_msg}", current_user)  # æ·»åŠ æ—¥å¿—
            result = {
                'status': 'failed',
                'message': error_msg,
                'error': error_msg  # ä¹ŸåŒ…å«errorå­—æ®µï¼Œç¡®ä¿å‰ç«¯èƒ½è·å–åˆ°
            }
            # æ¸…ç†ä¼šè¯
            del password_login_sessions[session_id]
            return result
        else:
            # å¤„ç†ä¸­
            return {
                'status': 'processing',
                'message': 'ç™»å½•å¤„ç†ä¸­ï¼Œè¯·ç¨å€™...'
            }
        
    except Exception as e:
        log_with_user('error', f"æ£€æŸ¥è´¦å·å¯†ç ç™»å½•çŠ¶æ€å¼‚å¸¸: {str(e)}", current_user)
        return {'status': 'error', 'message': str(e)}


# ========================= äººè„¸éªŒè¯æˆªå›¾ç›¸å…³æ¥å£ =========================

@app.get("/face-verification/screenshot/{account_id}")
async def get_account_face_verification_screenshot(
    account_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """è·å–æŒ‡å®šè´¦å·çš„äººè„¸éªŒè¯æˆªå›¾"""
    try:
        import glob
        from datetime import datetime
        
        # æ£€æŸ¥è´¦å·æ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        username = current_user['username']
        
        # å¦‚æœæ˜¯ç®¡ç†å‘˜ï¼Œå…è®¸è®¿é—®æ‰€æœ‰è´¦å·
        is_admin = username == 'admin'
        
        if not is_admin:
            cookie_info = db_manager.get_cookie_details(account_id)
            if not cookie_info:
                log_with_user('warning', f"è´¦å· {account_id} ä¸å­˜åœ¨", current_user)
                return {
                    'success': False,
                    'message': 'è´¦å·ä¸å­˜åœ¨'
                }
            
            cookie_user_id = cookie_info.get('user_id')
            if cookie_user_id != user_id:
                log_with_user('warning', f"ç”¨æˆ· {user_id} å°è¯•è®¿é—®è´¦å· {account_id}ï¼ˆå½’å±ç”¨æˆ·: {cookie_user_id}ï¼‰", current_user)
                return {
                    'success': False,
                    'message': 'æ— æƒè®¿é—®è¯¥è´¦å·'
                }
        
        # è·å–è¯¥è´¦å·çš„éªŒè¯æˆªå›¾
        screenshots_dir = os.path.join(static_dir, 'uploads', 'images')
        pattern = os.path.join(screenshots_dir, f'face_verify_{account_id}_*.jpg')
        screenshot_files = glob.glob(pattern)
        
        log_with_user('debug', f"æŸ¥æ‰¾æˆªå›¾: {pattern}, æ‰¾åˆ° {len(screenshot_files)} ä¸ªæ–‡ä»¶", current_user)
        
        if not screenshot_files:
            log_with_user('warning', f"è´¦å· {account_id} æ²¡æœ‰æ‰¾åˆ°éªŒè¯æˆªå›¾", current_user)
            return {
                'success': False,
                'message': 'æœªæ‰¾åˆ°éªŒè¯æˆªå›¾'
            }
        
        # è·å–æœ€æ–°çš„æˆªå›¾
        latest_file = max(screenshot_files, key=os.path.getmtime)
        filename = os.path.basename(latest_file)
        stat = os.stat(latest_file)
        
        screenshot_info = {
            'filename': filename,
            'account_id': account_id,
            'path': f'/static/uploads/images/{filename}',
            'size': stat.st_size,
            'created_time': stat.st_ctime,
            'created_time_str': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
        }
        
        log_with_user('info', f"è·å–è´¦å· {account_id} çš„éªŒè¯æˆªå›¾", current_user)
        
        return {
            'success': True,
            'screenshot': screenshot_info
        }
        
    except Exception as e:
        log_with_user('error', f"è·å–éªŒè¯æˆªå›¾å¤±è´¥: {str(e)}", current_user)
        return {
            'success': False,
            'message': str(e)
        }


@app.delete("/face-verification/screenshot/{account_id}")
async def delete_account_face_verification_screenshot(
    account_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """åˆ é™¤æŒ‡å®šè´¦å·çš„äººè„¸éªŒè¯æˆªå›¾"""
    try:
        import glob
        
        # æ£€æŸ¥è´¦å·æ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        cookie_info = db_manager.get_cookie_details(account_id)
        if not cookie_info or cookie_info.get('user_id') != user_id:
            return {
                'success': False,
                'message': 'æ— æƒè®¿é—®è¯¥è´¦å·'
            }
        
        # åˆ é™¤è¯¥è´¦å·çš„æ‰€æœ‰éªŒè¯æˆªå›¾
        screenshots_dir = os.path.join(static_dir, 'uploads', 'images')
        pattern = os.path.join(screenshots_dir, f'face_verify_{account_id}_*.jpg')
        screenshot_files = glob.glob(pattern)
        
        deleted_count = 0
        for file_path in screenshot_files:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    deleted_count += 1
                    log_with_user('info', f"åˆ é™¤è´¦å· {account_id} çš„éªŒè¯æˆªå›¾: {os.path.basename(file_path)}", current_user)
            except Exception as e:
                log_with_user('error', f"åˆ é™¤æˆªå›¾å¤±è´¥ {file_path}: {str(e)}", current_user)
        
        return {
            'success': True,
            'message': f'å·²åˆ é™¤ {deleted_count} ä¸ªéªŒè¯æˆªå›¾',
            'deleted_count': deleted_count
        }
        
    except Exception as e:
        log_with_user('error', f"åˆ é™¤éªŒè¯æˆªå›¾å¤±è´¥: {str(e)}", current_user)
        return {
            'success': False,
            'message': str(e)
        }


# ========================= æ‰«ç ç™»å½•ç›¸å…³æ¥å£ =========================

@app.post("/qr-login/generate")
async def generate_qr_code(current_user: Dict[str, Any] = Depends(get_current_user)):
    """ç”Ÿæˆæ‰«ç ç™»å½•äºŒç»´ç """
    try:
        log_with_user('info', "è¯·æ±‚ç”Ÿæˆæ‰«ç ç™»å½•äºŒç»´ç ", current_user)

        result = await qr_login_manager.generate_qr_code()

        if result['success']:
            log_with_user('info', f"æ‰«ç ç™»å½•äºŒç»´ç ç”ŸæˆæˆåŠŸ: {result['session_id']}", current_user)
        else:
            log_with_user('warning', f"æ‰«ç ç™»å½•äºŒç»´ç ç”Ÿæˆå¤±è´¥: {result.get('message', 'æœªçŸ¥é”™è¯¯')}", current_user)

        return result

    except Exception as e:
        log_with_user('error', f"ç”Ÿæˆæ‰«ç ç™»å½•äºŒç»´ç å¼‚å¸¸: {str(e)}", current_user)
        return {'success': False, 'message': f'ç”ŸæˆäºŒç»´ç å¤±è´¥: {str(e)}'}


@app.get("/qr-login/check/{session_id}")
async def check_qr_code_status(session_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ£€æŸ¥æ‰«ç ç™»å½•çŠ¶æ€"""
    try:
        # æ¸…ç†è¿‡æœŸè®°å½•
        cleanup_qr_check_records()

        # æ£€æŸ¥æ˜¯å¦å·²ç»å¤„ç†è¿‡
        if session_id in qr_check_processed:
            record = qr_check_processed[session_id]
            if record['processed']:
                log_with_user('debug', f"æ‰«ç ç™»å½•session {session_id} å·²å¤„ç†è¿‡ï¼Œç›´æ¥è¿”å›", current_user)
                # è¿”å›ç®€å•çš„æˆåŠŸçŠ¶æ€ï¼Œé¿å…é‡å¤å¤„ç†
                return {'status': 'already_processed', 'message': 'è¯¥ä¼šè¯å·²å¤„ç†å®Œæˆ'}

        # è·å–è¯¥sessionçš„é”
        session_lock = qr_check_locks[session_id]

        # ä½¿ç”¨éé˜»å¡æ–¹å¼å°è¯•è·å–é”
        if session_lock.locked():
            log_with_user('debug', f"æ‰«ç ç™»å½•session {session_id} æ­£åœ¨è¢«å…¶ä»–è¯·æ±‚å¤„ç†ï¼Œè·³è¿‡", current_user)
            return {'status': 'processing', 'message': 'æ­£åœ¨å¤„ç†ä¸­ï¼Œè¯·ç¨å€™...'}

        async with session_lock:
            # å†æ¬¡æ£€æŸ¥æ˜¯å¦å·²å¤„ç†ï¼ˆåŒé‡æ£€æŸ¥ï¼‰
            if session_id in qr_check_processed and qr_check_processed[session_id]['processed']:
                log_with_user('debug', f"æ‰«ç ç™»å½•session {session_id} åœ¨è·å–é”åå‘ç°å·²å¤„ç†ï¼Œç›´æ¥è¿”å›", current_user)
                return {'status': 'already_processed', 'message': 'è¯¥ä¼šè¯å·²å¤„ç†å®Œæˆ'}

            # æ¸…ç†è¿‡æœŸä¼šè¯
            qr_login_manager.cleanup_expired_sessions()

            # è·å–ä¼šè¯çŠ¶æ€
            status_info = qr_login_manager.get_session_status(session_id)
            log_with_user('info', f"è·å–ä¼šè¯çŠ¶æ€1111111: {status_info}", current_user)
            if status_info['status'] == 'success':
                log_with_user('info', f"è·å–ä¼šè¯çŠ¶æ€22222222: {status_info}", current_user)
                # ç™»å½•æˆåŠŸï¼Œå¤„ç†Cookieï¼ˆç°åœ¨åŒ…å«è·å–çœŸå®cookieçš„é€»è¾‘ï¼‰
                cookies_info = qr_login_manager.get_session_cookies(session_id)
                log_with_user('info', f"è·å–ä¼šè¯Cookie: {cookies_info}", current_user)
                if cookies_info:
                    account_info = await process_qr_login_cookies(
                        cookies_info['cookies'],
                        cookies_info['unb'],
                        current_user
                    )
                    status_info['account_info'] = account_info

                    log_with_user('info', f"æ‰«ç ç™»å½•å¤„ç†å®Œæˆ: {session_id}, è´¦å·: {account_info.get('account_id', 'unknown')}", current_user)

                    # æ ‡è®°è¯¥sessionå·²å¤„ç†
                    qr_check_processed[session_id] = {
                        'processed': True,
                        'timestamp': time.time()
                    }

            return status_info

    except Exception as e:
        log_with_user('error', f"æ£€æŸ¥æ‰«ç ç™»å½•çŠ¶æ€å¼‚å¸¸: {str(e)}", current_user)
        return {'status': 'error', 'message': str(e)}


async def process_qr_login_cookies(cookies: str, unb: str, current_user: Dict[str, Any]) -> Dict[str, Any]:
    """å¤„ç†æ‰«ç ç™»å½•è·å–çš„Cookie - å…ˆè·å–çœŸå®cookieå†ä¿å­˜åˆ°æ•°æ®åº“"""
    try:
        user_id = current_user['user_id']

        # æ£€æŸ¥æ˜¯å¦å·²å­˜åœ¨ç›¸åŒunbçš„è´¦å·
        existing_cookies = db_manager.get_all_cookies(user_id)
        existing_account_id = None

        for account_id, cookie_value in existing_cookies.items():
            try:
                # è§£æç°æœ‰Cookieä¸­çš„unb
                existing_cookie_dict = trans_cookies(cookie_value)
                if existing_cookie_dict.get('unb') == unb:
                    existing_account_id = account_id
                    break
            except:
                continue

        # ç¡®å®šè´¦å·ID
        if existing_account_id:
            account_id = existing_account_id
            is_new_account = False
            log_with_user('info', f"æ‰«ç ç™»å½•æ‰¾åˆ°ç°æœ‰è´¦å·: {account_id}, UNB: {unb}", current_user)
        else:
            # åˆ›å»ºæ–°è´¦å·ï¼Œä½¿ç”¨unbä½œä¸ºè´¦å·ID
            account_id = unb

            # ç¡®ä¿è´¦å·IDå”¯ä¸€
            counter = 1
            original_account_id = account_id
            while account_id in existing_cookies:
                account_id = f"{original_account_id}_{counter}"
                counter += 1

            is_new_account = True
            log_with_user('info', f"æ‰«ç ç™»å½•å‡†å¤‡åˆ›å»ºæ–°è´¦å·: {account_id}, UNB: {unb}", current_user)

        # ç¬¬ä¸€æ­¥ï¼šä½¿ç”¨æ‰«ç cookieè·å–çœŸå®cookie
        log_with_user('info', f"å¼€å§‹ä½¿ç”¨æ‰«ç cookieè·å–çœŸå®cookie: {account_id}", current_user)

        try:
            # åˆ›å»ºä¸€ä¸ªä¸´æ—¶çš„XianyuLiveå®ä¾‹æ¥æ‰§è¡Œcookieåˆ·æ–°
            from XianyuAutoAsync import XianyuLive

            # ä½¿ç”¨æ‰«ç ç™»å½•çš„cookieåˆ›å»ºä¸´æ—¶å®ä¾‹
            temp_instance = XianyuLive(
                cookies_str=cookies,
                cookie_id=account_id,
                user_id=user_id
            )

            # æ‰§è¡Œcookieåˆ·æ–°è·å–çœŸå®cookie
            refresh_success = await temp_instance.refresh_cookies_from_qr_login(
                qr_cookies_str=cookies,
                cookie_id=account_id,
                user_id=user_id
            )

            if refresh_success:
                log_with_user('info', f"æ‰«ç ç™»å½•çœŸå®cookieè·å–æˆåŠŸ: {account_id}", current_user)

                # ä»æ•°æ®åº“è·å–åˆšåˆšä¿å­˜çš„çœŸå®cookie
                updated_cookie_info = db_manager.get_cookie_by_id(account_id)
                if updated_cookie_info:
                    real_cookies = updated_cookie_info['cookies_str']
                    log_with_user('info', f"å·²è·å–çœŸå®cookieï¼Œé•¿åº¦: {len(real_cookies)}", current_user)

                    # ç¬¬äºŒæ­¥ï¼šå°†çœŸå®cookieæ·»åŠ åˆ°cookie_managerï¼ˆå¦‚æœæ˜¯æ–°è´¦å·ï¼‰æˆ–æ›´æ–°ç°æœ‰è´¦å·
                    if cookie_manager.manager:
                        if is_new_account:
                            cookie_manager.manager.add_cookie(account_id, real_cookies)
                            log_with_user('info', f"å·²å°†çœŸå®cookieæ·»åŠ åˆ°cookie_manager: {account_id}", current_user)
                        else:
                            # refresh_cookies_from_qr_login å·²ç»ä¿å­˜åˆ°æ•°æ®åº“äº†ï¼Œè¿™é‡Œä¸éœ€è¦å†ä¿å­˜
                            cookie_manager.manager.update_cookie(account_id, real_cookies, save_to_db=False)
                            log_with_user('info', f"å·²æ›´æ–°cookie_managerä¸­çš„çœŸå®cookie: {account_id}", current_user)

                    return {
                        'account_id': account_id,
                        'is_new_account': is_new_account,
                        'real_cookie_refreshed': True,
                        'cookie_length': len(real_cookies)
                    }
                else:
                    log_with_user('error', f"æ— æ³•ä»æ•°æ®åº“è·å–çœŸå®cookie: {account_id}", current_user)
                    # é™çº§å¤„ç†ï¼šä½¿ç”¨åŸå§‹æ‰«ç cookie
                    return await _fallback_save_qr_cookie(account_id, cookies, user_id, is_new_account, current_user, "æ— æ³•ä»æ•°æ®åº“è·å–çœŸå®cookie")
            else:
                log_with_user('warning', f"æ‰«ç ç™»å½•çœŸå®cookieè·å–å¤±è´¥: {account_id}", current_user)
                # é™çº§å¤„ç†ï¼šä½¿ç”¨åŸå§‹æ‰«ç cookie
                return await _fallback_save_qr_cookie(account_id, cookies, user_id, is_new_account, current_user, "çœŸå®cookieè·å–å¤±è´¥")

        except Exception as refresh_e:
            log_with_user('error', f"æ‰«ç ç™»å½•çœŸå®cookieè·å–å¼‚å¸¸: {str(refresh_e)}", current_user)
            # é™çº§å¤„ç†ï¼šä½¿ç”¨åŸå§‹æ‰«ç cookie
            return await _fallback_save_qr_cookie(account_id, cookies, user_id, is_new_account, current_user, f"è·å–çœŸå®cookieå¼‚å¸¸: {str(refresh_e)}")

    except Exception as e:
        log_with_user('error', f"å¤„ç†æ‰«ç ç™»å½•Cookieå¤±è´¥: {str(e)}", current_user)
        raise e


async def _fallback_save_qr_cookie(account_id: str, cookies: str, user_id: int, is_new_account: bool, current_user: Dict[str, Any], error_reason: str) -> Dict[str, Any]:
    """é™çº§å¤„ç†ï¼šå½“æ— æ³•è·å–çœŸå®cookieæ—¶ï¼Œä¿å­˜åŸå§‹æ‰«ç cookie"""
    try:
        log_with_user('warning', f"é™çº§å¤„ç† - ä¿å­˜åŸå§‹æ‰«ç cookie: {account_id}, åŸå› : {error_reason}", current_user)

        # ä¿å­˜åŸå§‹æ‰«ç cookieåˆ°æ•°æ®åº“
        if is_new_account:
            db_manager.save_cookie(account_id, cookies, user_id)
            log_with_user('info', f"é™çº§å¤„ç† - æ–°è´¦å·åŸå§‹cookieå·²ä¿å­˜: {account_id}", current_user)
        else:
            # ç°æœ‰è´¦å·ä½¿ç”¨ update_cookie_account_info é¿å…è¦†ç›–å…¶ä»–å­—æ®µ
            db_manager.update_cookie_account_info(account_id, cookie_value=cookies)
            log_with_user('info', f"é™çº§å¤„ç† - ç°æœ‰è´¦å·åŸå§‹cookieå·²æ›´æ–°: {account_id}", current_user)

        # æ·»åŠ åˆ°æˆ–æ›´æ–°cookie_manager
        if cookie_manager.manager:
            if is_new_account:
                cookie_manager.manager.add_cookie(account_id, cookies)
                log_with_user('info', f"é™çº§å¤„ç† - å·²å°†åŸå§‹cookieæ·»åŠ åˆ°cookie_manager: {account_id}", current_user)
            else:
                # update_cookie_account_info å·²ç»ä¿å­˜åˆ°æ•°æ®åº“äº†ï¼Œè¿™é‡Œä¸éœ€è¦å†ä¿å­˜
                cookie_manager.manager.update_cookie(account_id, cookies, save_to_db=False)
                log_with_user('info', f"é™çº§å¤„ç† - å·²æ›´æ–°cookie_managerä¸­çš„åŸå§‹cookie: {account_id}", current_user)

        return {
            'account_id': account_id,
            'is_new_account': is_new_account,
            'real_cookie_refreshed': False,
            'fallback_reason': error_reason,
            'cookie_length': len(cookies)
        }

    except Exception as fallback_e:
        log_with_user('error', f"é™çº§å¤„ç†å¤±è´¥: {str(fallback_e)}", current_user)
        raise fallback_e


@app.post("/qr-login/refresh-cookies")
async def refresh_cookies_from_qr_login(
    request: Dict[str, Any],
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """ä½¿ç”¨æ‰«ç ç™»å½•è·å–çš„cookieè®¿é—®æŒ‡å®šç•Œé¢è·å–çœŸå®cookieå¹¶å­˜å…¥æ•°æ®åº“"""
    try:
        qr_cookies = request.get('qr_cookies')
        cookie_id = request.get('cookie_id')

        if not qr_cookies:
            return {'success': False, 'message': 'ç¼ºå°‘æ‰«ç ç™»å½•cookie'}

        if not cookie_id:
            return {'success': False, 'message': 'ç¼ºå°‘cookie_id'}

        log_with_user('info', f"å¼€å§‹ä½¿ç”¨æ‰«ç cookieåˆ·æ–°çœŸå®cookie: {cookie_id}", current_user)

        # åˆ›å»ºä¸€ä¸ªä¸´æ—¶çš„XianyuLiveå®ä¾‹æ¥æ‰§è¡Œcookieåˆ·æ–°
        from XianyuAutoAsync import XianyuLive

        # ä½¿ç”¨æ‰«ç ç™»å½•çš„cookieåˆ›å»ºä¸´æ—¶å®ä¾‹
        temp_instance = XianyuLive(
            cookies_str=qr_cookies,
            cookie_id=cookie_id,
            user_id=current_user['user_id']
        )

        # æ‰§è¡Œcookieåˆ·æ–°
        success = await temp_instance.refresh_cookies_from_qr_login(
            qr_cookies_str=qr_cookies,
            cookie_id=cookie_id,
            user_id=current_user['user_id']
        )

        if success:
            log_with_user('info', f"æ‰«ç cookieåˆ·æ–°æˆåŠŸ: {cookie_id}", current_user)

            # å¦‚æœcookie_managerå­˜åœ¨ï¼Œæ›´æ–°å…¶ä¸­çš„cookie
            if cookie_manager.manager:
                # ä»æ•°æ®åº“è·å–æ›´æ–°åçš„cookie
                updated_cookie_info = db_manager.get_cookie_by_id(cookie_id)
                if updated_cookie_info:
                    # refresh_cookies_from_qr_login å·²ç»ä¿å­˜åˆ°æ•°æ®åº“äº†ï¼Œè¿™é‡Œä¸éœ€è¦å†ä¿å­˜
                    cookie_manager.manager.update_cookie(cookie_id, updated_cookie_info['cookies_str'], save_to_db=False)
                    log_with_user('info', f"å·²æ›´æ–°cookie_managerä¸­çš„cookie: {cookie_id}", current_user)

            return {
                'success': True,
                'message': 'çœŸå®cookieè·å–å¹¶ä¿å­˜æˆåŠŸ',
                'cookie_id': cookie_id
            }
        else:
            log_with_user('error', f"æ‰«ç cookieåˆ·æ–°å¤±è´¥: {cookie_id}", current_user)
            return {'success': False, 'message': 'è·å–çœŸå®cookieå¤±è´¥'}

    except Exception as e:
        log_with_user('error', f"æ‰«ç cookieåˆ·æ–°å¼‚å¸¸: {str(e)}", current_user)
        return {'success': False, 'message': f'åˆ·æ–°cookieå¤±è´¥: {str(e)}'}


@app.post("/qr-login/reset-cooldown/{cookie_id}")
async def reset_qr_cookie_refresh_cooldown(
    cookie_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """é‡ç½®æŒ‡å®šè´¦å·çš„æ‰«ç ç™»å½•Cookieåˆ·æ–°å†·å´æ—¶é—´"""
    try:
        log_with_user('info', f"é‡ç½®æ‰«ç ç™»å½•Cookieåˆ·æ–°å†·å´æ—¶é—´: {cookie_id}", current_user)

        # æ£€æŸ¥cookieæ˜¯å¦å­˜åœ¨
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {'success': False, 'message': 'è´¦å·ä¸å­˜åœ¨'}

        # å¦‚æœcookie_managerä¸­æœ‰å¯¹åº”çš„å®ä¾‹ï¼Œç›´æ¥é‡ç½®
        if cookie_manager.manager and cookie_id in cookie_manager.manager.instances:
            instance = cookie_manager.manager.instances[cookie_id]
            remaining_time_before = instance.get_qr_cookie_refresh_remaining_time()
            instance.reset_qr_cookie_refresh_flag()

            log_with_user('info', f"å·²é‡ç½®è´¦å· {cookie_id} çš„æ‰«ç ç™»å½•å†·å´æ—¶é—´ï¼ŒåŸå‰©ä½™æ—¶é—´: {remaining_time_before}ç§’", current_user)

            return {
                'success': True,
                'message': 'æ‰«ç ç™»å½•Cookieåˆ·æ–°å†·å´æ—¶é—´å·²é‡ç½®',
                'cookie_id': cookie_id,
                'previous_remaining_time': remaining_time_before
            }
        else:
            # å¦‚æœæ²¡æœ‰æ´»è·ƒå®ä¾‹ï¼Œè¿”å›æˆåŠŸï¼ˆå› ä¸ºæ²¡æœ‰å†·å´æ—¶é—´éœ€è¦é‡ç½®ï¼‰
            log_with_user('info', f"è´¦å· {cookie_id} æ²¡æœ‰æ´»è·ƒå®ä¾‹ï¼Œæ— éœ€é‡ç½®å†·å´æ—¶é—´", current_user)
            return {
                'success': True,
                'message': 'è´¦å·æ²¡æœ‰æ´»è·ƒå®ä¾‹ï¼Œæ— éœ€é‡ç½®å†·å´æ—¶é—´',
                'cookie_id': cookie_id
            }

    except Exception as e:
        log_with_user('error', f"é‡ç½®æ‰«ç ç™»å½•å†·å´æ—¶é—´å¼‚å¸¸: {str(e)}", current_user)
        return {'success': False, 'message': f'é‡ç½®å†·å´æ—¶é—´å¤±è´¥: {str(e)}'}


@app.get("/qr-login/cooldown-status/{cookie_id}")
async def get_qr_cookie_refresh_cooldown_status(
    cookie_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """è·å–æŒ‡å®šè´¦å·çš„æ‰«ç ç™»å½•Cookieåˆ·æ–°å†·å´çŠ¶æ€"""
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å­˜åœ¨
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {'success': False, 'message': 'è´¦å·ä¸å­˜åœ¨'}

        # å¦‚æœcookie_managerä¸­æœ‰å¯¹åº”çš„å®ä¾‹ï¼Œè·å–å†·å´çŠ¶æ€
        if cookie_manager.manager and cookie_id in cookie_manager.manager.instances:
            instance = cookie_manager.manager.instances[cookie_id]
            remaining_time = instance.get_qr_cookie_refresh_remaining_time()
            cooldown_duration = instance.qr_cookie_refresh_cooldown
            last_refresh_time = instance.last_qr_cookie_refresh_time

            return {
                'success': True,
                'cookie_id': cookie_id,
                'remaining_time': remaining_time,
                'cooldown_duration': cooldown_duration,
                'last_refresh_time': last_refresh_time,
                'is_in_cooldown': remaining_time > 0,
                'remaining_minutes': remaining_time // 60,
                'remaining_seconds': remaining_time % 60
            }
        else:
            return {
                'success': True,
                'cookie_id': cookie_id,
                'remaining_time': 0,
                'cooldown_duration': 600,  # é»˜è®¤10åˆ†é’Ÿ
                'last_refresh_time': 0,
                'is_in_cooldown': False,
                'message': 'è´¦å·æ²¡æœ‰æ´»è·ƒå®ä¾‹'
            }

    except Exception as e:
        log_with_user('error', f"è·å–æ‰«ç ç™»å½•å†·å´çŠ¶æ€å¼‚å¸¸: {str(e)}", current_user)
        return {'success': False, 'message': f'è·å–å†·å´çŠ¶æ€å¤±è´¥: {str(e)}'}


@app.put('/cookies/{cid}/status')
def update_cookie_status(cid: str, status_data: CookieStatusIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ›´æ–°è´¦å·çš„å¯ç”¨/ç¦ç”¨çŠ¶æ€"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager æœªå°±ç»ª')
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        cookie_manager.manager.update_cookie_status(cid, status_data.enabled)
        return {'msg': 'status updated', 'enabled': status_data.enabled}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ------------------------- é»˜è®¤å›å¤ç®¡ç†æ¥å£ -------------------------

@app.get('/default-replies/{cid}')
def get_default_reply(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–æŒ‡å®šè´¦å·çš„é»˜è®¤å›å¤è®¾ç½®"""
    from db_manager import db_manager
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

        result = db_manager.get_default_reply(cid)
        if result is None:
            # å¦‚æœæ²¡æœ‰è®¾ç½®ï¼Œè¿”å›é»˜è®¤å€¼
            return {'enabled': False, 'reply_content': '', 'reply_once': False}
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/default-replies/{cid}')
def update_default_reply(cid: str, reply_data: DefaultReplyIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ›´æ–°æŒ‡å®šè´¦å·çš„é»˜è®¤å›å¤è®¾ç½®"""
    from db_manager import db_manager
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        db_manager.save_default_reply(cid, reply_data.enabled, reply_data.reply_content, reply_data.reply_once)
        return {'msg': 'default reply updated', 'enabled': reply_data.enabled, 'reply_once': reply_data.reply_once}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/default-replies')
def get_all_default_replies(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å½“å‰ç”¨æˆ·æ‰€æœ‰è´¦å·çš„é»˜è®¤å›å¤è®¾ç½®"""
    from db_manager import db_manager
    try:
        # åªè¿”å›å½“å‰ç”¨æˆ·çš„é»˜è®¤å›å¤è®¾ç½®
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_replies = db_manager.get_all_default_replies()
        # è¿‡æ»¤åªå±äºå½“å‰ç”¨æˆ·çš„å›å¤è®¾ç½®
        user_replies = {cid: reply for cid, reply in all_replies.items() if cid in user_cookies}
        return user_replies
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/default-replies/{cid}')
def delete_default_reply(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """åˆ é™¤æŒ‡å®šè´¦å·çš„é»˜è®¤å›å¤è®¾ç½®"""
    from db_manager import db_manager
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        success = db_manager.delete_default_reply(cid)
        if success:
            return {'msg': 'default reply deleted'}
        else:
            raise HTTPException(status_code=400, detail='åˆ é™¤å¤±è´¥')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/default-replies/{cid}/clear-records')
def clear_default_reply_records(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ¸…ç©ºæŒ‡å®šè´¦å·çš„é»˜è®¤å›å¤è®°å½•"""
    from db_manager import db_manager
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        db_manager.clear_default_reply_records(cid)
        return {'msg': 'default reply records cleared'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- é€šçŸ¥æ¸ é“ç®¡ç†æ¥å£ -------------------------

@app.get('/notification-channels')
def get_notification_channels(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–æ‰€æœ‰é€šçŸ¥æ¸ é“"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        return db_manager.get_notification_channels(user_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/notification-channels')
def create_notification_channel(channel_data: NotificationChannelIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """åˆ›å»ºé€šçŸ¥æ¸ é“"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        channel_id = db_manager.create_notification_channel(
            channel_data.name,
            channel_data.type,
            channel_data.config,
            user_id
        )
        return {'msg': 'notification channel created', 'id': channel_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get('/notification-channels/{channel_id}')
def get_notification_channel(channel_id: int, _: None = Depends(require_auth)):
    """è·å–æŒ‡å®šé€šçŸ¥æ¸ é“"""
    from db_manager import db_manager
    try:
        channel = db_manager.get_notification_channel(channel_id)
        if not channel:
            raise HTTPException(status_code=404, detail='é€šçŸ¥æ¸ é“ä¸å­˜åœ¨')
        return channel
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/notification-channels/{channel_id}')
def update_notification_channel(channel_id: int, channel_data: NotificationChannelUpdate, _: None = Depends(require_auth)):
    """æ›´æ–°é€šçŸ¥æ¸ é“"""
    from db_manager import db_manager
    try:
        success = db_manager.update_notification_channel(
            channel_id,
            channel_data.name,
            channel_data.config,
            channel_data.enabled
        )
        if success:
            return {'msg': 'notification channel updated'}
        else:
            raise HTTPException(status_code=404, detail='é€šçŸ¥æ¸ é“ä¸å­˜åœ¨')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.delete('/notification-channels/{channel_id}')
def delete_notification_channel(channel_id: int, _: None = Depends(require_auth)):
    """åˆ é™¤é€šçŸ¥æ¸ é“"""
    from db_manager import db_manager
    try:
        success = db_manager.delete_notification_channel(channel_id)
        if success:
            return {'msg': 'notification channel deleted'}
        else:
            raise HTTPException(status_code=404, detail='é€šçŸ¥æ¸ é“ä¸å­˜åœ¨')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- æ¶ˆæ¯é€šçŸ¥é…ç½®æ¥å£ -------------------------

@app.get('/message-notifications')
def get_all_message_notifications(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å½“å‰ç”¨æˆ·æ‰€æœ‰è´¦å·çš„æ¶ˆæ¯é€šçŸ¥é…ç½®"""
    from db_manager import db_manager
    try:
        # åªè¿”å›å½“å‰ç”¨æˆ·çš„æ¶ˆæ¯é€šçŸ¥é…ç½®
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_notifications = db_manager.get_all_message_notifications()
        # è¿‡æ»¤åªå±äºå½“å‰ç”¨æˆ·çš„é€šçŸ¥é…ç½®
        user_notifications = {cid: notifications for cid, notifications in all_notifications.items() if cid in user_cookies}
        return user_notifications
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/message-notifications/{cid}')
def get_account_notifications(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–æŒ‡å®šè´¦å·çš„æ¶ˆæ¯é€šçŸ¥é…ç½®"""
    from db_manager import db_manager
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

        return db_manager.get_account_notifications(cid)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/message-notifications/{cid}')
def set_message_notification(cid: str, notification_data: MessageNotificationIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è®¾ç½®è´¦å·çš„æ¶ˆæ¯é€šçŸ¥"""
    from db_manager import db_manager
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # æ£€æŸ¥é€šçŸ¥æ¸ é“æ˜¯å¦å­˜åœ¨
        channel = db_manager.get_notification_channel(notification_data.channel_id)
        if not channel:
            raise HTTPException(status_code=404, detail='é€šçŸ¥æ¸ é“ä¸å­˜åœ¨')

        success = db_manager.set_message_notification(cid, notification_data.channel_id, notification_data.enabled)
        if success:
            return {'msg': 'message notification set'}
        else:
            raise HTTPException(status_code=400, detail='è®¾ç½®å¤±è´¥')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/message-notifications/account/{cid}')
def delete_account_notifications(cid: str, _: None = Depends(require_auth)):
    """åˆ é™¤è´¦å·çš„æ‰€æœ‰æ¶ˆæ¯é€šçŸ¥é…ç½®"""
    from db_manager import db_manager
    try:
        success = db_manager.delete_account_notifications(cid)
        if success:
            return {'msg': 'account notifications deleted'}
        else:
            raise HTTPException(status_code=404, detail='è´¦å·é€šçŸ¥é…ç½®ä¸å­˜åœ¨')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/message-notifications/{notification_id}')
def delete_message_notification(notification_id: int, _: None = Depends(require_auth)):
    """åˆ é™¤æ¶ˆæ¯é€šçŸ¥é…ç½®"""
    from db_manager import db_manager
    try:
        success = db_manager.delete_message_notification(notification_id)
        if success:
            return {'msg': 'message notification deleted'}
        else:
            raise HTTPException(status_code=404, detail='é€šçŸ¥é…ç½®ä¸å­˜åœ¨')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- ç³»ç»Ÿè®¾ç½®æ¥å£ -------------------------

@app.get('/system-settings')
def get_system_settings(_: None = Depends(require_auth)):
    """è·å–ç³»ç»Ÿè®¾ç½®ï¼ˆæ’é™¤æ•æ„Ÿä¿¡æ¯ï¼‰"""
    from db_manager import db_manager
    try:
        settings = db_manager.get_all_system_settings()
        # ç§»é™¤æ•æ„Ÿä¿¡æ¯
        if 'admin_password_hash' in settings:
            del settings['admin_password_hash']
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





@app.put('/system-settings/{key}')
def update_system_setting(key: str, setting_data: SystemSettingIn, _: None = Depends(require_auth)):
    """æ›´æ–°ç³»ç»Ÿè®¾ç½®"""
    from db_manager import db_manager
    try:
        # ç¦æ­¢ç›´æ¥ä¿®æ”¹å¯†ç å“ˆå¸Œ
        if key == 'admin_password_hash':
            raise HTTPException(status_code=400, detail='è¯·ä½¿ç”¨å¯†ç ä¿®æ”¹æ¥å£')

        success = db_manager.set_system_setting(key, setting_data.value, setting_data.description)
        if success:
            return {'msg': 'system setting updated'}
        else:
            raise HTTPException(status_code=400, detail='æ›´æ–°å¤±è´¥')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- æ³¨å†Œè®¾ç½®æ¥å£ -------------------------

@app.get('/registration-status')
def get_registration_status():
    """è·å–æ³¨å†Œå¼€å…³çŠ¶æ€ï¼ˆå…¬å¼€æ¥å£ï¼Œæ— éœ€è®¤è¯ï¼‰"""
    from db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('registration_enabled')
        logger.info(f"ä»æ•°æ®åº“è·å–çš„æ³¨å†Œè®¾ç½®å€¼: '{enabled_str}'")  # è°ƒè¯•ä¿¡æ¯

        # å¦‚æœè®¾ç½®ä¸å­˜åœ¨ï¼Œé»˜è®¤ä¸ºå¼€å¯
        if enabled_str is None:
            enabled_bool = True
            message = 'æ³¨å†ŒåŠŸèƒ½å·²å¼€å¯'
        else:
            enabled_bool = enabled_str == 'true'
            message = 'æ³¨å†ŒåŠŸèƒ½å·²å¼€å¯' if enabled_bool else 'æ³¨å†ŒåŠŸèƒ½å·²å…³é—­'

        logger.info(f"è§£æåçš„æ³¨å†ŒçŠ¶æ€: enabled={enabled_bool}, message='{message}'")  # è°ƒè¯•ä¿¡æ¯

        return {
            'enabled': enabled_bool,
            'message': message
        }
    except Exception as e:
        logger.error(f"è·å–æ³¨å†ŒçŠ¶æ€å¤±è´¥: {e}")
        return {'enabled': True, 'message': 'æ³¨å†ŒåŠŸèƒ½å·²å¼€å¯'}  # å‡ºé”™æ—¶é»˜è®¤å¼€å¯


@app.get('/login-info-status')
def get_login_info_status():
    """è·å–é»˜è®¤ç™»å½•ä¿¡æ¯æ˜¾ç¤ºçŠ¶æ€ï¼ˆå…¬å¼€æ¥å£ï¼Œæ— éœ€è®¤è¯ï¼‰"""
    from db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('show_default_login_info')
        logger.debug(f"ä»æ•°æ®åº“è·å–çš„ç™»å½•ä¿¡æ¯æ˜¾ç¤ºè®¾ç½®å€¼: '{enabled_str}'")

        # å¦‚æœè®¾ç½®ä¸å­˜åœ¨ï¼Œé»˜è®¤ä¸ºå¼€å¯
        if enabled_str is None:
            enabled_bool = True
        else:
            enabled_bool = enabled_str == 'true'

        return {"enabled": enabled_bool}
    except Exception as e:
        logger.error(f"è·å–ç™»å½•ä¿¡æ¯æ˜¾ç¤ºçŠ¶æ€å¤±è´¥: {e}")
        # å‡ºé”™æ—¶é»˜è®¤ä¸ºå¼€å¯
        return {"enabled": True}


class RegistrationSettingUpdate(BaseModel):
    enabled: bool


class LoginInfoSettingUpdate(BaseModel):
    enabled: bool


@app.put('/registration-settings')
def update_registration_settings(setting_data: RegistrationSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """æ›´æ–°æ³¨å†Œå¼€å…³è®¾ç½®ï¼ˆä»…ç®¡ç†å‘˜ï¼‰"""
    from db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'registration_enabled',
            'true' if enabled else 'false',
            'æ˜¯å¦å¼€å¯ç”¨æˆ·æ³¨å†Œ'
        )
        if success:
            log_with_user('info', f"æ›´æ–°æ³¨å†Œè®¾ç½®: {'å¼€å¯' if enabled else 'å…³é—­'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"æ³¨å†ŒåŠŸèƒ½å·²{'å¼€å¯' if enabled else 'å…³é—­'}"
            }
        else:
            raise HTTPException(status_code=500, detail='æ›´æ–°æ³¨å†Œè®¾ç½®å¤±è´¥')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"æ›´æ–°æ³¨å†Œè®¾ç½®å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/login-info-settings')
def update_login_info_settings(setting_data: LoginInfoSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """æ›´æ–°é»˜è®¤ç™»å½•ä¿¡æ¯æ˜¾ç¤ºè®¾ç½®ï¼ˆä»…ç®¡ç†å‘˜ï¼‰"""
    from db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'show_default_login_info',
            'true' if enabled else 'false',
            'æ˜¯å¦æ˜¾ç¤ºé»˜è®¤ç™»å½•ä¿¡æ¯'
        )
        if success:
            log_with_user('info', f"æ›´æ–°ç™»å½•ä¿¡æ¯æ˜¾ç¤ºè®¾ç½®: {'å¼€å¯' if enabled else 'å…³é—­'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"é»˜è®¤ç™»å½•ä¿¡æ¯æ˜¾ç¤ºå·²{'å¼€å¯' if enabled else 'å…³é—­'}"
            }
        else:
            raise HTTPException(status_code=500, detail='æ›´æ–°ç™»å½•ä¿¡æ¯æ˜¾ç¤ºè®¾ç½®å¤±è´¥')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"æ›´æ–°ç™»å½•ä¿¡æ¯æ˜¾ç¤ºè®¾ç½®å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=str(e))




@app.delete("/cookies/{cid}")
def remove_cookie(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        cookie_manager.manager.remove_cookie(cid)
        return {"msg": "removed"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class AutoConfirmUpdate(BaseModel):
    auto_confirm: bool


class RemarkUpdate(BaseModel):
    remark: str


class PauseDurationUpdate(BaseModel):
    pause_duration: int


@app.put("/cookies/{cid}/auto-confirm")
def update_auto_confirm(cid: str, update_data: AutoConfirmUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ›´æ–°è´¦å·çš„è‡ªåŠ¨ç¡®è®¤å‘è´§è®¾ç½®"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # æ›´æ–°æ•°æ®åº“ä¸­çš„auto_confirmè®¾ç½®
        success = db_manager.update_auto_confirm(cid, update_data.auto_confirm)
        if not success:
            raise HTTPException(status_code=500, detail="æ›´æ–°è‡ªåŠ¨ç¡®è®¤å‘è´§è®¾ç½®å¤±è´¥")

        # é€šçŸ¥CookieManageræ›´æ–°è®¾ç½®ï¼ˆå¦‚æœè´¦å·æ­£åœ¨è¿è¡Œï¼‰
        if hasattr(cookie_manager.manager, 'update_auto_confirm_setting'):
            cookie_manager.manager.update_auto_confirm_setting(cid, update_data.auto_confirm)

        return {
            "msg": "success",
            "auto_confirm": update_data.auto_confirm,
            "message": f"è‡ªåŠ¨ç¡®è®¤å‘è´§å·²{'å¼€å¯' if update_data.auto_confirm else 'å…³é—­'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/auto-confirm")
def get_auto_confirm(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–è´¦å·çš„è‡ªåŠ¨ç¡®è®¤å‘è´§è®¾ç½®"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # è·å–auto_confirmè®¾ç½®
        auto_confirm = db_manager.get_auto_confirm(cid)
        return {
            "auto_confirm": auto_confirm,
            "message": f"è‡ªåŠ¨ç¡®è®¤å‘è´§å½“å‰{'å¼€å¯' if auto_confirm else 'å…³é—­'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cookies/{cid}/remark")
def update_cookie_remark(cid: str, update_data: RemarkUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ›´æ–°è´¦å·å¤‡æ³¨"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # æ›´æ–°å¤‡æ³¨
        success = db_manager.update_cookie_remark(cid, update_data.remark)
        if success:
            log_with_user('info', f"æ›´æ–°è´¦å·å¤‡æ³¨: {cid} -> {update_data.remark}", current_user)
            return {
                "message": "å¤‡æ³¨æ›´æ–°æˆåŠŸ",
                "remark": update_data.remark
            }
        else:
            raise HTTPException(status_code=500, detail="å¤‡æ³¨æ›´æ–°å¤±è´¥")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/remark")
def get_cookie_remark(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–è´¦å·å¤‡æ³¨"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # è·å–Cookieè¯¦ç»†ä¿¡æ¯ï¼ˆåŒ…å«å¤‡æ³¨ï¼‰
        cookie_details = db_manager.get_cookie_details(cid)
        if cookie_details:
            return {
                "remark": cookie_details.get('remark', ''),
                "message": "è·å–å¤‡æ³¨æˆåŠŸ"
            }
        else:
            raise HTTPException(status_code=404, detail="è´¦å·ä¸å­˜åœ¨")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cookies/{cid}/pause-duration")
def update_cookie_pause_duration(cid: str, update_data: PauseDurationUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ›´æ–°è´¦å·è‡ªåŠ¨å›å¤æš‚åœæ—¶é—´"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # éªŒè¯æš‚åœæ—¶é—´èŒƒå›´ï¼ˆ0-60åˆ†é’Ÿï¼Œ0è¡¨ç¤ºä¸æš‚åœï¼‰
        if not (0 <= update_data.pause_duration <= 60):
            raise HTTPException(status_code=400, detail="æš‚åœæ—¶é—´å¿…é¡»åœ¨0-60åˆ†é’Ÿä¹‹é—´ï¼ˆ0è¡¨ç¤ºä¸æš‚åœï¼‰")

        # æ›´æ–°æš‚åœæ—¶é—´
        success = db_manager.update_cookie_pause_duration(cid, update_data.pause_duration)
        if success:
            log_with_user('info', f"æ›´æ–°è´¦å·è‡ªåŠ¨å›å¤æš‚åœæ—¶é—´: {cid} -> {update_data.pause_duration}åˆ†é’Ÿ", current_user)
            return {
                "message": "æš‚åœæ—¶é—´æ›´æ–°æˆåŠŸ",
                "pause_duration": update_data.pause_duration
            }
        else:
            raise HTTPException(status_code=500, detail="æš‚åœæ—¶é—´æ›´æ–°å¤±è´¥")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/pause-duration")
def get_cookie_pause_duration(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–è´¦å·è‡ªåŠ¨å›å¤æš‚åœæ—¶é—´"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # è·å–æš‚åœæ—¶é—´
        pause_duration = db_manager.get_cookie_pause_duration(cid)
        return {
            "pause_duration": pause_duration,
            "message": "è·å–æš‚åœæ—¶é—´æˆåŠŸ"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




class KeywordIn(BaseModel):
    keywords: Dict[str, str]  # key -> reply

class KeywordWithItemIdIn(BaseModel):
    keywords: List[Dict[str, Any]]  # [{"keyword": str, "reply": str, "item_id": str}]


@app.get("/keywords/{cid}")
def get_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")

    # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

    # ç›´æ¥ä»æ•°æ®åº“è·å–æ‰€æœ‰å…³é”®è¯ï¼ˆé¿å…é‡å¤è®¡ç®—ï¼‰
    item_keywords = db_manager.get_keywords_with_item_id(cid)

    # è½¬æ¢ä¸ºç»Ÿä¸€æ ¼å¼
    all_keywords = []
    for keyword, reply, item_id in item_keywords:
        all_keywords.append({
            "keyword": keyword,
            "reply": reply,
            "item_id": item_id,
            "type": "item" if item_id else "normal"
        })

    return all_keywords


@app.get("/keywords-with-item-id/{cid}")
def get_keywords_with_item_id(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–åŒ…å«å•†å“IDçš„å…³é”®è¯åˆ—è¡¨"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")

    # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

    # è·å–åŒ…å«ç±»å‹ä¿¡æ¯çš„å…³é”®è¯
    keywords = db_manager.get_keywords_with_type(cid)

    # è½¬æ¢ä¸ºå‰ç«¯éœ€è¦çš„æ ¼å¼
    result = []
    for keyword_data in keywords:
        result.append({
            "keyword": keyword_data['keyword'],
            "reply": keyword_data['reply'],
            "item_id": keyword_data['item_id'] or "",
            "type": keyword_data['type'],
            "image_url": keyword_data['image_url']
        })

    return result


@app.post("/keywords/{cid}")
def update_keywords(cid: str, body: KeywordIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")

    # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"å°è¯•æ“ä½œå…¶ä»–ç”¨æˆ·çš„Cookieå…³é”®å­—: {cid}", current_user)
        raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

    kw_list = [(k, v) for k, v in body.keywords.items()]
    log_with_user('info', f"æ›´æ–°Cookieå…³é”®å­—: {cid}, æ•°é‡: {len(kw_list)}", current_user)

    cookie_manager.manager.update_keywords(cid, kw_list)
    log_with_user('info', f"Cookieå…³é”®å­—æ›´æ–°æˆåŠŸ: {cid}", current_user)
    return {"msg": "updated", "count": len(kw_list)}


@app.post("/keywords-with-item-id/{cid}")
def update_keywords_with_item_id(cid: str, body: KeywordWithItemIdIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ›´æ–°åŒ…å«å•†å“IDçš„å…³é”®è¯åˆ—è¡¨"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")

    # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"å°è¯•æ“ä½œå…¶ä»–ç”¨æˆ·çš„Cookieå…³é”®å­—: {cid}", current_user)
        raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

    # éªŒè¯æ•°æ®æ ¼å¼
    keywords_to_save = []
    keyword_set = set()  # ç”¨äºæ£€æŸ¥å½“å‰æäº¤çš„å…³é”®è¯ä¸­æ˜¯å¦æœ‰é‡å¤

    for kw_data in body.keywords:
        keyword = kw_data.get('keyword', '').strip()
        reply = kw_data.get('reply', '').strip()
        item_id = kw_data.get('item_id', '').strip() or None

        if not keyword:
            raise HTTPException(status_code=400, detail="å…³é”®è¯ä¸èƒ½ä¸ºç©º")

        # æ£€æŸ¥å½“å‰æäº¤çš„å…³é”®è¯ä¸­æ˜¯å¦æœ‰é‡å¤
        keyword_key = f"{keyword}|{item_id or ''}"
        if keyword_key in keyword_set:
            item_id_text = f"ï¼ˆå•†å“ID: {item_id}ï¼‰" if item_id else "ï¼ˆé€šç”¨å…³é”®è¯ï¼‰"
            raise HTTPException(status_code=400, detail=f"å…³é”®è¯ '{keyword}' {item_id_text} åœ¨å½“å‰æäº¤ä¸­é‡å¤")
        keyword_set.add(keyword_key)

        keywords_to_save.append((keyword, reply, item_id))

    # ä¿å­˜å…³é”®è¯ï¼ˆåªä¿å­˜æ–‡æœ¬å…³é”®è¯ï¼Œä¿ç•™å›¾ç‰‡å…³é”®è¯ï¼‰
    try:
        success = db_manager.save_text_keywords_only(cid, keywords_to_save)
        if not success:
            raise HTTPException(status_code=500, detail="ä¿å­˜å…³é”®è¯å¤±è´¥")
    except Exception as e:
        error_msg = str(e)

        # æ£€æŸ¥æ˜¯å¦æ˜¯å›¾ç‰‡å…³é”®è¯å†²çª
        if "å·²å­˜åœ¨ï¼ˆå›¾ç‰‡å…³é”®è¯ï¼‰" in error_msg:
            # ç›´æ¥ä½¿ç”¨æ•°æ®åº“ç®¡ç†å™¨æä¾›çš„å‹å¥½é”™è¯¯ä¿¡æ¯
            raise HTTPException(status_code=400, detail=error_msg)
        elif "UNIQUE constraint failed" in error_msg or "å”¯ä¸€çº¦æŸå†²çª" in error_msg:
            # å°è¯•ä»é”™è¯¯ä¿¡æ¯ä¸­æå–å…·ä½“çš„å†²çªå…³é”®è¯
            conflict_keyword = None
            conflict_type = None

            # æ£€æŸ¥æ˜¯å¦æ˜¯æ•°æ®åº“ç®¡ç†å™¨æŠ›å‡ºçš„è¯¦ç»†é”™è¯¯
            if "å…³é”®è¯å”¯ä¸€çº¦æŸå†²çª" in error_msg:
                # è§£æè¯¦ç»†é”™è¯¯ä¿¡æ¯ï¼šå…³é”®è¯å”¯ä¸€çº¦æŸå†²çª: Cookie=xxx, å…³é”®è¯='xxx', é€šç”¨å…³é”®è¯/å•†å“ID: xxx
                import re
                keyword_match = re.search(r"å…³é”®è¯='([^']+)'", error_msg)
                if keyword_match:
                    conflict_keyword = keyword_match.group(1)

                if "é€šç”¨å…³é”®è¯" in error_msg:
                    conflict_type = "é€šç”¨å…³é”®è¯"
                elif "å•†å“ID:" in error_msg:
                    item_match = re.search(r"å•†å“ID: ([^\s,]+)", error_msg)
                    if item_match:
                        conflict_type = f"å•†å“å…³é”®è¯ï¼ˆå•†å“ID: {item_match.group(1)}ï¼‰"

            # æ„é€ ç”¨æˆ·å‹å¥½çš„é”™è¯¯ä¿¡æ¯
            if conflict_keyword and conflict_type:
                detail_msg = f'å…³é”®è¯ "{conflict_keyword}" ï¼ˆ{conflict_type}ï¼‰ å·²å­˜åœ¨ï¼Œè¯·ä½¿ç”¨å…¶ä»–å…³é”®è¯æˆ–å•†å“ID'
            elif "keywords.cookie_id, keywords.keyword" in error_msg:
                detail_msg = "å…³é”®è¯é‡å¤ï¼è¯¥å…³é”®è¯å·²å­˜åœ¨ï¼ˆå¯èƒ½æ˜¯å›¾ç‰‡å…³é”®è¯æˆ–æ–‡æœ¬å…³é”®è¯ï¼‰ï¼Œè¯·ä½¿ç”¨å…¶ä»–å…³é”®è¯"
            else:
                detail_msg = "å…³é”®è¯é‡å¤ï¼è¯·ä½¿ç”¨ä¸åŒçš„å…³é”®è¯æˆ–å•†å“IDç»„åˆ"

            raise HTTPException(status_code=400, detail=detail_msg)
        else:
            log_with_user('error', f"ä¿å­˜å…³é”®è¯æ—¶å‘ç”ŸæœªçŸ¥é”™è¯¯: {error_msg}", current_user)
            raise HTTPException(status_code=500, detail="ä¿å­˜å…³é”®è¯å¤±è´¥")

    log_with_user('info', f"æ›´æ–°Cookieå…³é”®å­—(å«å•†å“ID): {cid}, æ•°é‡: {len(keywords_to_save)}", current_user)
    return {"msg": "updated", "count": len(keywords_to_save)}


@app.get("/items/{cid}")
def get_items_list(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–æŒ‡å®šè´¦å·çš„å•†å“åˆ—è¡¨"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")

    # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

    try:
        # è·å–è¯¥è´¦å·çš„æ‰€æœ‰å•†å“
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute('''
            SELECT item_id, item_title, item_price, created_at
            FROM item_info
            WHERE cookie_id = ?
            ORDER BY created_at DESC
            ''', (cid,))

            items = []
            for row in cursor.fetchall():
                items.append({
                    'item_id': row[0],
                    'item_title': row[1] or 'æœªçŸ¥å•†å“',
                    'item_price': row[2] or 'ä»·æ ¼æœªçŸ¥',
                    'created_at': row[3]
                })

            return {"items": items, "count": len(items)}

    except Exception as e:
        logger.error(f"è·å–å•†å“åˆ—è¡¨å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail="è·å–å•†å“åˆ—è¡¨å¤±è´¥")


@app.get("/keywords-export/{cid}")
def export_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """å¯¼å‡ºæŒ‡å®šè´¦å·çš„å…³é”®è¯ä¸ºExcelæ–‡ä»¶"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")

    # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

    try:
        # è·å–å…³é”®è¯æ•°æ®ï¼ˆåŒ…å«ç±»å‹ä¿¡æ¯ï¼‰
        keywords = db_manager.get_keywords_with_type(cid)

        # åˆ›å»ºDataFrameï¼Œåªå¯¼å‡ºæ–‡æœ¬ç±»å‹çš„å…³é”®è¯
        data = []
        for keyword_data in keywords:
            # åªå¯¼å‡ºæ–‡æœ¬ç±»å‹çš„å…³é”®è¯
            if keyword_data.get('type', 'text') == 'text':
                data.append({
                    'å…³é”®è¯': keyword_data['keyword'],
                    'å•†å“ID': keyword_data['item_id'] or '',
                    'å…³é”®è¯å†…å®¹': keyword_data['reply']
                })

        # å¦‚æœæ²¡æœ‰æ•°æ®ï¼Œåˆ›å»ºç©ºçš„DataFrameä½†ä¿ç•™åˆ—åï¼ˆä½œä¸ºæ¨¡æ¿ï¼‰
        if not data:
            df = pd.DataFrame(columns=['å…³é”®è¯', 'å•†å“ID', 'å…³é”®è¯å†…å®¹'])
        else:
            df = pd.DataFrame(data)

        # åˆ›å»ºExcelæ–‡ä»¶
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='å…³é”®è¯æ•°æ®', index=False)

            # å¦‚æœæ˜¯ç©ºæ¨¡æ¿ï¼Œæ·»åŠ ä¸€äº›ç¤ºä¾‹è¯´æ˜
            if data == []:
                worksheet = writer.sheets['å…³é”®è¯æ•°æ®']
                # æ·»åŠ ç¤ºä¾‹æ•°æ®ä½œä¸ºæ³¨é‡Šï¼ˆä»ç¬¬2è¡Œå¼€å§‹ï¼‰
                worksheet['A2'] = 'ä½ å¥½'
                worksheet['B2'] = ''
                worksheet['C2'] = 'æ‚¨å¥½ï¼æ¬¢è¿å’¨è¯¢ï¼Œæœ‰ä»€ä¹ˆå¯ä»¥å¸®åŠ©æ‚¨çš„å—ï¼Ÿ'

                worksheet['A3'] = 'ä»·æ ¼'
                worksheet['B3'] = '123456'
                worksheet['C3'] = 'è¿™ä¸ªå•†å“çš„ä»·æ ¼æ˜¯99å…ƒï¼Œç°åœ¨æœ‰ä¼˜æƒ æ´»åŠ¨å“¦ï¼'

                worksheet['A4'] = 'å‘è´§'
                worksheet['B4'] = ''
                worksheet['C4'] = 'æˆ‘ä»¬ä¼šåœ¨24å°æ—¶å†…å‘è´§ï¼Œè¯·è€å¿ƒç­‰å¾…ã€‚'

                # è®¾ç½®ç¤ºä¾‹è¡Œçš„æ ·å¼ï¼ˆæµ…ç°è‰²èƒŒæ™¯ï¼‰
                from openpyxl.styles import PatternFill
                gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                for row in range(2, 5):
                    for col in range(1, 4):
                        worksheet.cell(row=row, column=col).fill = gray_fill

        output.seek(0)

        # ç”Ÿæˆæ–‡ä»¶åï¼ˆä½¿ç”¨URLç¼–ç å¤„ç†ä¸­æ–‡ï¼‰
        from urllib.parse import quote
        if not data:
            filename = f"keywords_template_{cid}_{int(time.time())}.xlsx"
        else:
            filename = f"keywords_{cid}_{int(time.time())}.xlsx"
        encoded_filename = quote(filename.encode('utf-8'))

        # è¿”å›æ–‡ä»¶
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        logger.error(f"å¯¼å‡ºå…³é”®è¯å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=f"å¯¼å‡ºå…³é”®è¯å¤±è´¥: {str(e)}")


@app.post("/keywords-import/{cid}")
async def import_keywords(cid: str, file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """å¯¼å…¥Excelæ–‡ä»¶ä¸­çš„å…³é”®è¯åˆ°æŒ‡å®šè´¦å·"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")

    # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

    # æ£€æŸ¥æ–‡ä»¶ç±»å‹
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="è¯·ä¸Šä¼ Excelæ–‡ä»¶(.xlsxæˆ–.xls)")

    try:
        # è¯»å–Excelæ–‡ä»¶
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))

        # æ£€æŸ¥å¿…è¦çš„åˆ—
        required_columns = ['å…³é”®è¯', 'å•†å“ID', 'å…³é”®è¯å†…å®¹']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Excelæ–‡ä»¶ç¼ºå°‘å¿…è¦çš„åˆ—: {', '.join(missing_columns)}")

        # è·å–ç°æœ‰çš„æ–‡æœ¬ç±»å‹å…³é”®è¯ï¼ˆç”¨äºæ¯”è¾ƒæ›´æ–°/æ–°å¢ï¼‰
        existing_keywords = db_manager.get_keywords_with_type(cid)
        existing_dict = {}
        for keyword_data in existing_keywords:
            # åªè€ƒè™‘æ–‡æœ¬ç±»å‹çš„å…³é”®è¯
            if keyword_data.get('type', 'text') == 'text':
                keyword = keyword_data['keyword']
                reply = keyword_data['reply']
                item_id = keyword_data['item_id']
                key = f"{keyword}|{item_id or ''}"
                existing_dict[key] = (keyword, reply, item_id)

        # å¤„ç†å¯¼å…¥æ•°æ®
        import_data = []
        update_count = 0
        add_count = 0

        for index, row in df.iterrows():
            keyword = str(row['å…³é”®è¯']).strip()
            item_id = str(row['å•†å“ID']).strip() if pd.notna(row['å•†å“ID']) and str(row['å•†å“ID']).strip() else None
            reply = str(row['å…³é”®è¯å†…å®¹']).strip()

            if not keyword:
                continue  # è·³è¿‡æ²¡æœ‰å…³é”®è¯çš„è¡Œ

            # æ£€æŸ¥æ˜¯å¦é‡å¤
            key = f"{keyword}|{item_id or ''}"
            if key in existing_dict:
                # æ›´æ–°ç°æœ‰å…³é”®è¯
                update_count += 1
            else:
                # æ–°å¢å…³é”®è¯
                add_count += 1

            import_data.append((keyword, reply, item_id))

        if not import_data:
            raise HTTPException(status_code=400, detail="Excelæ–‡ä»¶ä¸­æ²¡æœ‰æœ‰æ•ˆçš„å…³é”®è¯æ•°æ®")

        # ä¿å­˜åˆ°æ•°æ®åº“ï¼ˆåªå½±å“æ–‡æœ¬å…³é”®è¯ï¼Œä¿ç•™å›¾ç‰‡å…³é”®è¯ï¼‰
        success = db_manager.save_text_keywords_only(cid, import_data)
        if not success:
            raise HTTPException(status_code=500, detail="ä¿å­˜å…³é”®è¯åˆ°æ•°æ®åº“å¤±è´¥")

        log_with_user('info', f"å¯¼å…¥å…³é”®è¯æˆåŠŸ: {cid}, æ–°å¢: {add_count}, æ›´æ–°: {update_count}", current_user)

        return {
            "msg": "å¯¼å…¥æˆåŠŸ",
            "total": len(import_data),
            "added": add_count,
            "updated": update_count
        }

    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="Excelæ–‡ä»¶ä¸ºç©º")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Excelæ–‡ä»¶æ ¼å¼é”™è¯¯")
    except Exception as e:
        logger.error(f"å¯¼å…¥å…³é”®è¯å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=f"å¯¼å…¥å…³é”®è¯å¤±è´¥: {str(e)}")


@app.post("/keywords/{cid}/image")
async def add_image_keyword(
    cid: str,
    keyword: str = Form(...),
    item_id: str = Form(default=""),
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """æ·»åŠ å›¾ç‰‡å…³é”®è¯"""
    logger.info(f"æ¥æ”¶åˆ°å›¾ç‰‡å…³é”®è¯æ·»åŠ è¯·æ±‚: cid={cid}, keyword={keyword}, item_id={item_id}")

    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")

    # æ£€æŸ¥å‚æ•°
    if not keyword or not keyword.strip():
        raise HTTPException(status_code=400, detail="å…³é”®è¯ä¸èƒ½ä¸ºç©º")

    if not image or not image.filename:
        raise HTTPException(status_code=400, detail="è¯·é€‰æ‹©å›¾ç‰‡æ–‡ä»¶")

    # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="è´¦å·ä¸å­˜åœ¨æˆ–æ— æƒé™")

    try:
        logger.info(f"æ¥æ”¶åˆ°å›¾ç‰‡å…³é”®è¯æ·»åŠ è¯·æ±‚: cid={cid}, keyword={keyword}, item_id={item_id}, filename={image.filename}")

        # éªŒè¯å›¾ç‰‡æ–‡ä»¶
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"æ— æ•ˆçš„å›¾ç‰‡æ–‡ä»¶ç±»å‹: {image.content_type}")
            raise HTTPException(status_code=400, detail="è¯·ä¸Šä¼ å›¾ç‰‡æ–‡ä»¶")

        # è¯»å–å›¾ç‰‡æ•°æ®
        image_data = await image.read()
        logger.info(f"è¯»å–å›¾ç‰‡æ•°æ®æˆåŠŸï¼Œå¤§å°: {len(image_data)} bytes")

        # ä¿å­˜å›¾ç‰‡
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("å›¾ç‰‡ä¿å­˜å¤±è´¥")
            raise HTTPException(status_code=400, detail="å›¾ç‰‡ä¿å­˜å¤±è´¥")

        logger.info(f"å›¾ç‰‡ä¿å­˜æˆåŠŸ: {image_url}")

        # å…ˆæ£€æŸ¥å…³é”®è¯æ˜¯å¦å·²å­˜åœ¨
        normalized_item_id = item_id if item_id and item_id.strip() else None
        if db_manager.check_keyword_duplicate(cid, keyword, normalized_item_id):
            # åˆ é™¤å·²ä¿å­˜çš„å›¾ç‰‡
            image_manager.delete_image(image_url)
            if normalized_item_id:
                raise HTTPException(status_code=400, detail=f"å…³é”®è¯ '{keyword}' åœ¨å•†å“ '{normalized_item_id}' ä¸­å·²å­˜åœ¨")
            else:
                raise HTTPException(status_code=400, detail=f"é€šç”¨å…³é”®è¯ '{keyword}' å·²å­˜åœ¨")

        # ä¿å­˜å›¾ç‰‡å…³é”®è¯åˆ°æ•°æ®åº“
        success = db_manager.save_image_keyword(cid, keyword, image_url, item_id or None)
        if not success:
            # å¦‚æœæ•°æ®åº“ä¿å­˜å¤±è´¥ï¼Œåˆ é™¤å·²ä¿å­˜çš„å›¾ç‰‡
            logger.error("æ•°æ®åº“ä¿å­˜å¤±è´¥ï¼Œåˆ é™¤å·²ä¿å­˜çš„å›¾ç‰‡")
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=400, detail="å›¾ç‰‡å…³é”®è¯ä¿å­˜å¤±è´¥ï¼Œè¯·ç¨åé‡è¯•")

        log_with_user('info', f"æ·»åŠ å›¾ç‰‡å…³é”®è¯æˆåŠŸ: {cid}, å…³é”®è¯: {keyword}", current_user)

        return {
            "msg": "å›¾ç‰‡å…³é”®è¯æ·»åŠ æˆåŠŸ",
            "keyword": keyword,
            "image_url": image_url,
            "item_id": item_id or None
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"æ·»åŠ å›¾ç‰‡å…³é”®è¯å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=f"æ·»åŠ å›¾ç‰‡å…³é”®è¯å¤±è´¥: {str(e)}")


@app.post("/upload-image")
async def upload_image(
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """ä¸Šä¼ å›¾ç‰‡ï¼ˆç”¨äºå¡åˆ¸ç­‰åŠŸèƒ½ï¼‰"""
    try:
        logger.info(f"æ¥æ”¶åˆ°å›¾ç‰‡ä¸Šä¼ è¯·æ±‚: filename={image.filename}")

        # éªŒè¯å›¾ç‰‡æ–‡ä»¶
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"æ— æ•ˆçš„å›¾ç‰‡æ–‡ä»¶ç±»å‹: {image.content_type}")
            raise HTTPException(status_code=400, detail="è¯·ä¸Šä¼ å›¾ç‰‡æ–‡ä»¶")

        # è¯»å–å›¾ç‰‡æ•°æ®
        image_data = await image.read()
        logger.info(f"è¯»å–å›¾ç‰‡æ•°æ®æˆåŠŸï¼Œå¤§å°: {len(image_data)} bytes")

        # ä¿å­˜å›¾ç‰‡
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("å›¾ç‰‡ä¿å­˜å¤±è´¥")
            raise HTTPException(status_code=400, detail="å›¾ç‰‡ä¿å­˜å¤±è´¥")

        logger.info(f"å›¾ç‰‡ä¸Šä¼ æˆåŠŸ: {image_url}")

        return {
            "message": "å›¾ç‰‡ä¸Šä¼ æˆåŠŸ",
            "image_url": image_url
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"å›¾ç‰‡ä¸Šä¼ å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=f"å›¾ç‰‡ä¸Šä¼ å¤±è´¥: {str(e)}")


@app.get("/keywords-with-type/{cid}")
def get_keywords_with_type(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–åŒ…å«ç±»å‹ä¿¡æ¯çš„å…³é”®è¯åˆ—è¡¨"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")

    # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="è´¦å·ä¸å­˜åœ¨æˆ–æ— æƒé™")

    try:
        keywords = db_manager.get_keywords_with_type(cid)
        return keywords
    except Exception as e:
        logger.error(f"è·å–å…³é”®è¯åˆ—è¡¨å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=f"è·å–å…³é”®è¯åˆ—è¡¨å¤±è´¥: {str(e)}")


@app.delete("/keywords/{cid}/{index}")
def delete_keyword_by_index(cid: str, index: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ ¹æ®ç´¢å¼•åˆ é™¤å…³é”®è¯"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager æœªå°±ç»ª")

    # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="è´¦å·ä¸å­˜åœ¨æˆ–æ— æƒé™")

    try:
        # å…ˆè·å–è¦åˆ é™¤çš„å…³é”®è¯ä¿¡æ¯ï¼ˆç”¨äºåˆ é™¤å›¾ç‰‡æ–‡ä»¶ï¼‰
        keywords = db_manager.get_keywords_with_type(cid)
        if 0 <= index < len(keywords):
            keyword_data = keywords[index]

            # åˆ é™¤å…³é”®è¯
            success = db_manager.delete_keyword_by_index(cid, index)
            if not success:
                raise HTTPException(status_code=400, detail="åˆ é™¤å…³é”®è¯å¤±è´¥")

            # å¦‚æœæ˜¯å›¾ç‰‡å…³é”®è¯ï¼Œåˆ é™¤å¯¹åº”çš„å›¾ç‰‡æ–‡ä»¶
            if keyword_data.get('type') == 'image' and keyword_data.get('image_url'):
                image_manager.delete_image(keyword_data['image_url'])

            log_with_user('info', f"åˆ é™¤å…³é”®è¯æˆåŠŸ: {cid}, ç´¢å¼•: {index}, å…³é”®è¯: {keyword_data.get('keyword')}", current_user)

            return {"msg": "åˆ é™¤æˆåŠŸ"}
        else:
            raise HTTPException(status_code=400, detail="å…³é”®è¯ç´¢å¼•æ— æ•ˆ")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"åˆ é™¤å…³é”®è¯å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=f"åˆ é™¤å…³é”®è¯å¤±è´¥: {str(e)}")


@app.get("/debug/keywords-table-info")
def debug_keywords_table_info(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è°ƒè¯•ï¼šæ£€æŸ¥keywordsè¡¨ç»“æ„"""
    try:
        import sqlite3
        conn = sqlite3.connect(db_manager.db_path)
        cursor = conn.cursor()

        # è·å–è¡¨ç»“æ„ä¿¡æ¯
        cursor.execute("PRAGMA table_info(keywords)")
        columns = cursor.fetchall()

        # è·å–æ•°æ®åº“ç‰ˆæœ¬
        cursor.execute("SELECT value FROM system_settings WHERE key = 'db_version'")
        version_result = cursor.fetchone()
        db_version = version_result[0] if version_result else "æœªçŸ¥"

        conn.close()

        return {
            "db_version": db_version,
            "table_columns": [{"name": col[1], "type": col[2], "default": col[4]} for col in columns]
        }
    except Exception as e:
        logger.error(f"æ£€æŸ¥è¡¨ç»“æ„å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=f"æ£€æŸ¥è¡¨ç»“æ„å¤±è´¥: {str(e)}")


# å¡åˆ¸ç®¡ç†API
@app.get("/cards")
def get_cards(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å½“å‰ç”¨æˆ·çš„å¡åˆ¸åˆ—è¡¨"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        cards = db_manager.get_all_cards(user_id)
        return cards
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/cards")
def create_card(card_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """åˆ›å»ºæ–°å¡åˆ¸"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        card_name = card_data.get('name', 'æœªå‘½åå¡åˆ¸')

        log_with_user('info', f"åˆ›å»ºå¡åˆ¸: {card_name}", current_user)

        # éªŒè¯å¤šè§„æ ¼å­—æ®µ
        is_multi_spec = card_data.get('is_multi_spec', False)
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="å¤šè§„æ ¼å¡åˆ¸å¿…é¡»æä¾›è§„æ ¼åç§°å’Œè§„æ ¼å€¼")

        card_id = db_manager.create_card(
            name=card_data.get('name'),
            card_type=card_data.get('type'),
            api_config=card_data.get('api_config'),
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds', 0),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name') if is_multi_spec else None,
            spec_value=card_data.get('spec_value') if is_multi_spec else None,
            user_id=user_id
        )

        log_with_user('info', f"å¡åˆ¸åˆ›å»ºæˆåŠŸ: {card_name} (ID: {card_id})", current_user)
        return {"id": card_id, "message": "å¡åˆ¸åˆ›å»ºæˆåŠŸ"}
    except Exception as e:
        log_with_user('error', f"åˆ›å»ºå¡åˆ¸å¤±è´¥: {card_data.get('name', 'æœªçŸ¥')} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cards/{card_id}")
def get_card(card_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å•ä¸ªå¡åˆ¸è¯¦æƒ…"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        card = db_manager.get_card_by_id(card_id, user_id)
        if card:
            return card
        else:
            raise HTTPException(status_code=404, detail="å¡åˆ¸ä¸å­˜åœ¨")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cards/{card_id}")
def update_card(card_id: int, card_data: dict, _: None = Depends(require_auth)):
    """æ›´æ–°å¡åˆ¸"""
    try:
        from db_manager import db_manager
        # éªŒè¯å¤šè§„æ ¼å­—æ®µ
        is_multi_spec = card_data.get('is_multi_spec')
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="å¤šè§„æ ¼å¡åˆ¸å¿…é¡»æä¾›è§„æ ¼åç§°å’Œè§„æ ¼å€¼")

        success = db_manager.update_card(
            card_id=card_id,
            name=card_data.get('name'),
            card_type=card_data.get('type'),
            api_config=card_data.get('api_config'),
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds'),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name'),
            spec_value=card_data.get('spec_value')
        )
        if success:
            return {"message": "å¡åˆ¸æ›´æ–°æˆåŠŸ"}
        else:
            raise HTTPException(status_code=404, detail="å¡åˆ¸ä¸å­˜åœ¨")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cards/{card_id}/image")
async def update_card_with_image(
    card_id: int,
    image: UploadFile = File(...),
    name: str = Form(...),
    type: str = Form(...),
    description: str = Form(default=""),
    delay_seconds: int = Form(default=0),
    enabled: bool = Form(default=True),
    is_multi_spec: bool = Form(default=False),
    spec_name: str = Form(default=""),
    spec_value: str = Form(default=""),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """æ›´æ–°å¸¦å›¾ç‰‡çš„å¡åˆ¸"""
    try:
        logger.info(f"æ¥æ”¶åˆ°å¸¦å›¾ç‰‡çš„å¡åˆ¸æ›´æ–°è¯·æ±‚: card_id={card_id}, name={name}, type={type}")

        # éªŒè¯å›¾ç‰‡æ–‡ä»¶
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"æ— æ•ˆçš„å›¾ç‰‡æ–‡ä»¶ç±»å‹: {image.content_type}")
            raise HTTPException(status_code=400, detail="è¯·ä¸Šä¼ å›¾ç‰‡æ–‡ä»¶")

        # éªŒè¯å¤šè§„æ ¼å­—æ®µ
        if is_multi_spec:
            if not spec_name or not spec_value:
                raise HTTPException(status_code=400, detail="å¤šè§„æ ¼å¡åˆ¸å¿…é¡»æä¾›è§„æ ¼åç§°å’Œè§„æ ¼å€¼")

        # è¯»å–å›¾ç‰‡æ•°æ®
        image_data = await image.read()
        logger.info(f"è¯»å–å›¾ç‰‡æ•°æ®æˆåŠŸï¼Œå¤§å°: {len(image_data)} bytes")

        # ä¿å­˜å›¾ç‰‡
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("å›¾ç‰‡ä¿å­˜å¤±è´¥")
            raise HTTPException(status_code=400, detail="å›¾ç‰‡ä¿å­˜å¤±è´¥")

        logger.info(f"å›¾ç‰‡ä¿å­˜æˆåŠŸ: {image_url}")

        # æ›´æ–°å¡åˆ¸
        from db_manager import db_manager
        success = db_manager.update_card(
            card_id=card_id,
            name=name,
            card_type=type,
            image_url=image_url,
            description=description,
            enabled=enabled,
            delay_seconds=delay_seconds,
            is_multi_spec=is_multi_spec,
            spec_name=spec_name if is_multi_spec else None,
            spec_value=spec_value if is_multi_spec else None
        )

        if success:
            logger.info(f"å¡åˆ¸æ›´æ–°æˆåŠŸ: {name} (ID: {card_id})")
            return {"message": "å¡åˆ¸æ›´æ–°æˆåŠŸ", "image_url": image_url}
        else:
            # å¦‚æœæ•°æ®åº“æ›´æ–°å¤±è´¥ï¼Œåˆ é™¤å·²ä¿å­˜çš„å›¾ç‰‡
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=404, detail="å¡åˆ¸ä¸å­˜åœ¨")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"æ›´æ–°å¸¦å›¾ç‰‡çš„å¡åˆ¸å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# è‡ªåŠ¨å‘è´§è§„åˆ™API
@app.get("/delivery-rules")
def get_delivery_rules(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å‘è´§è§„åˆ™åˆ—è¡¨"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rules = db_manager.get_all_delivery_rules(user_id)
        return rules
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/delivery-rules")
def create_delivery_rule(rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """åˆ›å»ºæ–°å‘è´§è§„åˆ™"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rule_id = db_manager.create_delivery_rule(
            keyword=rule_data.get('keyword'),
            card_id=rule_data.get('card_id'),
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        return {"id": rule_id, "message": "å‘è´§è§„åˆ™åˆ›å»ºæˆåŠŸ"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/delivery-rules/{rule_id}")
def get_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å•ä¸ªå‘è´§è§„åˆ™è¯¦æƒ…"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rule = db_manager.get_delivery_rule_by_id(rule_id, user_id)
        if rule:
            return rule
        else:
            raise HTTPException(status_code=404, detail="å‘è´§è§„åˆ™ä¸å­˜åœ¨")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/delivery-rules/{rule_id}")
def update_delivery_rule(rule_id: int, rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ›´æ–°å‘è´§è§„åˆ™"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.update_delivery_rule(
            rule_id=rule_id,
            keyword=rule_data.get('keyword'),
            card_id=rule_data.get('card_id'),
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        if success:
            return {"message": "å‘è´§è§„åˆ™æ›´æ–°æˆåŠŸ"}
        else:
            raise HTTPException(status_code=404, detail="å‘è´§è§„åˆ™ä¸å­˜åœ¨")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/cards/{card_id}")
def delete_card(card_id: int, _: None = Depends(require_auth)):
    """åˆ é™¤å¡åˆ¸"""
    try:
        from db_manager import db_manager
        success = db_manager.delete_card(card_id)
        if success:
            return {"message": "å¡åˆ¸åˆ é™¤æˆåŠŸ"}
        else:
            raise HTTPException(status_code=404, detail="å¡åˆ¸ä¸å­˜åœ¨")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/delivery-rules/{rule_id}")
def delete_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """åˆ é™¤å‘è´§è§„åˆ™"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.delete_delivery_rule(rule_id, user_id)
        if success:
            return {"message": "å‘è´§è§„åˆ™åˆ é™¤æˆåŠŸ"}
        else:
            raise HTTPException(status_code=404, detail="å‘è´§è§„åˆ™ä¸å­˜åœ¨")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== å¤‡ä»½å’Œæ¢å¤ API ====================

@app.get("/backup/export")
def export_backup(current_user: Dict[str, Any] = Depends(get_current_user)):
    """å¯¼å‡ºç”¨æˆ·å¤‡ä»½"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        username = current_user['username']

        # å¯¼å‡ºå½“å‰ç”¨æˆ·çš„æ•°æ®
        backup_data = db_manager.export_backup(user_id)

        # ç”Ÿæˆæ–‡ä»¶å
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"xianyu_backup_{username}_{timestamp}.json"

        # è¿”å›JSONå“åº”ï¼Œè®¾ç½®ä¸‹è½½å¤´
        response = JSONResponse(content=backup_data)
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/json"

        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"å¯¼å‡ºå¤‡ä»½å¤±è´¥: {str(e)}")


@app.post("/backup/import")
def import_backup(file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """å¯¼å…¥ç”¨æˆ·å¤‡ä»½"""
    try:
        # éªŒè¯æ–‡ä»¶ç±»å‹
        if not file.filename.endswith('.json'):
            raise HTTPException(status_code=400, detail="åªæ”¯æŒJSONæ ¼å¼çš„å¤‡ä»½æ–‡ä»¶")

        # è¯»å–æ–‡ä»¶å†…å®¹
        content = file.file.read()
        backup_data = json.loads(content.decode('utf-8'))

        # å¯¼å…¥å¤‡ä»½åˆ°å½“å‰ç”¨æˆ·
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.import_backup(backup_data, user_id)

        if success:
            # å¤‡ä»½å¯¼å…¥æˆåŠŸåï¼Œåˆ·æ–° CookieManager çš„å†…å­˜ç¼“å­˜
            import cookie_manager
            if cookie_manager.manager:
                try:
                    cookie_manager.manager.reload_from_db()
                    logger.info("å¤‡ä»½å¯¼å…¥åå·²åˆ·æ–° CookieManager ç¼“å­˜")
                except Exception as e:
                    logger.error(f"åˆ·æ–° CookieManager ç¼“å­˜å¤±è´¥: {e}")

            return {"message": "å¤‡ä»½å¯¼å…¥æˆåŠŸ"}
        else:
            raise HTTPException(status_code=400, detail="å¤‡ä»½å¯¼å…¥å¤±è´¥")

    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="å¤‡ä»½æ–‡ä»¶æ ¼å¼æ— æ•ˆ")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"å¯¼å…¥å¤‡ä»½å¤±è´¥: {str(e)}")


@app.post("/system/reload-cache")
def reload_cache(_: None = Depends(require_auth)):
    """é‡æ–°åŠ è½½ç³»ç»Ÿç¼“å­˜ï¼ˆç”¨äºæ‰‹åŠ¨åˆ·æ–°æ•°æ®ï¼‰"""
    try:
        import cookie_manager
        if cookie_manager.manager:
            success = cookie_manager.manager.reload_from_db()
            if success:
                return {"message": "ç³»ç»Ÿç¼“å­˜å·²åˆ·æ–°", "success": True}
            else:
                raise HTTPException(status_code=500, detail="ç¼“å­˜åˆ·æ–°å¤±è´¥")
        else:
            raise HTTPException(status_code=500, detail="CookieManager æœªåˆå§‹åŒ–")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"åˆ·æ–°ç¼“å­˜å¤±è´¥: {str(e)}")


# ==================== å•†å“ç®¡ç† API ====================

@app.get("/items")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å½“å‰ç”¨æˆ·çš„æ‰€æœ‰å•†å“ä¿¡æ¯"""
    try:
        # åªè¿”å›å½“å‰ç”¨æˆ·çš„å•†å“ä¿¡æ¯
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_items_by_cookie(cookie_id)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"è·å–å•†å“ä¿¡æ¯å¤±è´¥: {str(e)}")


# ==================== å•†å“æœç´¢ API ====================

class ItemSearchRequest(BaseModel):
    keyword: str
    page: int = 1
    page_size: int = 20

class ItemSearchMultipleRequest(BaseModel):
    keyword: str
    total_pages: int = 1

@app.post("/items/search")
async def search_items(
    search_request: ItemSearchRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """æœç´¢é—²é±¼å•†å“"""
    user_info = f"ã€{current_user.get('username', 'unknown')}#{current_user.get('user_id', 'unknown')}ã€‘" if current_user else "ã€æœªç™»å½•ã€‘"

    try:
        logger.info(f"{user_info} å¼€å§‹å•é¡µæœç´¢: å…³é”®è¯='{search_request.keyword}', é¡µç ={search_request.page}, æ¯é¡µ={search_request.page_size}")

        from utils.item_search import search_xianyu_items

        # æ‰§è¡Œæœç´¢
        result = await search_xianyu_items(
            keyword=search_request.keyword,
            page=search_request.page,
            page_size=search_request.page_size
        )

        # æ£€æŸ¥æ˜¯å¦æœ‰é”™è¯¯
        has_error = result.get("error")
        items_count = len(result.get("items", []))

        logger.info(f"{user_info} å•é¡µæœç´¢å®Œæˆ: è·å–åˆ° {items_count} æ¡æ•°æ®" +
                   (f", é”™è¯¯: {has_error}" if has_error else ""))

        response_data = {
            "success": True,
            "data": result.get("items", []),
            "total": result.get("total", 0),
            "page": search_request.page,
            "page_size": search_request.page_size,
            "keyword": search_request.keyword,
            "is_real_data": result.get("is_real_data", False),
            "source": result.get("source", "unknown")
        }

        # å¦‚æœæœ‰é”™è¯¯ä¿¡æ¯ï¼Œä¹ŸåŒ…å«åœ¨å“åº”ä¸­
        if has_error:
            response_data["error"] = has_error

        return response_data

    except Exception as e:
        error_msg = str(e)
        logger.error(f"{user_info} å•†å“æœç´¢å¤±è´¥: {error_msg}")
        raise HTTPException(status_code=500, detail=f"å•†å“æœç´¢å¤±è´¥: {error_msg}")


@app.get("/cookies/check")
async def check_valid_cookies(
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """æ£€æŸ¥æ˜¯å¦æœ‰æœ‰æ•ˆçš„cookiesè´¦æˆ·ï¼ˆå¿…é¡»æ˜¯å¯ç”¨çŠ¶æ€ï¼‰"""
    try:
        if cookie_manager.manager is None:
            return {
                "success": True,
                "hasValidCookies": False,
                "validCount": 0,
                "enabledCount": 0,
                "totalCount": 0
            }

        from db_manager import db_manager

        # è·å–æ‰€æœ‰cookies
        all_cookies = db_manager.get_all_cookies()

        # æ£€æŸ¥å¯ç”¨çŠ¶æ€å’Œæœ‰æ•ˆæ€§
        valid_cookies = []
        enabled_cookies = []

        for cookie_id, cookie_value in all_cookies.items():
            # æ£€æŸ¥æ˜¯å¦å¯ç”¨
            is_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
            if is_enabled:
                enabled_cookies.append(cookie_id)
                # æ£€æŸ¥æ˜¯å¦æœ‰æ•ˆï¼ˆé•¿åº¦å¤§äº50ï¼‰
                if len(cookie_value) > 50:
                    valid_cookies.append(cookie_id)

        return {
            "success": True,
            "hasValidCookies": len(valid_cookies) > 0,
            "validCount": len(valid_cookies),
            "enabledCount": len(enabled_cookies),
            "totalCount": len(all_cookies)
        }

    except Exception as e:
        logger.error(f"æ£€æŸ¥cookieså¤±è´¥: {str(e)}")
        return {
            "success": False,
            "hasValidCookies": False,
            "error": str(e)
        }

@app.post("/items/search_multiple")
async def search_multiple_pages(
    search_request: ItemSearchMultipleRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """æœç´¢å¤šé¡µé—²é±¼å•†å“"""
    user_info = f"ã€{current_user.get('username', 'unknown')}#{current_user.get('user_id', 'unknown')}ã€‘" if current_user else "ã€æœªç™»å½•ã€‘"

    try:
        logger.info(f"{user_info} å¼€å§‹å¤šé¡µæœç´¢: å…³é”®è¯='{search_request.keyword}', é¡µæ•°={search_request.total_pages}")

        from utils.item_search import search_multiple_pages_xianyu

        # æ‰§è¡Œå¤šé¡µæœç´¢
        result = await search_multiple_pages_xianyu(
            keyword=search_request.keyword,
            total_pages=search_request.total_pages
        )

        # æ£€æŸ¥æ˜¯å¦æœ‰é”™è¯¯
        has_error = result.get("error")
        items_count = len(result.get("items", []))

        logger.info(f"{user_info} å¤šé¡µæœç´¢å®Œæˆ: è·å–åˆ° {items_count} æ¡æ•°æ®" +
                   (f", é”™è¯¯: {has_error}" if has_error else ""))

        response_data = {
            "success": True,
            "data": result.get("items", []),
            "total": result.get("total", 0),
            "total_pages": search_request.total_pages,
            "keyword": search_request.keyword,
            "is_real_data": result.get("is_real_data", False),
            "is_fallback": result.get("is_fallback", False),
            "source": result.get("source", "unknown")
        }

        # å¦‚æœæœ‰é”™è¯¯ä¿¡æ¯ï¼Œä¹ŸåŒ…å«åœ¨å“åº”ä¸­
        if has_error:
            response_data["error"] = has_error

        return response_data

    except Exception as e:
        error_msg = str(e)
        logger.error(f"{user_info} å¤šé¡µå•†å“æœç´¢å¤±è´¥: {error_msg}")
        raise HTTPException(status_code=500, detail=f"å¤šé¡µå•†å“æœç´¢å¤±è´¥: {error_msg}")



@app.get("/items/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–æŒ‡å®šCookieçš„å•†å“ä¿¡æ¯"""
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

        items = db_manager.get_items_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"è·å–å•†å“ä¿¡æ¯å¤±è´¥: {str(e)}")


@app.get("/items/{cookie_id}/{item_id}")
def get_item_detail(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å•†å“è¯¦æƒ…"""
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

        item = db_manager.get_item_info(cookie_id, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="å•†å“ä¸å­˜åœ¨")
        return {"item": item}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"è·å–å•†å“è¯¦æƒ…å¤±è´¥: {str(e)}")


class ItemDetailUpdate(BaseModel):
    item_detail: str


@app.put("/items/{cookie_id}/{item_id}")
def update_item_detail(
    cookie_id: str,
    item_id: str,
    update_data: ItemDetailUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """æ›´æ–°å•†å“è¯¦æƒ…"""
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        success = db_manager.update_item_detail(cookie_id, item_id, update_data.item_detail)
        if success:
            return {"message": "å•†å“è¯¦æƒ…æ›´æ–°æˆåŠŸ"}
        else:
            raise HTTPException(status_code=400, detail="æ›´æ–°å¤±è´¥")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"æ›´æ–°å•†å“è¯¦æƒ…å¤±è´¥: {str(e)}")


@app.delete("/items/{cookie_id}/{item_id}")
def delete_item_info(
    cookie_id: str,
    item_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """åˆ é™¤å•†å“ä¿¡æ¯"""
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        success = db_manager.delete_item_info(cookie_id, item_id)
        if success:
            return {"message": "å•†å“ä¿¡æ¯åˆ é™¤æˆåŠŸ"}
        else:
            raise HTTPException(status_code=404, detail="å•†å“ä¿¡æ¯ä¸å­˜åœ¨")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"åˆ é™¤å•†å“ä¿¡æ¯å¼‚å¸¸: {e}")
        raise HTTPException(status_code=500, detail=f"æœåŠ¡å™¨é”™è¯¯: {str(e)}")


class BatchDeleteRequest(BaseModel):
    items: List[dict]  # [{"cookie_id": "xxx", "item_id": "yyy"}, ...]


class AIReplySettings(BaseModel):
    ai_enabled: bool
    model_name: str = "qwen-plus"
    api_key: str = ""
    base_url: str = "https://dashscope.aliyuncs.com/compatible-mode/v1"
    max_discount_percent: int = 10
    max_discount_amount: int = 100
    max_bargain_rounds: int = 3
    custom_prompts: str = ""


@app.delete("/items/batch")
def batch_delete_items(
    request: BatchDeleteRequest,
    _: None = Depends(require_auth)
):
    """æ‰¹é‡åˆ é™¤å•†å“ä¿¡æ¯"""
    try:
        if not request.items:
            raise HTTPException(status_code=400, detail="åˆ é™¤åˆ—è¡¨ä¸èƒ½ä¸ºç©º")

        success_count = db_manager.batch_delete_item_info(request.items)
        total_count = len(request.items)

        return {
            "message": f"æ‰¹é‡åˆ é™¤å®Œæˆ",
            "success_count": success_count,
            "total_count": total_count,
            "failed_count": total_count - success_count
        }
    except Exception as e:
        logger.error(f"æ‰¹é‡åˆ é™¤å•†å“ä¿¡æ¯å¼‚å¸¸: {e}")
        raise HTTPException(status_code=500, detail=f"æœåŠ¡å™¨é”™è¯¯: {str(e)}")


# ==================== AIå›å¤ç®¡ç†API ====================

@app.get("/ai-reply-settings/{cookie_id}")
def get_ai_reply_settings(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–æŒ‡å®šè´¦å·çš„AIå›å¤è®¾ç½®"""
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

        settings = db_manager.get_ai_reply_settings(cookie_id)
        return settings
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"è·å–AIå›å¤è®¾ç½®å¼‚å¸¸: {e}")
        raise HTTPException(status_code=500, detail=f"æœåŠ¡å™¨é”™è¯¯: {str(e)}")


@app.put("/ai-reply-settings/{cookie_id}")
def update_ai_reply_settings(cookie_id: str, settings: AIReplySettings, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ›´æ–°æŒ‡å®šè´¦å·çš„AIå›å¤è®¾ç½®"""
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™æ“ä½œè¯¥Cookie")

        # æ£€æŸ¥è´¦å·æ˜¯å¦å­˜åœ¨
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager æœªå°±ç»ª')

        # ä¿å­˜è®¾ç½®
        settings_dict = settings.dict()
        success = db_manager.save_ai_reply_settings(cookie_id, settings_dict)

        if success:

            # å¦‚æœå¯ç”¨äº†AIå›å¤ï¼Œè®°å½•æ—¥å¿—
            if settings.ai_enabled:
                logger.info(f"è´¦å· {cookie_id} å¯ç”¨AIå›å¤")
            else:
                logger.info(f"è´¦å· {cookie_id} ç¦ç”¨AIå›å¤")

            return {"message": "AIå›å¤è®¾ç½®æ›´æ–°æˆåŠŸ"}
        else:
            raise HTTPException(status_code=400, detail="æ›´æ–°å¤±è´¥")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"æ›´æ–°AIå›å¤è®¾ç½®å¼‚å¸¸: {e}")
        raise HTTPException(status_code=500, detail=f"æœåŠ¡å™¨é”™è¯¯: {str(e)}")


@app.get("/ai-reply-settings")
def get_all_ai_reply_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å½“å‰ç”¨æˆ·æ‰€æœ‰è´¦å·çš„AIå›å¤è®¾ç½®"""
    try:
        # åªè¿”å›å½“å‰ç”¨æˆ·çš„AIå›å¤è®¾ç½®
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_settings = db_manager.get_all_ai_reply_settings()
        # è¿‡æ»¤åªå±äºå½“å‰ç”¨æˆ·çš„è®¾ç½®
        user_settings = {cid: settings for cid, settings in all_settings.items() if cid in user_cookies}
        return user_settings
    except Exception as e:
        logger.error(f"è·å–æ‰€æœ‰AIå›å¤è®¾ç½®å¼‚å¸¸: {e}")
        raise HTTPException(status_code=500, detail=f"æœåŠ¡å™¨é”™è¯¯: {str(e)}")


@app.post("/ai-reply-test/{cookie_id}")
def test_ai_reply(cookie_id: str, test_data: dict, _: None = Depends(require_auth)):
    """æµ‹è¯•AIå›å¤åŠŸèƒ½"""
    try:
        # æ£€æŸ¥è´¦å·æ˜¯å¦å­˜åœ¨
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager æœªå°±ç»ª')

        if cookie_id not in cookie_manager.manager.cookies:
            raise HTTPException(status_code=404, detail='è´¦å·ä¸å­˜åœ¨')

        # æ£€æŸ¥æ˜¯å¦å¯ç”¨AIå›å¤
        if not ai_reply_engine.is_ai_enabled(cookie_id):
            raise HTTPException(status_code=400, detail='è¯¥è´¦å·æœªå¯ç”¨AIå›å¤')

        # æ„é€ æµ‹è¯•æ•°æ®
        test_message = test_data.get('message', 'ä½ å¥½')
        test_item_info = {
            'title': test_data.get('item_title', 'æµ‹è¯•å•†å“'),
            'price': test_data.get('item_price', 100),
            'desc': test_data.get('item_desc', 'è¿™æ˜¯ä¸€ä¸ªæµ‹è¯•å•†å“')
        }

        # ç”Ÿæˆæµ‹è¯•å›å¤
        reply = ai_reply_engine.generate_reply(
            message=test_message,
            item_info=test_item_info,
            chat_id=f"test_{int(time.time())}",
            cookie_id=cookie_id,
            user_id="test_user",
            item_id="test_item"
        )

        if reply:
            return {"message": "æµ‹è¯•æˆåŠŸ", "reply": reply}
        else:
            raise HTTPException(status_code=400, detail="AIå›å¤ç”Ÿæˆå¤±è´¥")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"æµ‹è¯•AIå›å¤å¼‚å¸¸: {e}")
        raise HTTPException(status_code=500, detail=f"æœåŠ¡å™¨é”™è¯¯: {str(e)}")


# ==================== æ—¥å¿—ç®¡ç†API ====================

@app.get("/logs")
async def get_logs(lines: int = 200, level: str = None, source: str = None, _: None = Depends(require_auth)):
    """è·å–å®æ—¶ç³»ç»Ÿæ—¥å¿—"""
    try:
        # è·å–æ–‡ä»¶æ—¥å¿—æ”¶é›†å™¨
        collector = get_file_log_collector()

        # è·å–æ—¥å¿—
        logs = collector.get_logs(lines=lines, level_filter=level, source_filter=source)

        return {"success": True, "logs": logs}

    except Exception as e:
        return {"success": False, "message": f"è·å–æ—¥å¿—å¤±è´¥: {str(e)}", "logs": []}


@app.get("/risk-control-logs")
async def get_risk_control_logs(
    cookie_id: str = None,
    limit: int = 100,
    offset: int = 0,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """è·å–é£æ§æ—¥å¿—ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    try:
        log_with_user('info', f"æŸ¥è¯¢é£æ§æ—¥å¿—: cookie_id={cookie_id}, limit={limit}, offset={offset}", admin_user)

        # è·å–é£æ§æ—¥å¿—
        logs = db_manager.get_risk_control_logs(cookie_id=cookie_id, limit=limit, offset=offset)
        total_count = db_manager.get_risk_control_logs_count(cookie_id=cookie_id)

        log_with_user('info', f"é£æ§æ—¥å¿—æŸ¥è¯¢æˆåŠŸï¼Œå…± {len(logs)} æ¡è®°å½•ï¼Œæ€»è®¡ {total_count} æ¡", admin_user)

        return {
            "success": True,
            "data": logs,
            "total": total_count,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        log_with_user('error', f"è·å–é£æ§æ—¥å¿—å¤±è´¥: {str(e)}", admin_user)
        return {
            "success": False,
            "message": f"è·å–é£æ§æ—¥å¿—å¤±è´¥: {str(e)}",
            "data": [],
            "total": 0
        }


@app.delete("/risk-control-logs/{log_id}")
async def delete_risk_control_log(
    log_id: int,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """åˆ é™¤é£æ§æ—¥å¿—è®°å½•ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    try:
        log_with_user('info', f"åˆ é™¤é£æ§æ—¥å¿—è®°å½•: {log_id}", admin_user)

        success = db_manager.delete_risk_control_log(log_id)

        if success:
            log_with_user('info', f"é£æ§æ—¥å¿—åˆ é™¤æˆåŠŸ: {log_id}", admin_user)
            return {"success": True, "message": "åˆ é™¤æˆåŠŸ"}
        else:
            log_with_user('warning', f"é£æ§æ—¥å¿—åˆ é™¤å¤±è´¥: {log_id}", admin_user)
            return {"success": False, "message": "åˆ é™¤å¤±è´¥ï¼Œè®°å½•å¯èƒ½ä¸å­˜åœ¨"}

    except Exception as e:
        log_with_user('error', f"åˆ é™¤é£æ§æ—¥å¿—å¤±è´¥: {log_id} - {str(e)}", admin_user)
        return {"success": False, "message": f"åˆ é™¤å¤±è´¥: {str(e)}"}


@app.get("/logs/stats")
async def get_log_stats(_: None = Depends(require_auth)):
    """è·å–æ—¥å¿—ç»Ÿè®¡ä¿¡æ¯"""
    try:
        collector = get_file_log_collector()
        stats = collector.get_stats()

        return {"success": True, "stats": stats}

    except Exception as e:
        return {"success": False, "message": f"è·å–æ—¥å¿—ç»Ÿè®¡å¤±è´¥: {str(e)}", "stats": {}}


@app.post("/logs/clear")
async def clear_logs(_: None = Depends(require_auth)):
    """æ¸…ç©ºæ—¥å¿—"""
    try:
        collector = get_file_log_collector()
        collector.clear_logs()

        return {"success": True, "message": "æ—¥å¿—å·²æ¸…ç©º"}

    except Exception as e:
        return {"success": False, "message": f"æ¸…ç©ºæ—¥å¿—å¤±è´¥: {str(e)}"}


# ==================== å•†å“ç®¡ç†API ====================

@app.post("/items/get-all-from-account")
async def get_all_items_from_account(request: dict, _: None = Depends(require_auth)):
    """ä»æŒ‡å®šè´¦å·è·å–æ‰€æœ‰å•†å“ä¿¡æ¯"""
    try:
        cookie_id = request.get('cookie_id')
        if not cookie_id:
            return {"success": False, "message": "ç¼ºå°‘cookie_idå‚æ•°"}

        # è·å–æŒ‡å®šè´¦å·çš„cookieä¿¡æ¯
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {"success": False, "message": "æœªæ‰¾åˆ°æŒ‡å®šçš„è´¦å·ä¿¡æ¯"}

        cookies_str = cookie_info.get('cookies_str', '')
        if not cookies_str:
            return {"success": False, "message": "è´¦å·cookieä¿¡æ¯ä¸ºç©º"}

        # åˆ›å»ºXianyuLiveå®ä¾‹ï¼Œä¼ å…¥æ­£ç¡®çš„cookie_id
        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cookie_id)

        # è°ƒç”¨è·å–æ‰€æœ‰å•†å“ä¿¡æ¯çš„æ–¹æ³•ï¼ˆè‡ªåŠ¨åˆ†é¡µï¼‰
        logger.info(f"å¼€å§‹è·å–è´¦å· {cookie_id} çš„æ‰€æœ‰å•†å“ä¿¡æ¯")
        result = await xianyu_instance.get_all_items()

        # å…³é—­session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"è·å–å•†å“ä¿¡æ¯å¤±è´¥: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            total_count = result.get('total_count', 0)
            total_pages = result.get('total_pages', 1)
            logger.info(f"æˆåŠŸè·å–è´¦å· {cookie_id} çš„ {total_count} ä¸ªå•†å“ï¼ˆå…±{total_pages}é¡µï¼‰")
            return {
                "success": True,
                "message": f"æˆåŠŸè·å– {total_count} ä¸ªå•†å“ï¼ˆå…±{total_pages}é¡µï¼‰ï¼Œè¯¦ç»†ä¿¡æ¯å·²æ‰“å°åˆ°æ§åˆ¶å°",
                "total_count": total_count,
                "total_pages": total_pages
            }

    except Exception as e:
        logger.error(f"è·å–è´¦å·å•†å“ä¿¡æ¯å¼‚å¸¸: {str(e)}")
        return {"success": False, "message": f"è·å–å•†å“ä¿¡æ¯å¼‚å¸¸: {str(e)}"}


@app.post("/items/get-by-page")
async def get_items_by_page(request: dict, _: None = Depends(require_auth)):
    """ä»æŒ‡å®šè´¦å·æŒ‰é¡µè·å–å•†å“ä¿¡æ¯"""
    try:
        # éªŒè¯å‚æ•°
        cookie_id = request.get('cookie_id')
        page_number = request.get('page_number', 1)
        page_size = request.get('page_size', 20)

        if not cookie_id:
            return {"success": False, "message": "ç¼ºå°‘cookie_idå‚æ•°"}

        # éªŒè¯åˆ†é¡µå‚æ•°
        try:
            page_number = int(page_number)
            page_size = int(page_size)
        except (ValueError, TypeError):
            return {"success": False, "message": "é¡µç å’Œæ¯é¡µæ•°é‡å¿…é¡»æ˜¯æ•°å­—"}

        if page_number < 1:
            return {"success": False, "message": "é¡µç å¿…é¡»å¤§äº0"}

        if page_size < 1 or page_size > 100:
            return {"success": False, "message": "æ¯é¡µæ•°é‡å¿…é¡»åœ¨1-100ä¹‹é—´"}

        # è·å–è´¦å·ä¿¡æ¯
        account = db_manager.get_cookie_by_id(cookie_id)
        if not account:
            return {"success": False, "message": "è´¦å·ä¸å­˜åœ¨"}

        cookies_str = account['cookies_str']
        if not cookies_str:
            return {"success": False, "message": "è´¦å·cookiesä¸ºç©º"}

        # åˆ›å»ºXianyuLiveå®ä¾‹ï¼Œä¼ å…¥æ­£ç¡®çš„cookie_id
        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cookie_id)

        # è°ƒç”¨è·å–æŒ‡å®šé¡µå•†å“ä¿¡æ¯çš„æ–¹æ³•
        logger.info(f"å¼€å§‹è·å–è´¦å· {cookie_id} ç¬¬{page_number}é¡µå•†å“ä¿¡æ¯ï¼ˆæ¯é¡µ{page_size}æ¡ï¼‰")
        result = await xianyu_instance.get_item_list_info(page_number, page_size)

        # å…³é—­session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"è·å–å•†å“ä¿¡æ¯å¤±è´¥: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            current_count = result.get('current_count', 0)
            logger.info(f"æˆåŠŸè·å–è´¦å· {cookie_id} ç¬¬{page_number}é¡µ {current_count} ä¸ªå•†å“")
            return {
                "success": True,
                "message": f"æˆåŠŸè·å–ç¬¬{page_number}é¡µ {current_count} ä¸ªå•†å“ï¼Œè¯¦ç»†ä¿¡æ¯å·²æ‰“å°åˆ°æ§åˆ¶å°",
                "page_number": page_number,
                "page_size": page_size,
                "current_count": current_count
            }

    except Exception as e:
        logger.error(f"è·å–è´¦å·å•†å“ä¿¡æ¯å¼‚å¸¸: {str(e)}")
        return {"success": False, "message": f"è·å–å•†å“ä¿¡æ¯å¼‚å¸¸: {str(e)}"}


# ------------------------- ç”¨æˆ·è®¾ç½®æ¥å£ -------------------------

@app.get('/user-settings')
def get_user_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å½“å‰ç”¨æˆ·çš„è®¾ç½®"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        settings = db_manager.get_user_settings(user_id)
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/user-settings/{key}')
def update_user_setting(key: str, setting_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """æ›´æ–°ç”¨æˆ·è®¾ç½®"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        value = setting_data.get('value')
        description = setting_data.get('description', '')

        log_with_user('info', f"æ›´æ–°ç”¨æˆ·è®¾ç½®: {key} = {value}", current_user)

        success = db_manager.set_user_setting(user_id, key, value, description)
        if success:
            log_with_user('info', f"ç”¨æˆ·è®¾ç½®æ›´æ–°æˆåŠŸ: {key}", current_user)
            return {'msg': 'setting updated', 'key': key, 'value': value}
        else:
            log_with_user('error', f"ç”¨æˆ·è®¾ç½®æ›´æ–°å¤±è´¥: {key}", current_user)
            raise HTTPException(status_code=400, detail='æ›´æ–°å¤±è´¥')
    except Exception as e:
        log_with_user('error', f"æ›´æ–°ç”¨æˆ·è®¾ç½®å¼‚å¸¸: {key} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/user-settings/{key}')
def get_user_setting(key: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–ç”¨æˆ·ç‰¹å®šè®¾ç½®"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        setting = db_manager.get_user_setting(user_id, key)
        if setting:
            return setting
        else:
            raise HTTPException(status_code=404, detail='è®¾ç½®ä¸å­˜åœ¨')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- ç®¡ç†å‘˜ä¸“ç”¨æ¥å£ -------------------------

@app.get('/admin/users')
def get_all_users(admin_user: Dict[str, Any] = Depends(require_admin)):
    """è·å–æ‰€æœ‰ç”¨æˆ·ä¿¡æ¯ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    from db_manager import db_manager
    try:
        log_with_user('info', "æŸ¥è¯¢æ‰€æœ‰ç”¨æˆ·ä¿¡æ¯", admin_user)
        users = db_manager.get_all_users()

        # ä¸ºæ¯ä¸ªç”¨æˆ·æ·»åŠ ç»Ÿè®¡ä¿¡æ¯
        for user in users:
            user_id = user['id']
            # ç»Ÿè®¡ç”¨æˆ·çš„Cookieæ•°é‡
            user_cookies = db_manager.get_all_cookies(user_id)
            user['cookie_count'] = len(user_cookies)

            # ç»Ÿè®¡ç”¨æˆ·çš„å¡åˆ¸æ•°é‡
            user_cards = db_manager.get_all_cards(user_id)
            user['card_count'] = len(user_cards) if user_cards else 0

            # éšè—å¯†ç å­—æ®µ
            if 'password_hash' in user:
                del user['password_hash']

        log_with_user('info', f"è¿”å›ç”¨æˆ·ä¿¡æ¯ï¼Œå…± {len(users)} ä¸ªç”¨æˆ·", admin_user)
        return {"users": users}
    except Exception as e:
        log_with_user('error', f"è·å–ç”¨æˆ·ä¿¡æ¯å¤±è´¥: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/users/{user_id}')
def delete_user(user_id: int, admin_user: Dict[str, Any] = Depends(require_admin)):
    """åˆ é™¤ç”¨æˆ·ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    from db_manager import db_manager
    try:
        # ä¸èƒ½åˆ é™¤ç®¡ç†å‘˜è‡ªå·±
        if user_id == admin_user['user_id']:
            log_with_user('warning', "å°è¯•åˆ é™¤ç®¡ç†å‘˜è‡ªå·±", admin_user)
            raise HTTPException(status_code=400, detail="ä¸èƒ½åˆ é™¤ç®¡ç†å‘˜è‡ªå·±")

        # è·å–è¦åˆ é™¤çš„ç”¨æˆ·ä¿¡æ¯
        user_to_delete = db_manager.get_user_by_id(user_id)
        if not user_to_delete:
            raise HTTPException(status_code=404, detail="ç”¨æˆ·ä¸å­˜åœ¨")

        log_with_user('info', f"å‡†å¤‡åˆ é™¤ç”¨æˆ·: {user_to_delete['username']} (ID: {user_id})", admin_user)

        # åˆ é™¤ç”¨æˆ·åŠå…¶ç›¸å…³æ•°æ®
        success = db_manager.delete_user_and_data(user_id)

        if success:
            log_with_user('info', f"ç”¨æˆ·åˆ é™¤æˆåŠŸ: {user_to_delete['username']} (ID: {user_id})", admin_user)
            return {"message": f"ç”¨æˆ· {user_to_delete['username']} åˆ é™¤æˆåŠŸ"}
        else:
            log_with_user('error', f"ç”¨æˆ·åˆ é™¤å¤±è´¥: {user_to_delete['username']} (ID: {user_id})", admin_user)
            raise HTTPException(status_code=400, detail="åˆ é™¤å¤±è´¥")
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"åˆ é™¤ç”¨æˆ·å¼‚å¸¸: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/risk-control-logs')
async def get_admin_risk_control_logs(
    cookie_id: str = None,
    limit: int = 100,
    offset: int = 0,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """è·å–é£æ§æ—¥å¿—ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    try:
        log_with_user('info', f"æŸ¥è¯¢é£æ§æ—¥å¿—: cookie_id={cookie_id}, limit={limit}, offset={offset}", admin_user)

        # è·å–é£æ§æ—¥å¿—
        logs = db_manager.get_risk_control_logs(cookie_id=cookie_id, limit=limit, offset=offset)
        total_count = db_manager.get_risk_control_logs_count(cookie_id=cookie_id)

        log_with_user('info', f"é£æ§æ—¥å¿—æŸ¥è¯¢æˆåŠŸï¼Œå…± {len(logs)} æ¡è®°å½•ï¼Œæ€»è®¡ {total_count} æ¡", admin_user)

        return {
            "success": True,
            "data": logs,
            "total": total_count,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        log_with_user('error', f"æŸ¥è¯¢é£æ§æ—¥å¿—å¤±è´¥: {str(e)}", admin_user)
        return {"success": False, "message": f"æŸ¥è¯¢å¤±è´¥: {str(e)}", "data": [], "total": 0}


@app.get('/admin/cookies')
def get_admin_cookies(admin_user: Dict[str, Any] = Depends(require_admin)):
    """è·å–æ‰€æœ‰Cookieä¿¡æ¯ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    try:
        log_with_user('info', "æŸ¥è¯¢æ‰€æœ‰Cookieä¿¡æ¯", admin_user)

        if cookie_manager.manager is None:
            return {
                "success": True,
                "cookies": [],
                "message": "CookieManager æœªå°±ç»ª"
            }

        # è·å–æ‰€æœ‰ç”¨æˆ·çš„cookies
        from db_manager import db_manager
        all_users = db_manager.get_all_users()
        all_cookies = []

        for user in all_users:
            user_id = user['id']
            user_cookies = db_manager.get_all_cookies(user_id)
            for cookie_id, cookie_value in user_cookies.items():
                # è·å–cookieè¯¦ç»†ä¿¡æ¯
                cookie_details = db_manager.get_cookie_details(cookie_id)
                cookie_info = {
                    'cookie_id': cookie_id,
                    'user_id': user_id,
                    'username': user['username'],
                    'nickname': cookie_details.get('remark', '') if cookie_details else '',
                    'enabled': cookie_manager.manager.get_cookie_status(cookie_id)
                }
                all_cookies.append(cookie_info)

        log_with_user('info', f"è·å–åˆ° {len(all_cookies)} ä¸ªCookie", admin_user)
        return {
            "success": True,
            "cookies": all_cookies,
            "total": len(all_cookies)
        }

    except Exception as e:
        log_with_user('error', f"è·å–Cookieä¿¡æ¯å¤±è´¥: {str(e)}", admin_user)
        return {
            "success": False,
            "cookies": [],
            "message": f"è·å–å¤±è´¥: {str(e)}"
        }


@app.get('/admin/logs')
def get_system_logs(admin_user: Dict[str, Any] = Depends(require_admin),
                   lines: int = 100,
                   level: str = None):
    """è·å–ç³»ç»Ÿæ—¥å¿—ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', f"æŸ¥è¯¢ç³»ç»Ÿæ—¥å¿—ï¼Œè¡Œæ•°: {lines}, çº§åˆ«: {level}", admin_user)

        # æŸ¥æ‰¾æ—¥å¿—æ–‡ä»¶
        log_files = glob.glob("logs/xianyu_*.log")
        logger.info(f"æ‰¾åˆ°æ—¥å¿—æ–‡ä»¶: {log_files}")

        if not log_files:
            logger.warning("æœªæ‰¾åˆ°æ—¥å¿—æ–‡ä»¶")
            return {"logs": [], "message": "æœªæ‰¾åˆ°æ—¥å¿—æ–‡ä»¶", "success": False}

        # è·å–æœ€æ–°çš„æ—¥å¿—æ–‡ä»¶
        latest_log_file = max(log_files, key=os.path.getctime)
        logger.info(f"ä½¿ç”¨æœ€æ–°æ—¥å¿—æ–‡ä»¶: {latest_log_file}")

        logs = []
        try:
            with open(latest_log_file, 'r', encoding='utf-8') as f:
                all_lines = f.readlines()
                logger.info(f"è¯»å–åˆ° {len(all_lines)} è¡Œæ—¥å¿—")

                # å¦‚æœæŒ‡å®šäº†æ—¥å¿—çº§åˆ«ï¼Œè¿›è¡Œè¿‡æ»¤
                if level:
                    filtered_lines = [line for line in all_lines if f"| {level.upper()} |" in line]
                    logger.info(f"æŒ‰çº§åˆ« {level} è¿‡æ»¤åå‰©ä½™ {len(filtered_lines)} è¡Œ")
                else:
                    filtered_lines = all_lines

                # è·å–æœ€åNè¡Œ
                recent_lines = filtered_lines[-lines:] if len(filtered_lines) > lines else filtered_lines
                logger.info(f"å–æœ€å {len(recent_lines)} è¡Œæ—¥å¿—")

                for line in recent_lines:
                    logs.append(line.strip())

        except Exception as e:
            logger.error(f"è¯»å–æ—¥å¿—æ–‡ä»¶å¤±è´¥: {str(e)}")
            log_with_user('error', f"è¯»å–æ—¥å¿—æ–‡ä»¶å¤±è´¥: {str(e)}", admin_user)
            return {"logs": [], "message": f"è¯»å–æ—¥å¿—æ–‡ä»¶å¤±è´¥: {str(e)}", "success": False}

        log_with_user('info', f"è¿”å›æ—¥å¿—è®°å½• {len(logs)} æ¡", admin_user)
        logger.info(f"æˆåŠŸè¿”å› {len(logs)} æ¡æ—¥å¿—è®°å½•")

        return {
            "logs": logs,
            "log_file": latest_log_file,
            "total_lines": len(logs),
            "success": True
        }

    except Exception as e:
        logger.error(f"è·å–ç³»ç»Ÿæ—¥å¿—å¤±è´¥: {str(e)}")
        log_with_user('error', f"è·å–ç³»ç»Ÿæ—¥å¿—å¤±è´¥: {str(e)}", admin_user)
        return {"logs": [], "message": f"è·å–ç³»ç»Ÿæ—¥å¿—å¤±è´¥: {str(e)}", "success": False}

@app.get('/admin/log-files')
def list_log_files(admin_user: Dict[str, Any] = Depends(require_admin)):
    """åˆ—å‡ºæ‰€æœ‰å¯ç”¨çš„ç³»ç»Ÿæ—¥å¿—æ–‡ä»¶"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', "æŸ¥è¯¢æ—¥å¿—æ–‡ä»¶åˆ—è¡¨", admin_user)

        log_dir = "logs"
        if not os.path.exists(log_dir):
            logger.warning("æ—¥å¿—ç›®å½•ä¸å­˜åœ¨")
            return {"success": True, "files": []}

        log_pattern = os.path.join(log_dir, "xianyu_*.log")
        log_files = glob.glob(log_pattern)

        files_info = []
        for file_path in log_files:
            try:
                stat_info = os.stat(file_path)
                files_info.append({
                    "name": os.path.basename(file_path),
                    "size": stat_info.st_size,
                    "modified_at": datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                    "modified_ts": stat_info.st_mtime
                })
            except OSError as e:
                logger.warning(f"è¯»å–æ—¥å¿—æ–‡ä»¶ä¿¡æ¯å¤±è´¥ {file_path}: {e}")

        # æŒ‰ä¿®æ”¹æ—¶é—´å€’åºæ’åº
        files_info.sort(key=lambda item: item.get("modified_ts", 0), reverse=True)

        logger.info(f"è¿”å›æ—¥å¿—æ–‡ä»¶åˆ—è¡¨ï¼Œå…± {len(files_info)} ä¸ªæ–‡ä»¶")
        return {"success": True, "files": files_info}

    except Exception as e:
        logger.error(f"è·å–æ—¥å¿—æ–‡ä»¶åˆ—è¡¨å¤±è´¥: {str(e)}")
        log_with_user('error', f"è·å–æ—¥å¿—æ–‡ä»¶åˆ—è¡¨å¤±è´¥: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/logs/export')
def export_log_file(file: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """å¯¼å‡ºæŒ‡å®šçš„æ—¥å¿—æ–‡ä»¶"""
    import os
    from fastapi.responses import StreamingResponse

    try:
        if not file:
            raise HTTPException(status_code=400, detail="ç¼ºå°‘æ–‡ä»¶å‚æ•°")

        safe_name = os.path.basename(file)
        log_dir = os.path.abspath("logs")
        target_path = os.path.abspath(os.path.join(log_dir, safe_name))

        # é˜²æ­¢ç›®å½•éå†
        if not target_path.startswith(log_dir):
            log_with_user('warning', f"å°è¯•è®¿é—®éæ³•æ—¥å¿—æ–‡ä»¶: {file}", admin_user)
            raise HTTPException(status_code=400, detail="éæ³•çš„æ—¥å¿—æ–‡ä»¶è·¯å¾„")

        if not os.path.exists(target_path):
            log_with_user('warning', f"æ—¥å¿—æ–‡ä»¶ä¸å­˜åœ¨: {file}", admin_user)
            raise HTTPException(status_code=404, detail="æ—¥å¿—æ–‡ä»¶ä¸å­˜åœ¨")

        log_with_user('info', f"å¯¼å‡ºæ—¥å¿—æ–‡ä»¶: {safe_name}", admin_user)
        def iter_file(path: str):
            file_handle = open(path, 'rb')
            try:
                while True:
                    chunk = file_handle.read(8192)
                    if not chunk:
                        break
                    yield chunk
            finally:
                file_handle.close()

        headers = {
            "Content-Disposition": f'attachment; filename="{safe_name}"'
        }
        return StreamingResponse(
            iter_file(target_path),
            media_type='text/plain; charset=utf-8',
            headers=headers
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"å¯¼å‡ºæ—¥å¿—æ–‡ä»¶å¤±è´¥: {str(e)}")
        log_with_user('error', f"å¯¼å‡ºæ—¥å¿—æ–‡ä»¶å¤±è´¥: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/stats')
def get_system_stats(admin_user: Dict[str, Any] = Depends(require_admin)):
    """è·å–ç³»ç»Ÿç»Ÿè®¡ä¿¡æ¯ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    from db_manager import db_manager
    try:
        log_with_user('info', "æŸ¥è¯¢ç³»ç»Ÿç»Ÿè®¡ä¿¡æ¯", admin_user)

        stats = {
            "users": {
                "total": 0,
                "active_today": 0
            },
            "cookies": {
                "total": 0,
                "enabled": 0
            },
            "cards": {
                "total": 0,
                "enabled": 0
            },
            "system": {
                "uptime": "æœªçŸ¥",
                "version": "1.0.0"
            }
        }

        # ç”¨æˆ·ç»Ÿè®¡
        all_users = db_manager.get_all_users()
        stats["users"]["total"] = len(all_users)

        # Cookieç»Ÿè®¡
        all_cookies = db_manager.get_all_cookies()
        stats["cookies"]["total"] = len(all_cookies)

        # å¡åˆ¸ç»Ÿè®¡
        all_cards = db_manager.get_all_cards()
        if all_cards:
            stats["cards"]["total"] = len(all_cards)
            stats["cards"]["enabled"] = len([card for card in all_cards if card.get('enabled', True)])

        log_with_user('info', "ç³»ç»Ÿç»Ÿè®¡ä¿¡æ¯æŸ¥è¯¢å®Œæˆ", admin_user)
        return stats

    except Exception as e:
        log_with_user('error', f"è·å–ç³»ç»Ÿç»Ÿè®¡ä¿¡æ¯å¤±è´¥: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

# ------------------------- æŒ‡å®šå•†å“å›å¤æ¥å£ -------------------------

@app.get("/itemReplays")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å½“å‰ç”¨æˆ·çš„æ‰€æœ‰å•†å“å›å¤ä¿¡æ¯"""
    try:
        # åªè¿”å›å½“å‰ç”¨æˆ·çš„å•†å“ä¿¡æ¯
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_itemReplays_by_cookie(cookie_id)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"è·å–å•†å“å›å¤ä¿¡æ¯å¤±è´¥: {str(e)}")

@app.get("/itemReplays/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–æŒ‡å®šCookieçš„å•†å“ä¿¡æ¯"""
    try:
        # æ£€æŸ¥cookieæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

        items = db_manager.get_itemReplays_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"è·å–å•†å“ä¿¡æ¯å¤±è´¥: {str(e)}")

@app.put("/item-reply/{cookie_id}/{item_id}")
def update_item_reply(
    cookie_id: str,
    item_id: str,
    data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    æ›´æ–°æŒ‡å®šè´¦å·å’Œå•†å“çš„å›å¤å†…å®¹
    """
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager

        # éªŒè¯cookieæ˜¯å¦å±äºç”¨æˆ·
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

        reply_content = data.get("reply_content", "").strip()
        if not reply_content:
            raise HTTPException(status_code=400, detail="å›å¤å†…å®¹ä¸èƒ½ä¸ºç©º")

        db_manager.update_item_reply(cookie_id=cookie_id, item_id=item_id, reply_content=reply_content)

        return {"message": "å•†å“å›å¤æ›´æ–°æˆåŠŸ"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"æ›´æ–°å•†å“å›å¤å¤±è´¥: {str(e)}")

@app.delete("/item-reply/{cookie_id}/{item_id}")
def delete_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    åˆ é™¤æŒ‡å®šè´¦å·cookie_idå’Œå•†å“item_idçš„å•†å“å›å¤
    """
    try:
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

        success = db_manager.delete_item_reply(cookie_id, item_id)
        if not success:
            raise HTTPException(status_code=404, detail="å•†å“å›å¤ä¸å­˜åœ¨")

        return {"message": "å•†å“å›å¤åˆ é™¤æˆåŠŸ"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"åˆ é™¤å•†å“å›å¤å¤±è´¥: {str(e)}")

class ItemToDelete(BaseModel):
    cookie_id: str
    item_id: str

class BatchDeleteRequest(BaseModel):
    items: List[ItemToDelete]

@app.delete("/item-reply/batch")
async def batch_delete_item_reply(
    req: BatchDeleteRequest,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    æ‰¹é‡åˆ é™¤å•†å“å›å¤
    """
    user_id = current_user['user_id']
    from db_manager import db_manager

    # å…ˆæ ¡éªŒå½“å‰ç”¨æˆ·æ˜¯å¦æœ‰æƒé™åˆ é™¤æ¯ä¸ªcookieå¯¹åº”çš„å›å¤
    user_cookies = db_manager.get_all_cookies(user_id)
    for item in req.items:
        if item.cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail=f"æ— æƒé™è®¿é—®Cookie {item.cookie_id}")

    result = db_manager.batch_delete_item_replies([item.dict() for item in req.items])
    return {
        "success_count": result["success_count"],
        "failed_count": result["failed_count"]
    }

@app.get("/item-reply/{cookie_id}/{item_id}")
def get_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    è·å–æŒ‡å®šè´¦å·cookie_idå’Œå•†å“item_idçš„å•†å“å›å¤å†…å®¹
    """
    try:
        user_id = current_user['user_id']
        # æ ¡éªŒcookie_idæ˜¯å¦å±äºå½“å‰ç”¨æˆ·
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="æ— æƒé™è®¿é—®è¯¥Cookie")

        # è·å–æŒ‡å®šå•†å“å›å¤
        item_replies = db_manager.get_itemReplays_by_cookie(cookie_id)
        # æ‰¾å¯¹åº”item_idçš„å›å¤
        item_reply = next((r for r in item_replies if r['item_id'] == item_id), None)

        if item_reply is None:
            raise HTTPException(status_code=404, detail="å•†å“å›å¤ä¸å­˜åœ¨")

        return item_reply

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"è·å–å•†å“å›å¤å¤±è´¥: {str(e)}")


# ------------------------- æ•°æ®åº“å¤‡ä»½å’Œæ¢å¤æ¥å£ -------------------------

@app.get('/admin/backup/download')
def download_database_backup(admin_user: Dict[str, Any] = Depends(require_admin)):
    """ä¸‹è½½æ•°æ®åº“å¤‡ä»½æ–‡ä»¶ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    import os
    from fastapi.responses import FileResponse
    from datetime import datetime

    try:
        log_with_user('info', "è¯·æ±‚ä¸‹è½½æ•°æ®åº“å¤‡ä»½", admin_user)

        # ä½¿ç”¨db_managerçš„å®é™…æ•°æ®åº“è·¯å¾„
        from db_manager import db_manager
        db_file_path = db_manager.db_path

        # æ£€æŸ¥æ•°æ®åº“æ–‡ä»¶æ˜¯å¦å­˜åœ¨
        if not os.path.exists(db_file_path):
            log_with_user('error', f"æ•°æ®åº“æ–‡ä»¶ä¸å­˜åœ¨: {db_file_path}", admin_user)
            raise HTTPException(status_code=404, detail="æ•°æ®åº“æ–‡ä»¶ä¸å­˜åœ¨")

        # ç”Ÿæˆå¸¦æ—¶é—´æˆ³çš„æ–‡ä»¶å
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"xianyu_backup_{timestamp}.db"

        log_with_user('info', f"å¼€å§‹ä¸‹è½½æ•°æ®åº“å¤‡ä»½: {download_filename}", admin_user)

        return FileResponse(
            path=db_file_path,
            filename=download_filename,
            media_type='application/octet-stream'
        )

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"ä¸‹è½½æ•°æ®åº“å¤‡ä»½å¤±è´¥: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.post('/admin/backup/upload')
async def upload_database_backup(admin_user: Dict[str, Any] = Depends(require_admin),
                                backup_file: UploadFile = File(...)):
    """ä¸Šä¼ å¹¶æ¢å¤æ•°æ®åº“å¤‡ä»½æ–‡ä»¶ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    import os
    import shutil
    import sqlite3
    from datetime import datetime

    try:
        log_with_user('info', f"å¼€å§‹ä¸Šä¼ æ•°æ®åº“å¤‡ä»½: {backup_file.filename}", admin_user)

        # éªŒè¯æ–‡ä»¶ç±»å‹
        if not backup_file.filename.endswith('.db'):
            log_with_user('warning', f"æ— æ•ˆçš„å¤‡ä»½æ–‡ä»¶ç±»å‹: {backup_file.filename}", admin_user)
            raise HTTPException(status_code=400, detail="åªæ”¯æŒ.dbæ ¼å¼çš„æ•°æ®åº“æ–‡ä»¶")

        # éªŒè¯æ–‡ä»¶å¤§å°ï¼ˆé™åˆ¶100MBï¼‰
        content = await backup_file.read()
        if len(content) > 100 * 1024 * 1024:  # 100MB
            log_with_user('warning', f"å¤‡ä»½æ–‡ä»¶è¿‡å¤§: {len(content)} bytes", admin_user)
            raise HTTPException(status_code=400, detail="å¤‡ä»½æ–‡ä»¶å¤§å°ä¸èƒ½è¶…è¿‡100MB")

        # éªŒè¯æ˜¯å¦ä¸ºæœ‰æ•ˆçš„SQLiteæ•°æ®åº“æ–‡ä»¶
        temp_file_path = f"temp_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"

        try:
            # ä¿å­˜ä¸´æ—¶æ–‡ä»¶
            with open(temp_file_path, 'wb') as temp_file:
                temp_file.write(content)

            # éªŒè¯æ•°æ®åº“æ–‡ä»¶å®Œæ•´æ€§
            conn = sqlite3.connect(temp_file_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()
            conn.close()

            # æ£€æŸ¥æ˜¯å¦åŒ…å«å¿…è¦çš„è¡¨
            table_names = [table[0] for table in tables]
            required_tables = ['users', 'cookies']  # æœ€åŸºæœ¬çš„è¡¨

            missing_tables = [table for table in required_tables if table not in table_names]
            if missing_tables:
                log_with_user('warning', f"å¤‡ä»½æ–‡ä»¶ç¼ºå°‘å¿…è¦çš„è¡¨: {missing_tables}", admin_user)
                raise HTTPException(status_code=400, detail=f"å¤‡ä»½æ–‡ä»¶ä¸å®Œæ•´ï¼Œç¼ºå°‘è¡¨: {', '.join(missing_tables)}")

            log_with_user('info', f"å¤‡ä»½æ–‡ä»¶éªŒè¯é€šè¿‡ï¼ŒåŒ…å« {len(table_names)} ä¸ªè¡¨", admin_user)

        except sqlite3.Error as e:
            log_with_user('error', f"å¤‡ä»½æ–‡ä»¶éªŒè¯å¤±è´¥: {str(e)}", admin_user)
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            raise HTTPException(status_code=400, detail="æ— æ•ˆçš„æ•°æ®åº“æ–‡ä»¶")

        # å¤‡ä»½å½“å‰æ•°æ®åº“
        from db_manager import db_manager
        current_db_path = db_manager.db_path

        # ç”Ÿæˆå¤‡ä»½æ–‡ä»¶è·¯å¾„ï¼ˆä¸åŸæ•°æ®åº“åœ¨åŒä¸€ç›®å½•ï¼‰
        db_dir = os.path.dirname(current_db_path)
        backup_filename = f"xianyu_data_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        backup_current_path = os.path.join(db_dir, backup_filename)

        if os.path.exists(current_db_path):
            shutil.copy2(current_db_path, backup_current_path)
            log_with_user('info', f"å½“å‰æ•°æ®åº“å·²å¤‡ä»½ä¸º: {backup_current_path}", admin_user)

        # å…³é—­å½“å‰æ•°æ®åº“è¿æ¥
        if hasattr(db_manager, 'conn') and db_manager.conn:
            db_manager.conn.close()
            log_with_user('info', "å·²å…³é—­å½“å‰æ•°æ®åº“è¿æ¥", admin_user)

        # æ›¿æ¢æ•°æ®åº“æ–‡ä»¶
        shutil.move(temp_file_path, current_db_path)
        log_with_user('info', f"æ•°æ®åº“æ–‡ä»¶å·²æ›¿æ¢: {current_db_path}", admin_user)

        # é‡æ–°åˆå§‹åŒ–æ•°æ®åº“è¿æ¥ï¼ˆä½¿ç”¨åŸæœ‰çš„db_pathï¼‰
        db_manager.__init__(db_manager.db_path)
        log_with_user('info', "æ•°æ®åº“è¿æ¥å·²é‡æ–°åˆå§‹åŒ–", admin_user)

        # éªŒè¯æ–°æ•°æ®åº“
        try:
            test_users = db_manager.get_all_users()
            log_with_user('info', f"æ•°æ®åº“æ¢å¤æˆåŠŸï¼ŒåŒ…å« {len(test_users)} ä¸ªç”¨æˆ·", admin_user)
        except Exception as e:
            log_with_user('error', f"æ•°æ®åº“æ¢å¤åéªŒè¯å¤±è´¥: {str(e)}", admin_user)
            # å¦‚æœéªŒè¯å¤±è´¥ï¼Œå°è¯•æ¢å¤åŸæ•°æ®åº“
            if os.path.exists(backup_current_path):
                shutil.copy2(backup_current_path, current_db_path)
                db_manager.__init__()
                log_with_user('info', "å·²æ¢å¤åŸæ•°æ®åº“", admin_user)
            raise HTTPException(status_code=500, detail="æ•°æ®åº“æ¢å¤å¤±è´¥ï¼Œå·²å›æ»šåˆ°åŸæ•°æ®åº“")

        return {
            "success": True,
            "message": "æ•°æ®åº“æ¢å¤æˆåŠŸ",
            "backup_file": backup_current_path,
            "user_count": len(test_users)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"ä¸Šä¼ æ•°æ®åº“å¤‡ä»½å¤±è´¥: {str(e)}", admin_user)
        # æ¸…ç†ä¸´æ—¶æ–‡ä»¶
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/backup/list')
def list_backup_files(admin_user: Dict[str, Any] = Depends(require_admin)):
    """åˆ—å‡ºæœåŠ¡å™¨ä¸Šçš„å¤‡ä»½æ–‡ä»¶ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', "æŸ¥è¯¢å¤‡ä»½æ–‡ä»¶åˆ—è¡¨", admin_user)

        # æŸ¥æ‰¾å¤‡ä»½æ–‡ä»¶ï¼ˆåœ¨dataç›®å½•ä¸­ï¼‰
        backup_files = glob.glob("data/xianyu_data_backup_*.db")

        backup_list = []
        for file_path in backup_files:
            try:
                stat = os.stat(file_path)
                backup_list.append({
                    'filename': os.path.basename(file_path),
                    'size': stat.st_size,
                    'size_mb': round(stat.st_size / (1024 * 1024), 2),
                    'created_time': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'modified_time': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
            except Exception as e:
                log_with_user('warning', f"è¯»å–å¤‡ä»½æ–‡ä»¶ä¿¡æ¯å¤±è´¥: {file_path} - {str(e)}", admin_user)

        # æŒ‰ä¿®æ”¹æ—¶é—´å€’åºæ’åˆ—
        backup_list.sort(key=lambda x: x['modified_time'], reverse=True)

        log_with_user('info', f"æ‰¾åˆ° {len(backup_list)} ä¸ªå¤‡ä»½æ–‡ä»¶", admin_user)

        return {
            "backups": backup_list,
            "total": len(backup_list)
        }

    except Exception as e:
        log_with_user('error', f"æŸ¥è¯¢å¤‡ä»½æ–‡ä»¶åˆ—è¡¨å¤±è´¥: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- æ•°æ®ç®¡ç†æ¥å£ -------------------------

@app.get('/admin/data/{table_name}')
def get_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """è·å–æŒ‡å®šè¡¨çš„æ‰€æœ‰æ•°æ®ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"æŸ¥è¯¢è¡¨æ•°æ®: {table_name}", admin_user)

        # éªŒè¯è¡¨åå®‰å…¨æ€§
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', "item_replay"
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"å°è¯•è®¿é—®ä¸å…è®¸çš„è¡¨: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="ä¸å…è®¸è®¿é—®è¯¥è¡¨")

        # è·å–è¡¨æ•°æ®
        data, columns = db_manager.get_table_data(table_name)

        log_with_user('info', f"è¡¨ {table_name} æŸ¥è¯¢æˆåŠŸï¼Œå…± {len(data)} æ¡è®°å½•", admin_user)

        return {
            "success": True,
            "data": data,
            "columns": columns,
            "count": len(data)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"æŸ¥è¯¢è¡¨æ•°æ®å¤±è´¥: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/data/{table_name}/{record_id}')
def delete_table_record(table_name: str, record_id: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """åˆ é™¤æŒ‡å®šè¡¨çš„æŒ‡å®šè®°å½•ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"åˆ é™¤è¡¨è®°å½•: {table_name}.{record_id}", admin_user)

        # éªŒè¯è¡¨åå®‰å…¨æ€§
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders','item_replay'
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"å°è¯•åˆ é™¤ä¸å…è®¸çš„è¡¨è®°å½•: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="ä¸å…è®¸æ“ä½œè¯¥è¡¨")

        # ç‰¹æ®Šä¿æŠ¤ï¼šä¸èƒ½åˆ é™¤ç®¡ç†å‘˜ç”¨æˆ·
        if table_name == 'users' and record_id == str(admin_user['user_id']):
            log_with_user('warning', "å°è¯•åˆ é™¤ç®¡ç†å‘˜è‡ªå·±", admin_user)
            raise HTTPException(status_code=400, detail="ä¸èƒ½åˆ é™¤ç®¡ç†å‘˜è‡ªå·±")

        # åˆ é™¤è®°å½•
        success = db_manager.delete_table_record(table_name, record_id)

        if success:
            log_with_user('info', f"è¡¨è®°å½•åˆ é™¤æˆåŠŸ: {table_name}.{record_id}", admin_user)
            return {"success": True, "message": "åˆ é™¤æˆåŠŸ"}
        else:
            log_with_user('warning', f"è¡¨è®°å½•åˆ é™¤å¤±è´¥: {table_name}.{record_id}", admin_user)
            raise HTTPException(status_code=400, detail="åˆ é™¤å¤±è´¥ï¼Œè®°å½•å¯èƒ½ä¸å­˜åœ¨")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"åˆ é™¤è¡¨è®°å½•å¼‚å¸¸: {table_name}.{record_id} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/data/{table_name}')
def clear_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """æ¸…ç©ºæŒ‡å®šè¡¨çš„æ‰€æœ‰æ•°æ®ï¼ˆç®¡ç†å‘˜ä¸“ç”¨ï¼‰"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"æ¸…ç©ºè¡¨æ•°æ®: {table_name}", admin_user)

        # éªŒè¯è¡¨åå®‰å…¨æ€§
        allowed_tables = [
            'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', 'item_replay',
            'risk_control_logs'
        ]

        # ä¸å…è®¸æ¸…ç©ºç”¨æˆ·è¡¨
        if table_name == 'users':
            log_with_user('warning', "å°è¯•æ¸…ç©ºç”¨æˆ·è¡¨", admin_user)
            raise HTTPException(status_code=400, detail="ä¸å…è®¸æ¸…ç©ºç”¨æˆ·è¡¨")

        if table_name not in allowed_tables:
            log_with_user('warning', f"å°è¯•æ¸…ç©ºä¸å…è®¸çš„è¡¨: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="ä¸å…è®¸æ¸…ç©ºè¯¥è¡¨")

        # æ¸…ç©ºè¡¨æ•°æ®
        success = db_manager.clear_table_data(table_name)

        if success:
            log_with_user('info', f"è¡¨æ•°æ®æ¸…ç©ºæˆåŠŸ: {table_name}", admin_user)
            return {"success": True, "message": "æ¸…ç©ºæˆåŠŸ"}
        else:
            log_with_user('warning', f"è¡¨æ•°æ®æ¸…ç©ºå¤±è´¥: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="æ¸…ç©ºå¤±è´¥")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"æ¸…ç©ºè¡¨æ•°æ®å¼‚å¸¸: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# å•†å“å¤šè§„æ ¼ç®¡ç†API
@app.put("/items/{cookie_id}/{item_id}/multi-spec")
def update_item_multi_spec(cookie_id: str, item_id: str, spec_data: dict, _: None = Depends(require_auth)):
    """æ›´æ–°å•†å“çš„å¤šè§„æ ¼çŠ¶æ€"""
    try:
        from db_manager import db_manager

        is_multi_spec = spec_data.get('is_multi_spec', False)

        success = db_manager.update_item_multi_spec_status(cookie_id, item_id, is_multi_spec)

        if success:
            return {"message": f"å•†å“å¤šè§„æ ¼çŠ¶æ€å·²{'å¼€å¯' if is_multi_spec else 'å…³é—­'}"}
        else:
            raise HTTPException(status_code=404, detail="å•†å“ä¸å­˜åœ¨")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# å•†å“å¤šæ•°é‡å‘è´§ç®¡ç†API
@app.put("/items/{cookie_id}/{item_id}/multi-quantity-delivery")
def update_item_multi_quantity_delivery(cookie_id: str, item_id: str, delivery_data: dict, _: None = Depends(require_auth)):
    """æ›´æ–°å•†å“çš„å¤šæ•°é‡å‘è´§çŠ¶æ€"""
    try:
        from db_manager import db_manager

        multi_quantity_delivery = delivery_data.get('multi_quantity_delivery', False)

        success = db_manager.update_item_multi_quantity_delivery_status(cookie_id, item_id, multi_quantity_delivery)

        if success:
            return {"message": f"å•†å“å¤šæ•°é‡å‘è´§çŠ¶æ€å·²{'å¼€å¯' if multi_quantity_delivery else 'å…³é—­'}"}
        else:
            raise HTTPException(status_code=404, detail="å•†å“ä¸å­˜åœ¨")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





# ==================== è®¢å•ç®¡ç†æ¥å£ ====================

@app.get('/api/orders')
def get_user_orders(current_user: Dict[str, Any] = Depends(get_current_user)):
    """è·å–å½“å‰ç”¨æˆ·çš„è®¢å•ä¿¡æ¯"""
    try:
        from db_manager import db_manager

        user_id = current_user['user_id']
        log_with_user('info', "æŸ¥è¯¢ç”¨æˆ·è®¢å•ä¿¡æ¯", current_user)

        # è·å–ç”¨æˆ·çš„æ‰€æœ‰Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        # è·å–æ‰€æœ‰è®¢å•æ•°æ®
        all_orders = []
        for cookie_id in user_cookies.keys():
            orders = db_manager.get_orders_by_cookie(cookie_id, limit=1000)  # å¢åŠ é™åˆ¶æ•°é‡
            # ä¸ºæ¯ä¸ªè®¢å•æ·»åŠ cookie_idä¿¡æ¯
            for order in orders:
                order['cookie_id'] = cookie_id
                all_orders.append(order)

        # æŒ‰åˆ›å»ºæ—¶é—´å€’åºæ’åˆ—
        all_orders.sort(key=lambda x: x.get('created_at', ''), reverse=True)

        log_with_user('info', f"ç”¨æˆ·è®¢å•æŸ¥è¯¢æˆåŠŸï¼Œå…± {len(all_orders)} æ¡è®°å½•", current_user)
        return {"success": True, "data": all_orders}

    except Exception as e:
        log_with_user('error', f"æŸ¥è¯¢ç”¨æˆ·è®¢å•å¤±è´¥: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"æŸ¥è¯¢è®¢å•å¤±è´¥: {str(e)}")


# ç§»é™¤è‡ªåŠ¨å¯åŠ¨ï¼Œç”±Start.pyæˆ–æ‰‹åŠ¨å¯åŠ¨
# if __name__ == "__main__":
#     uvicorn.run(app, host="0.0.0.0", port=8080)